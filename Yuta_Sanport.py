
# ==============================
# IMPORTS
# ==============================
import sys
import re
import ssl
import certifi
import urllib.request
import shutil
import tempfile
import pdfplumber
import os
import msvcrt

from pathlib import Path
from itertools import cycle
from datetime import datetime, date, timedelta, timezone

import tkinter as tk
from tkinter import Tk, filedialog

import pandas as pd
import xlwings as xw
import openpyxl
from copy import copy  # para copiar estilos
from openpyxl.styles import Font
from tempfile import gettempdir
import shutil
import holidays
from docx import Document
from num2words import num2words
import comtypes.client
import unicodedata
import locale
import unicodedata

from docx import Document
from shutil import copy2
from num2words import num2words
from datetime import datetime
import calendar


from pdf2image import convert_from_path
import pytesseract



# ==============================
# FERIADOS
# ==============================
feriados_br = holidays.Brazil()

feriados_personalizados = [
    date(2025, 1, 1),
    date(2025, 4, 21),
    date(2025, 5, 1),
    # ... outros feriados locais
]

for d in feriados_personalizados:
    feriados_br[d] = "Feriado personalizado"


# ==============================
# FUNÇÕES AUXILIARES GLOBAIS
# ==============================  

# ---------------------------
# 1️⃣ Copiar arquivo para pasta temporária e ler Excel
# ---------------------------
def copiar_para_temp_xlwings(caminho_original: Path) -> Path:
    if not caminho_original.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {caminho_original}")

    temp_dir = Path(tempfile.mkdtemp(prefix="faturamento_"))
    caminho_temp = temp_dir / caminho_original.name

    print(f"📄 Copiando para local temporário:")
    print(f"   {caminho_original.name}")
    shutil.copy2(caminho_original, caminho_temp)

    return caminho_temp



def copiar_para_temp_word(caminho_original: Path) -> Path:
    if not caminho_original.exists():
        raise FileNotFoundError(f"Arquivo Word não encontrado: {caminho_original}")

    temp_dir = Path(tempfile.mkdtemp(prefix="recibo_"))
    caminho_temp = temp_dir / caminho_original.name

    print(f"📄 Copiando modelo Word para temporário:")
    print(f"   {caminho_original.name}")
    shutil.copy2(caminho_original, caminho_temp)

    return caminho_temp



# ---------------------------
# 2️⃣ Localizar pasta FATURAMENTOS automaticamente
# ---------------------------
def obter_pasta_faturamentos() -> Path:
    print("\n=== BUSCANDO PASTA FATURAMENTOS AUTOMATICAMENTE ===")

    bases = [
        Path.home() / "SANPORT LOGÍSTICA PORTUÁRIA LTDA",
        Path.home() / "OneDrive - SANPORT LOGÍSTICA PORTUÁRIA LTDA",
    ]

    for base in bases:
        if base.exists():
            candidatos = list(base.rglob("FATURAMENTOS"))
            for c in candidatos:
                if "01. FATURAMENTOS" in c.parent.as_posix():
                    print(f"✅ Pasta FATURAMENTOS encontrada em:\n   {c}")
                    return c

    raise FileNotFoundError("Pasta FATURAMENTOS não localizada")



# ---------------------------
# 3️⃣ Abrir workbooks NAVIO e cliente com xlwings

# ---------------------------


#================DE ACORDO====================#

def abrir_workbooks_de_acordo(pasta_faturamentos: Path, pasta_navio: Path):
    pasta_cliente = pasta_navio.parent
    nome_cliente = pasta_cliente.name.strip()

    caminho_cliente_rede = pasta_faturamentos / f"{nome_cliente}.xlsx"
    if not caminho_cliente_rede.exists():
        raise FileNotFoundError(f"Faturamento não encontrado: {caminho_cliente_rede}")

    caminho_cliente_local = copiar_para_temp_xlwings(caminho_cliente_rede)

    app = xw.App(visible=False, add_book=False)
    wb = None

    try:
        wb = app.books.open(str(caminho_cliente_local))

        nomes_abas = [s.name for s in wb.sheets]
        if nome_cliente in nomes_abas:
            ws_front = wb.sheets[nome_cliente]
        elif "FRONT VIGIA" in nomes_abas:
            ws_front = wb.sheets["FRONT VIGIA"]
        else:
            raise RuntimeError("Aba FRONT não encontrada")

        return app, wb, ws_front

    except Exception:
        if wb:
            wb.close()
        app.quit()
        raise



def montar_nome_faturamento(dn: str, nome_navio: str) -> str:
    """
    Ex: dn=1, nome_navio='SANPORT'
    -> 'FATURAMENTO - ND 001 - MV SANPORT'
    """
    nd_formatado = str(dn).zfill(3)
    return f"FATURAMENTO - ND {nd_formatado} - MV {nome_navio}"


def escrever_de_acordo_nf(wb, nome_navio, dn, ano):
    """
    Escreve o texto DE ACORDO na aba NF (A1:E2).
    """

    ws_nf = None
    for sheet in wb.sheets:
        if sheet.name.strip().lower() == "nf":
            ws_nf = sheet
            break

    if ws_nf is None:
        print("⚠️ Aba NF não encontrada (DE ACORDO).")
        return

    texto = (
        f'SERVIÇO DE ATENDIMENTO/APOIO NO "DE ACORDO" '
        f'DA RAP DO {nome_navio} DN {dn}/{ano}'
    )


    rng = ws_nf.range("A1:E2")

    # segurança: desfaz merge anterior
    if rng.api.MergeCells:
        rng.api.UnMerge()

    rng.merge()
    rng.value = texto

    cel = ws_nf.range("A1")
    cel.api.HorizontalAlignment = -4108  # Center
    cel.api.VerticalAlignment = -4108
    cel.api.WrapText = True
    cel.api.Font.Bold = True
    cel.api.Font.Size = 14


def obter_nome_navio_da_pasta(pasta_navio: Path) -> str:
    """
    Ex: '054 - sanport' -> 'SANPORT'
    """
    nome = re.sub(r"^\s*\d+\s*[-–—]?\s*", "", pasta_navio.name).strip()
    return nome.upper() if nome else "NAVIO NÃO IDENTIFICADO"


#====================================================================================#



#===================SISTEMA=========================================#


def abrir_workbooks(pasta_faturamentos: Path):
    caminho_navio_rede = selecionar_arquivo_navio()
    if not caminho_navio_rede:
        raise FileNotFoundError("Arquivo do NAVIO não selecionado")

    caminho_navio_rede = Path(caminho_navio_rede)
    pasta_navio = caminho_navio_rede.parent
    pasta_cliente = pasta_navio.parent
    nome_cliente = pasta_cliente.name.strip()

    caminho_cliente_rede = pasta_faturamentos / f"{nome_cliente}.xlsx"
    if not caminho_cliente_rede.exists():
        raise FileNotFoundError(
            f"Arquivo de faturamento não encontrado:\n{caminho_cliente_rede}"
        )

    # 🔥 COPIA AMBOS PARA LOCAL
    caminho_navio_local = copiar_para_temp_xlwings(caminho_navio_rede)
    caminho_cliente_local = copiar_para_temp_xlwings(caminho_cliente_rede)

    app = xw.App(visible=False, add_book=False)
    wb1 = wb2 = None

    try:
        wb1 = app.books.open(str(caminho_navio_local))
        wb2 = app.books.open(str(caminho_cliente_local))

        ws1 = wb1.sheets[0]
        nomes_abas = [s.name for s in wb2.sheets]

        if nome_cliente in nomes_abas:
            ws_front = wb2.sheets[nome_cliente]
        elif "FRONT VIGIA" in nomes_abas:
            ws_front = wb2.sheets["FRONT VIGIA"]
        else:
            raise RuntimeError("Aba FRONT não encontrada")

        # ✅ RETURN PADRONIZADO (5 valores)
        return app, wb1, wb2, ws1, ws_front, pasta_navio


    except Exception:
        if wb1:
            wb1.close()
        if wb2:
            wb2.close()
        app.quit()
        raise


def selecionar_pasta_navio() -> Path:
    root = Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    pasta = filedialog.askdirectory(title="Selecione a pasta do NAVIO")

    root.destroy()

    if not pasta:
        raise RuntimeError("Nenhuma pasta de navio selecionada")

    pasta = Path(pasta)
    print(f"📁 Pasta do navio selecionada: {pasta.name}")
    return pasta


def obter_nome_navio(pasta_navio: Path, caminho_navio: Path | None = None) -> str:
    """
    Prioridade:
    1) Nome no arquivo
    2) Nome da pasta
    """
    if caminho_navio:
        nome_arquivo = obter_nome_navio_de_arquivo(caminho_navio)
        if nome_arquivo:
            return nome_arquivo

    return obter_nome_navio_da_pasta(pasta_navio)




def escrever_nf_faturamento_completo(wb_faturamento, nome_navio, dn, celula="A1", area_merge="A1:E10"):
    ws_nf = None
    for sheet in wb_faturamento.sheets:
        if sheet.name.strip().lower() == "nf":
            ws_nf = sheet
            break

    if ws_nf is None:
        print("⚠️ Aba NF não encontrada.")
        return False

    ano = datetime.now().strftime("%y")

    texto = f"SERVIÇO PRESTADO DE ATENDIMENTO/APOIO AO M/V {nome_navio}\nDN {dn}/{ano}"

    rng = ws_nf.range(area_merge)

    # ✅ desfaz merges com segurança (mesmo se a área tiver merge parcial)
    try:
        rng.api.UnMerge()
    except Exception:
        pass

    rng.merge()
    rng.value = texto

    cel = ws_nf.range(celula)
    cel.api.HorizontalAlignment = -4108  # xlCenter
    cel.api.VerticalAlignment = -4108    # xlCenter
    cel.api.WrapText = True
    cel.api.Font.Bold = True
    cel.api.Font.Size = 12

    print("✅ NF preenchida (A1:E10)")
    return True




def obter_dn_da_pasta(pasta_navio: Path) -> str:
    """
    Extrai o DN do início do nome da pasta.
    Ex: '054 - SANPORT' -> '054'
    """
    match = re.match(r"^\s*(\d+)", pasta_navio.name)
    if not match:
        print(
            f"⚠️ DN não encontrado no início da pasta "
            f"'{pasta_navio.name}', usando '0000'"
        )
        return "0000"

    return match.group(1)


def obter_nome_navio_de_arquivo(caminho_navio: Path) -> str:
    """
    Ex: 'FATURAMENTO - ND 001 - MV HOS REMINGTON.xlsx'
    -> 'MV HOS REMINGTON'
    """
    nome = re.sub(
        r"^.*?(?:DN|ND)\s*\d+\s*[-–—]?\s*",
        "",
        caminho_navio.stem,
        flags=re.IGNORECASE
    ).strip()

    return nome.upper() if nome else "NAVIO NÃO IDENTIFICADO"



def fechar_workbooks(app=None, wb_navio=None, wb_cliente=None, arquivo_saida: Path | None = None):
    try:
        if wb_navio and arquivo_saida:
            if arquivo_saida.exists():
                arquivo_saida.unlink()
            wb_navio.save(str(arquivo_saida))
            print(f"💾 Arquivo Excel salvo em: {arquivo_saida}")
    except Exception as e:
        print(f"⚠️ Erro ao salvar wb_navio: {e}")

    try:
        if wb_navio:
            wb_navio.close()
    except Exception as e:
        print(f"⚠️ Erro ao fechar wb_navio: {e}")

    try:
        if wb_cliente:
            wb_cliente.close()
    except Exception as e:
        print(f"⚠️ Erro ao fechar wb_cliente: {e}")

    try:
        if app:
            app.quit()
    except Exception as e:
        print(f"⚠️ Erro ao fechar Excel: {e}")


def selecionar_arquivo_navio() -> str | None:
    root = Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    root.update_idletasks()

    caminho = filedialog.askopenfilename(
        title="Selecione o arquivo do NAVIO",
        filetypes=[("Arquivos Excel", "*.xlsx")]
    )

    root.destroy()

    if not caminho:
        return None

    print(f"📂 Arquivo NAVIO selecionado: {Path(caminho).name}")
    return caminho

def salvar_excel_com_nome(wb, pasta_saida: Path, nome_base: str) -> Path:
    """
    Salva SEM usar SaveAs (evita erro Excel COM).
    """
    caminho_final = pasta_saida / f"{nome_base}.xlsx"

    # 🧠 Se existir, apaga
    if caminho_final.exists():
        caminho_final.unlink()

    # 🔥 ESSENCIAL: SaveCopyAs (não SaveAs)
    wb.api.SaveCopyAs(str(caminho_final))

    return caminho_final

def obter_modelo_word_cargonave(pasta_faturamentos: Path, cliente: str) -> Path:
    caminhos_teste = [
        pasta_faturamentos / cliente,
        pasta_faturamentos / "CARGONAVE",  # fallback
    ]

    for caminho in caminhos_teste:
            arquivos = list(caminho.glob("RECIBO - YUTA.docx"))
            if arquivos:
                return arquivos[0]

    raise FileNotFoundError(f"Modelo Word não encontrado em {caminhos_teste}")


def gerar_pdf(caminho_excel, pasta_saida, nome_base, ws=None):
    app = xw.App(visible=False, add_book=False)
    wb = app.books.open(str(caminho_excel))

    try:
        caminho_pdf = pasta_saida / f"{nome_base}.pdf"

        if ws is not None:
            ws.api.ExportAsFixedFormat(Type=0, Filename=str(caminho_pdf))
        else:
            wb.api.ExportAsFixedFormat(Type=0, Filename=str(caminho_pdf))

        print(f"📄 PDF gerado: {caminho_pdf}")
        return caminho_pdf

    finally:
        wb.close()
        app.quit()





def gerar_pdf_workbook_inteiro(wb, pasta_saida: Path, nome_base: str) -> Path:
    caminho_pdf = pasta_saida / f"{nome_base}.pdf"

    if caminho_pdf.exists():
        caminho_pdf.unlink()

    wb.api.ExportAsFixedFormat(
        Type=0,  # PDF
        Filename=str(caminho_pdf),
        Quality=0,
        IncludeDocProperties=True,
        IgnorePrintAreas=False,  # respeita área de impressão de cada aba
        OpenAfterPublish=False
    )

    return caminho_pdf


def gerar_pdf_faturamento_completo(wb, pasta_saida: Path, nome_base: str) -> Path:
    caminho_pdf = pasta_saida / f"{nome_base}.pdf"

    if caminho_pdf.exists():
        caminho_pdf.unlink()

    # 🔒 Oculta aba NF (se existir)
    aba_nf = None
    for ws in wb.sheets:
        if ws.name.strip().upper() == "NF":
            aba_nf = ws
            ws.api.Visible = False
            break

    # 🔥 Remove qualquer Print_Area escondido
    try:
        for nome in list(wb.api.Names):
            if nome.Name.lower() == "print_area":
                nome.Delete()
    except:
        pass

    # 📄 Exporta workbook inteiro
    wb.api.ExportAsFixedFormat(
        Type=0,  # PDF
        Filename=str(caminho_pdf),
        Quality=0,
        IncludeDocProperties=True,
        IgnorePrintAreas=True,
        OpenAfterPublish=False
    )

    # 🔓 Reexibe NF
    if aba_nf:
        aba_nf.api.Visible = True

    return caminho_pdf



def extrair_identidade_navio(pasta_navio: Path) -> tuple[str, str]:
    """
    Retorna (dn, nome_navio) a partir da pasta do navio
    Ex: '123 - UNIMAR' -> ('123', 'UNIMAR')
    """
    dn = obter_dn_da_pasta(pasta_navio)
    nome_navio = obter_nome_navio_da_pasta(pasta_navio)
    return dn, nome_navio


#===================FATURAMENTO SÃO SEBASTIÃO=========================================#


def gerar_pdf_do_wb_aberto(wb, pasta_saida, nome_base, ignorar_abas=("nf",)):
    caminho_pdf = Path(pasta_saida) / f"{nome_base}.pdf"

    # 1) se existir e estiver aberto, já avisa o motivo
    if caminho_pdf.exists():
        try:
            caminho_pdf.unlink()
        except Exception as e:
            raise RuntimeError(f"PDF está aberto/travado e não pode ser sobrescrito: {caminho_pdf}") from e

    app = wb.app
    app.api.DisplayAlerts = False

    # 2) guarda visibilidade, oculta as que não devem sair no PDF
    vis_orig = {}
    for sh in wb.sheets:
        nome_norm = sh.name.strip().lower()
        vis_orig[sh.name] = sh.api.Visible
        if nome_norm in {x.strip().lower() for x in ignorar_abas}:
            sh.api.Visible = False  # oculta NF

    try:
        # 3) ativa uma aba visível (Excel odeia export sem sheet ativa)
        aba_ativa = None
        for sh in wb.sheets:
            if sh.api.Visible:  # True / -1
                aba_ativa = sh
                break
        if aba_ativa:
            aba_ativa.activate()

        # 4) exporta o workbook (sem as abas ocultas)
        wb.api.ExportAsFixedFormat(
            Type=0,  # xlTypePDF
            Filename=str(caminho_pdf),
            Quality=0,  # xlQualityStandard
            IncludeDocProperties=True,
            IgnorePrintAreas=False,
            OpenAfterPublish=False
        )

        print(f"📄 PDF gerado: {caminho_pdf}")
        return caminho_pdf

    finally:
        # 5) restaura visibilidade original
        for sh in wb.sheets:
            if sh.name in vis_orig:
                sh.api.Visible = vis_orig[sh.name]






# ==============================
# CLASSE 1: FATURAMENTO COMPLETO

class FaturamentoCompleto:
    def __init__(self, g_logic=1):
        self.app = None
        self.wb1 = None
        self.wb2 = None
        self.ws1 = None
        self.ws_front = None
        self.nome_navio = None
        self.g_logic = g_logic
        self.pasta_saida_final = None
        self.dn = None
        self.pdf_path = None
        self.pasta_faturamentos = None  # <--- GUARDA PASTA AQUI



    def executar(self):
        print("🚀 Iniciando execução...")

        # 🔹 1️⃣ Buscar pasta FATURAMENTOS apenas 1x
        self.pasta_faturamentos = obter_pasta_faturamentos()
        resultado = abrir_workbooks(self.pasta_faturamentos)

        if not resultado:
            raise SystemExit("❌ Erro ou pasta inválida")

        (
            self.app,
            self.wb1,
            self.wb2,
            self.ws1,
            self.ws_front,
            pasta_navio_rede,
        ) = resultado

        self.pasta_saida_final = pasta_navio_rede

        # 🔹 Extrair DN e nome do navio
        self.dn, self.nome_navio = extrair_identidade_navio(pasta_navio_rede)

        # caminho PDF OGMO
        self.pdf_path = pasta_navio_rede / "FOLHAS OGMO.pdf"

        print(f"📌 DN: {self.dn}")
        print(f"🚢 NAVIO: {self.nome_navio}")
        print(f"📑 PDF OGMO: {self.pdf_path}")

        escrever_nf_faturamento_completo(self.wb2, self.nome_navio, self.dn)

        nome_base = f"FATURAMENTO - DN {self.dn} - MV {self.nome_navio}"

        try:
            self.processar()

            caminho_excel = pasta_navio_rede / f"{nome_base}.xlsx"
            caminho_pdf = pasta_navio_rede / f"{nome_base}.pdf"

            if caminho_excel.exists():
                caminho_excel.unlink()
            gerar_pdf_faturamento_completo(
                self.wb2,
                pasta_navio_rede,
                nome_base
            )

            # SALVAR EXCEL (local → rede)
            temp_excel = Path(gettempdir()) / f"{nome_base}.xlsx"
            if temp_excel.exists():
                temp_excel.unlink()

            self.wb2.save(str(temp_excel))
            shutil.copy2(temp_excel, caminho_excel)

            # PDF REPORT separado
            self.gerar_pdf_report_vigia_separado(
                pasta_navio_rede, self.dn, self.nome_navio
            )

            fechar_workbooks(self.app, self.wb1, self.wb2)

            print(f"💾 Excel salvo em: {caminho_excel}")
            print(f"📑 PDF FRONT salvo em: {caminho_pdf}")

        except Exception as e:
            print(f"❌ ERRO NO FATURAMENTO: {e}")
            fechar_workbooks(self.app, self.wb1, self.wb2)
            raise


    def processar(self):
        # ---------- FRONT ----------
        self.preencher_front_vigia()

        # ---------- REPORT ----------
        if "REPORT VIGIA" not in [s.name for s in self.wb2.sheets]:
            raise RuntimeError("Aba 'REPORT VIGIA' não encontrada")

        ws_report = self.wb2.sheets["REPORT VIGIA"]

        self.processar_MMO(self.wb1, self.wb2)

        qtd_periodos = self.obter_periodos(self.ws1)

        self.inserir_linhas_report(
            ws_report,
            linha_inicial=22,
            periodos=qtd_periodos
        )

        periodos = self.preencher_coluna_E(
            ws_report,
            linha_inicial=22,
            debug=True
        )

        self.preencher_coluna_G(
            ws_report,
            self.ws1,
            linha_inicial=22,
            periodos=periodos,
            debug=True
        )

        self.montar_datas_report_vigia(
            ws_report,
            self.ws1,
            linha_inicial=22,
            periodos=len(periodos)
        )


        valor_arredondado = self.arredondar_para_baixo_50_se_cargonave()

        # 🔹 GERAR RECIBO CARGONAVE (Word + PDF)
        self.gerar_recibo_cargonave_word()


        # 🔹 GERAR PLANILHA DE CÁLCULO
        self.gerar_planilha_calculo_cargonave()

        self.gerar_planilha_calculo_conesul()

        print("✅ REPORT VIGIA atualizado com sucesso!")


    def escrever_cn_credit_note(self, texto_cn):
            ws_credit = None

            for sheet in self.wb2.sheets:
                if sheet.name.strip().lower() == "credit note":
                    ws_credit = sheet
                    break

            if ws_credit is None:
                print("ℹ️ Aba Credit Note não existe — seguindo fluxo.")
                return

            ws_credit.range("C21").value = texto_cn


    # ===== FRONT ======#
    def extrair_berco(self):
        """Extrai o valor do campo 'Berço' do PDF FOLHAS OGMO."""
        if not self.pdf_path or not Path(self.pdf_path).exists():
            print("⚠️ PDF FOLHAS OGMO não encontrado")
            return None

        with pdfplumber.open(str(self.pdf_path)) as pdf:
            for page in pdf.pages:
                words = page.extract_words()
                for w in words:
                    if w["text"] == "Berço":
                        x_ref = w["x0"]
                        y_ref = w["top"]

                        # pega palavras logo abaixo, alinhadas na mesma coluna
                        candidatos = [
                            wd for wd in words
                            if abs(wd["x0"] - x_ref) < 50 and wd["top"] > y_ref
                        ]

                        if candidatos:
                            candidatos.sort(key=lambda wd: wd["top"])
                            y_target = candidatos[0]["top"]

                            # junta todas as palavras dessa mesma linha
                            linha = [
                                wd["text"] for wd in candidatos
                                if abs(wd["top"] - y_target) < 5
                            ]
                            return " ".join(linha).strip()
        return None

    def preencher_front_vigia(self):
        try:
            ano_curto = datetime.now().strftime('%y')

            # FRONT VIGIA
            texto_dn = f"DN {self.dn}/{ano_curto}"
            self.ws_front.range("D15").value = self.nome_navio
            self.ws_front.range("C21").value = texto_dn

            # CREDIT NOTE (se existir)
            texto_cn = f"CN {self.dn}/{ano_curto}"
            self.escrever_cn_credit_note(texto_cn)


            # ======================

            # automatiza a leitura do BERÇO
            berco = self.extrair_berco()
            if berco:
                self.ws_front.range("D18").value = berco.upper()
            else:
                self.ws_front.range("D18").value = "NÃO ENCONTRADO"

            # -------- DATAS --------
            data_min, data_max = self.obter_datas_extremos(self.ws1)
            if data_min:
                self.ws_front.range("D16").value = self.data_por_extenso(data_min)
            if data_max:
                self.ws_front.range("D17").value = self.data_por_extenso(data_max)

            # -------- RODAPÉ --------
            hoje = datetime.now()
            meses = [
                "", "janeiro","fevereiro","março","abril","maio","junho",
                "julho","agosto","setembro","outubro","novembro","dezembro"
            ]
            self.ws_front.range("C39").value = (
                f"  Santos, {hoje.day} de {meses[hoje.month]} de {hoje.year}"
            )

            print("✅ FRONT VIGIA preenchido com sucesso!")

        except Exception as e:
            print(f"❌ Erro ao preencher FRONT VIGIA: {e}")
            raise


#==================== REPORT =====================#

    def inserir_linhas_report(self, ws_report, linha_inicial, periodos):
        if periodos <= 1:
            return
        app = ws_report.book.app
        app.screen_updating = False
        app.display_alerts = False
        app.enable_events = False
        app.calculation = 'manual'
        try:
            linha_modelo = linha_inicial
            for i in range(periodos - 1):
                destino = linha_inicial + 1 + i
                ws_report.api.Rows(destino).Insert()
                ws_report.api.Rows(linha_modelo).Copy(ws_report.api.Rows(destino))
        finally:
            app.calculation = 'automatic'
            app.enable_events = True
            app.display_alerts = True
            app.screen_updating = True


    # ===== LINHA E=====#


    def gerar_ciclos_coluna_E(self, ws_resumo, linha_inicial=2):
        """
        Gera a lista de períodos para a coluna E do REPORT, baseada na data mais antiga.
        Sequência: 06x12 -> 12x18 -> 18x24 -> 00x06
        """
        sequencia_padrao = ["06x12", "12x18", "18x24", "00x06"]

        # 1️⃣ Encontrar a primeira data válida (mais antiga que hoje)
        last_row = ws_resumo.used_range.last_cell.row
        valores = ws_resumo.range(f"B{linha_inicial}:B{last_row}").value
        hoje = date.today()
        primeira_linha_data = None

        for idx, v in enumerate(valores):
            if v in (None, "", "Total"):
                continue
            if isinstance(v, datetime):
                d = v.date()
            elif isinstance(v, str):
                try:
                    d = datetime.strptime(v.strip(), "%d/%m/%Y").date()
                except:
                    continue
            else:
                continue

            if d < hoje:
                primeira_linha_data = linha_inicial + idx
                break

        if primeira_linha_data is None:
            return []  # nenhuma data antiga encontrada

        # 2️⃣ Contar espaços vazios ou "Total" antes da próxima data
        contador_vazio = 0
        for i in range(primeira_linha_data + 1, last_row + 1):
            valor = ws_resumo.range(f"B{i}").value
            if valor in (None, "", "Total"):
                contador_vazio += 1
            else:
                break

        # 3️⃣ Definir primeiro período
        if contador_vazio >= 4:
            primeiro_periodo = "06x12"
        elif contador_vazio == 3:
            primeiro_periodo = "12x18"
        elif contador_vazio == 2:
            primeiro_periodo = "18x24"
        else:
            primeiro_periodo = "00x06"

        # 4️⃣ Sequência cíclica
        idx_inicio = sequencia_padrao.index(primeiro_periodo)
        sequencia_ciclica = sequencia_padrao[idx_inicio:] + sequencia_padrao[:idx_inicio]

        # 5️⃣ Gerar lista completa de períodos
        total_periodos = self.obter_periodos(ws_resumo)
        ciclos = [sequencia_ciclica[i % 4] for i in range(total_periodos)]

        return ciclos


    
    def preencher_coluna_E(self, ws_report, linha_inicial=22, debug=False):
        """
        Preenche a coluna E do REPORT VIGIA com os períodos gerados.
        """
        try:
            ciclos = self.gerar_ciclos_coluna_E(self.ws1)
            for idx, p in enumerate(ciclos):
                ws_report.range(f"E{linha_inicial + idx}").value = p
            if debug:

                return ciclos
        except Exception as e:
            print(f"❌ Erro ao preencher coluna E: {e}")
            raise

 

    # ===== LINHA G=====#


    def normalizar_periodo(self, valor_c):
        if not valor_c:
            return None

        s = str(valor_c).strip().lower()
        if s.startswith("06"):
            return "06x12"
        if s.startswith("12"):
            return "12x18"
        if s.startswith("18"):
            return "18x24"
        if s.startswith("00"):
            return "00x06"
        return None
        

    def gerar_valores_coluna_G(self, ws_resumo, periodos_E, debug=False):
        mapa = self.extrair_valores_por_periodo(ws_resumo, debug=debug)
        contadores = {k: 0 for k in mapa}
        valores_g = []

        for p in periodos_E:
            if p in mapa and contadores[p] < len(mapa[p]):
                valor = mapa[p][contadores[p]]
                contadores[p] += 1
            else:
                valor = 0.0

            valores_g.append(valor)

        if debug:


         return valores_g


    def preencher_coluna_G(self, ws_report, ws_resumo, linha_inicial=22, periodos=None, debug=False):
        """
        Preenche a coluna G seguindo EXATAMENTE a ordem da coluna E
        e formatando como moeda com 2 casas decimais.
        """

        if not periodos:
            raise ValueError("periodos (lista da coluna E) é obrigatório")

        valores = self.gerar_valores_coluna_G(
            ws_resumo,
            periodos,
            debug=debug
        )


        for i, valor in enumerate(valores):
                cell = ws_report.range(f"G{linha_inicial + i}")
                cell.value = valor              # número cru, sem arredondar
                cell.api.NumberFormatLocal = 'R$ #.##0,00'


    def extrair_valores_por_periodo(self, ws_resumo, debug=False):
        last_row = ws_resumo.used_range.last_cell.row

        mapa = {
            "00x06": [],
            "06x12": [],
            "12x18": [],
            "18x24": []
        }

        for i in range(2, last_row + 1):
            c = ws_resumo.range(f"C{i}").value
            z = ws_resumo.range(f"Z{i}").value

            if not c or z is None:
                continue

            s_c = str(c).strip().lower()
            if s_c.startswith("total"):
                continue

            periodo = self.normalizar_periodo(s_c)
            if not periodo:
                continue

            try:
                # ✅ CASO 1: já é número no Excel
                if isinstance(z, (int, float)):
                    valor = float(z)

                # ✅ CASO 2: veio como texto "R$ 1.144,70"
                else:
                    valor = (
                        str(z)
                        .replace("R$", "")
                        .replace(".", "")
                        .replace(",", ".")
                        .strip()
                    )
                    valor = float(valor)

            except:
                continue


            mapa[periodo].append(valor)
            

        return mapa
    

    def extrair_numero_excel(self, z):
        """
        Garante conversão correta de valores do Excel
        independente de vir como float ou string pt-BR.
        """

        # 👉 Caso 1: Excel já entregou número
        if isinstance(z, (int, float)):
            return float(z)

        # 👉 Caso 2: Veio como texto (ex: "1.144,70")
        s = str(z).strip()

        if not s:
            raise ValueError("valor vazio")

        s = (
            s.replace("R$", "")
            .replace(" ", "")
            .replace(".", "")
            .replace(",", ".")
        )

        return float(s)


    # ===== LINHA C =====#


    def montar_datas_report_vigia(self, ws_report, ws_resumo, linha_inicial=22, periodos=None):
        if periodos is None:
            raise ValueError("É necessário informar 'periodos' para preencher as datas")
        data_inicio, _ = self.obter_datas_extremos(ws_resumo)
        if not data_inicio:
            raise ValueError("Não foi possível determinar a data inicial na aba RESUMO")
        data_atual = data_inicio
        for i in range(periodos):
            linha = linha_inicial + i
            ciclo = ws_report.range(f"E{linha}").value
            if ciclo in (None, ""):
                break
            ws_report.range(f"C{linha}").value = data_atual
            if isinstance(ciclo, str) and ciclo.strip().lower() == "00x06":
                data_atual += timedelta(days=1)
        return periodos


    # ===== ABAS ESPECIFICAS =====#






    def gerar_pdf_report_vigia_separado(self, pasta_navio: Path, dn: str, navio: str):
        ws_report = self.wb2.sheets["REPORT VIGIA"]

        nome_pdf = f"REPORT VIGIA - DN {dn} - MV {navio}.pdf"
        caminho_pdf = pasta_navio / nome_pdf

        ws_report.api.ExportAsFixedFormat(
            Type=0,
            Filename=str(caminho_pdf),
            Quality=0,
            IncludeDocProperties=True,
            IgnorePrintAreas=False,
            OpenAfterPublish=False
        )

        print(f"📑 PDF REPORT VIGIA salvo em: {caminho_pdf}")



    def processar_MMO(self, wb_navio, wb_cliente):
        """
        MMO:
        - LÊ: último valor da coluna G da aba 'Resumo' (NAVIO)
        - ESCREVE: F25 da aba 'REPORT VIGIA' (CLIENTE)
        """

        print("   🔹 Iniciando MMO...")

        # ---------- REPORT VIGIA (CLIENTE) ----------
        try:
            ws_report = wb_cliente.sheets["REPORT VIGIA"]
        except:
            print("   ⚠️ Aba 'REPORT VIGIA' não encontrada no CLIENTE. MMO ignorado.")
            return

        if str(ws_report.range("E25").value).strip().upper() != "MMO":
            print("   ℹ️ MMO não aplicável (E25 != 'MMO').")
            return

        # ---------- RESUMO (NAVIO) ----------
        try:
            ws_resumo = wb_navio.sheets["Resumo"]
        except:
            print("   ⚠️ Aba 'Resumo' não encontrada no NAVIO. MMO ignorado.")
            return

        # ---------- LÊ COLUNA G ----------
        valores = ws_resumo.range("G1:G1000").value
        valores_validos = [v for v in valores if v not in (None, "")]

        if not valores_validos:
            print("   ℹ️ Coluna G vazia no Resumo. MMO ignorado.")
            return

        ultimo_valor = valores_validos[-1]

        # ---------- CONVERSÃO CORRETA (IGUAL COLUNA G) ----------
        try:
            valor_float = float(ultimo_valor)
        except:
            print(f"   ⚠️ Valor inválido '{ultimo_valor}'. Usando 0.")
            valor_float = 0.0

        # 🔥 correção de escala (quando vem gigante)
        if valor_float > 1_000_000:
            valor_float = valor_float / 100

        # ---------- ESCREVE ----------
        celula = ws_report.range("F25")
        celula.value = valor_float
        celula.api.NumberFormatLocal = 'R$ #.##0,00'

        print(f"   ✅ MMO concluído → R$ {valor_float:,.2f}")



    def arredondar_para_baixo_50_se_cargonave(self):
        """
        Arredonda para baixo em múltiplos de 50.
        Somente para A/C AGÊNCIA MARÍTIMA CARGONAVE LTDA.
        Coloca o resultado em H28 do FRONT.
        """
        ws_front_vigia = self.ws_front
        valor_empresa = ws_front_vigia.range("C9").value
        if not valor_empresa:
            return

        if str(valor_empresa).strip().upper() != "A/C AGÊNCIA MARÍTIMA CARGONAVE LTDA.":
            return

        valor = ws_front_vigia.range("E37").value
        if valor is None:
            return

        try:
            resultado = (int(valor) // 50) * 50
        except (ValueError, TypeError):
            return

        ws_front_vigia.range("H28").value = resultado

        # Para gerar o Word, você pode usar esse mesmo valor:
        return resultado
    

    def gerar_recibo_cargonave_word(self):

        word = None
        doc = None

        try:
            # ==========================
            # 🔒 TRAVA DE SEGURANÇA
            # ==========================
            ws = self.ws_front

            empresa = ws.range("C9").value
            if not empresa or str(empresa).strip().upper() != "A/C AGÊNCIA MARÍTIMA CARGONAVE LTDA.":
                print("ℹ️ Recibo não gerado (empresa não é CARGONAVE).")
                return

            valor_h28 = ws.range("H28").value
            if valor_h28 in (None, "", 0):
                print("ℹ️ Recibo não gerado (adiantamento não acionado ou valor zero).")
                return

            # ==========================
            # 📄 MODELO WORD
            # ==========================
            pasta_modelos = self.pasta_faturamentos.parent / "CARGONAVE"
            modelos = list(pasta_modelos.glob("RECIBO - YUTA.doc"))

            if not modelos:
                print(f"❌ Modelo Word não encontrado em {pasta_modelos}")
                return

            modelo_word = modelos[0]

            # ==========================
            # 📂 COPIAR PARA TEMP
            # ==========================
            temp_doc = Path(tempfile.gettempdir()) / f"RECIBO - {self.dn}.doc"
            shutil.copy2(modelo_word, temp_doc)

            # ==========================
            # 📝 ABRIR WORD
            # ==========================
            word = comtypes.client.CreateObject("Word.Application")
            word.Visible = False
            doc = word.Documents.Open(str(temp_doc))

            # ==========================
            # 💰 VALOR
            # ==========================
            valor = float(valor_h28)

            hoje = datetime.now()
            meses = [
                "", "janeiro", "fevereiro", "março", "abril", "maio", "junho",
                "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"
            ]

            data_extenso = f"Santos, {hoje.day} de {meses[hoje.month].capitalize()} de {hoje.year}"

            placeholders = {
                "{{VALOR}}": f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
                "{{VALOR_EXTENSO}}": num2words(valor, lang="pt_BR") + " reais",
                "{{DN}}": self.dn,
                "{{NAVIO}}": self.nome_navio,
                "{{DATA}}": data_extenso,
            }

            # ==========================
            # 🔁 SUBSTITUIR (TUDO NEGRITO)
            # ==========================
            find = doc.Content.Find

            for key, val in placeholders.items():
                find.ClearFormatting()
                find.Replacement.ClearFormatting()

                find.Text = key
                find.Replacement.Text = str(val)

                # 👉 FORÇA NEGRITO SEM EXCEÇÃO
                find.Replacement.Font.Bold = True

                find.Forward = True
                find.Wrap = 1  # wdFindContinue
                find.MatchCase = False
                find.MatchWholeWord = False
                find.Execute(Replace=2)

                print(f"🔄 Substituído {key} → {val} (NEGRITO)")

            # ==========================
            # 💾 SALVAR WORD + PDF
            # ==========================
            word_saida = self.pasta_saida_final / f"RECIBO - DN {self.dn} - MV {self.nome_navio}.doc"
            doc.SaveAs(str(word_saida))
            print(f"💾 Word do recibo salvo em: {word_saida}")

            pdf_saida = word_saida.with_suffix(".pdf")
            doc.SaveAs(str(pdf_saida), FileFormat=17)
            print(f"📑 PDF do recibo salvo em: {pdf_saida}")

            doc.Close(False)
            word.Quit()

        except Exception as e:
            if doc:
                try:
                    doc.Close(False)
                except:
                    pass
            if word:
                try:
                    word.Quit()
                except:
                    pass
            print(f"❌ Erro ao gerar recibo CARGONAVE: {e}")

    def gerar_planilha_calculo_cargonave(self):
        try:
            # ==========================
            # 🔒 TRAVA DE SEGURANÇA
            # ==========================
            ws = self.ws_front
            empresa = ws.range("C9").value

            if not empresa or str(empresa).strip().upper() != "A/C AGÊNCIA MARÍTIMA CARGONAVE LTDA.":
                print("ℹ️ Planilha de cálculo não gerada (empresa não é CARGONAVE).")
                return


            # ==========================
            # 🔤 FUNÇÃO AUXILIAR
            # ==========================
            def remover_acentos(txt: str) -> str:
                return unicodedata.normalize("NFD", txt).encode("ascii", "ignore").decode("utf-8")

            # ==========================
            # 📂 PASTA DO MODELO (BASE)
            # ==========================
            # ✅ PASTA DO MODELO (dinâmica - funciona no cliente)
            pasta_modelo = self.encontrar_pasta_modelo("CARGONAVE")


            if not pasta_modelo.exists():
                raise FileNotFoundError(f"Pasta modelo não encontrada: {pasta_modelo}")

            # ==========================
            # 📂 PASTA DO NAVIO (DESTINO)
            # ==========================
            pasta_navio = self.pasta_saida_final
            pasta_navio.mkdir(parents=True, exist_ok=True)

            # ==========================
            # 🔎 LOCALIZAR MODELO EXCEL
            # ==========================
            modelo = None

            for arq in pasta_modelo.glob("*.xlsx"):
                nome_limpo = remover_acentos(arq.name.lower())
                if "calculo" in nome_limpo:
                    modelo = arq
                    break

            if not modelo:
                raise FileNotFoundError(
                    f"Nenhum modelo de cálculo encontrado em {pasta_modelo}"
                )
            

            # ==========================
            # 📄 COPIAR MODELO
            # ==========================
            destino = pasta_navio / "CALCULO - YUTA.xlsx"
            shutil.copy2(modelo, destino)

            # ==========================
            # 📊 ABRIR PLANILHA
            # ==========================
            wb = openpyxl.load_workbook(destino)
            ws = wb.active  # ou wb["Cálculo"] se quiser fixar

            # ==========================
            # 📥 PEGAR ÚLTIMA LINHA DO OGMO
            # ==========================
            ws1 = self.ws1  # ✔ CONFIRMADO no teu fluxo

            ultima_linha = self.ultima_linha_com_valor(ws1, "G")



            print(f"📊 Última linha detectada no NAVIO: {ultima_linha}")

            mapa = {
                "C5": "G",
                "D5": "H",
                "E5": "I",
                "F5": "N",
                "G5": "O",
                "H5": "P",
                "I5": "Q",
                "J5": "S",
                "K5": "T",
                "L5": "V",
                "M5": "Z",
            }

            for destino_cell, origem_col in mapa.items():
                valor = ws1[f"{origem_col}{ultima_linha}"].value
                ws[destino_cell] = valor
                print(f"   🔹 {origem_col}{ultima_linha} → {destino_cell} | Valor: {valor}")

            # ==========================
            # ➕ CAMPOS ADICIONAIS
            # ==========================

            # AA (última linha OGMO) → B3
            valor_aa = ws1[f"AA{ultima_linha}"].value
            ws["B3"] = valor_aa
            print(f"   🔹 AA{ultima_linha} → B3 | Valor: {valor_aa}")

            # ==========================
            # 🚢 NOME DO NAVIO
            # ==========================
            nome_navio = self.nome_navio  # ajuste se o atributo tiver outro nome

            ws["A4"] = nome_navio

            print(f"   🔹 NAVIO → A2 e A4 | Valor: {nome_navio}")





            # ==========================
            # 💾 SALVAR
            # ==========================
            wb.save(destino)

            print("✅ Planilha CÁLCULO CARGONAVE gerada com sucesso!")

        except Exception as e:
            print(f"❌ Erro ao gerar planilha CÁLCULO CARGONAVE: {e}")
            raise




    def ultima_linha_com_valor(self, ws, coluna):
        for linha in range(ws.used_range.last_cell.row, 0, -1):
            if ws[f"{coluna}{linha}"].value not in (None, ""):
                return linha
        return None



    from pathlib import Path

    def encontrar_pasta_modelo(self, nome_cliente: str) -> Path:
        """
        Encontra ...\01. FATURAMENTOS\<nome_cliente> usando como base
        as pastas que já funcionam no PC atual.
        """
        bases = []
        if getattr(self, "pasta_saida_final", None):
            bases.append(Path(self.pasta_saida_final))
        if getattr(self, "pasta_faturamentos", None):
            bases.append(Path(self.pasta_faturamentos))

        for base in bases:
            for p in [base] + list(base.parents):
                nome_pasta = p.name.strip().upper()
                if "01. FATURAMENTOS" in nome_pasta:
                    pasta = p / nome_cliente
                    if pasta.exists():
                        return pasta

        raise FileNotFoundError(
            f"Não encontrei a pasta de modelos em ...\\01. FATURAMENTOS\\{nome_cliente} "
            f"(base testada: {[str(b) for b in bases]})"
        )



    def gerar_planilha_calculo_conesul(self):
        try:
            # ==========================
            # 🔒 TRAVA DE SEGURANÇA
            # ==========================
            ws = self.ws_front
            empresa = ws.range("C9").value

            if not empresa or str(empresa).strip().upper() != "A/C CONE SUL AGÊNCIA DE NAVEGAÇÃO LTDA.":
                print("ℹ️ Planilha de cálculo não gerada (empresa não é CONESUL).")
                return


            # ==========================
            # 🔤 FUNÇÃO AUXILIAR
            # ==========================
            def remover_acentos(txt: str) -> str:
                return unicodedata.normalize("NFD", txt).encode("ascii", "ignore").decode("utf-8")

            # ==========================
            # 📂 PASTA DO MODELO (BASE)
            # ==========================
            pasta_modelo = self.encontrar_pasta_modelo("CONESUL")


            if not pasta_modelo.exists():
                raise FileNotFoundError(f"Pasta modelo não encontrada: {pasta_modelo}")

            # ==========================
            # 📂 PASTA DO NAVIO (DESTINO)
            # ==========================
            pasta_navio = self.pasta_saida_final
            pasta_navio.mkdir(parents=True, exist_ok=True)

            # ==========================
            # 🔎 LOCALIZAR MODELO EXCEL
            # ==========================
            modelo = None
 
            for arq in pasta_modelo.glob("*.xlsx"):
                nome_limpo = remover_acentos(arq.name.lower())
                if "calculo" in nome_limpo:
                    modelo = arq
                    break

            if not modelo:
                raise FileNotFoundError(
                    f"Nenhum modelo de cálculo encontrado em {pasta_modelo}"
                )
            

            # ==========================
            # 📄 COPIAR MODELO
            # ==========================
            destino = pasta_navio / "CALCULO - YUTA.xlsx"
            shutil.copy2(modelo, destino)

            # ==========================
            # 📊 ABRIR PLANILHA
            # ==========================
            wb = openpyxl.load_workbook(destino)
            ws = wb.active  # ou wb["Cálculo"] se quiser fixar

            # ==========================
            # 📥 PEGAR ÚLTIMA LINHA DO OGMO
            # ==========================
            ws1 = self.ws1  # ✔ CONFIRMADO no teu fluxo

            ultima_linha = self.ultima_linha_com_valor(ws1, "G")



            print(f"📊 Última linha detectada no NAVIO: {ultima_linha}")

            mapa = {
                "C5": "G",
                "D5": "H",
                "E5": "I",
                "F5": "N",
                "G5": "O",
                "H5": "P",
                "I5": "Q",
                "J5": "S",
                "K5": "T",
                "L5": "V",
                "M5": "Z",
            }

            for destino_cell, origem_col in mapa.items():
                valor = ws1[f"{origem_col}{ultima_linha}"].value
                ws[destino_cell] = valor
                print(f"   🔹 {origem_col}{ultima_linha} → {destino_cell} | Valor: {valor}")

            # ==========================
            # ➕ CAMPOS ADICIONAIS
            # ==========================

            # AA (última linha OGMO) → B3
            valor_aa = ws1[f"AA{ultima_linha}"].value
            ws["B3"] = valor_aa
            print(f"   🔹 AA{ultima_linha} → B3 | Valor: {valor_aa}")

            # ==========================
            # 🚢 NOME DO NAVIO
            # ==========================
            nome_navio = self.nome_navio  # ajuste se o atributo tiver outro nome

            ws["A4"] = nome_navio

            print(f"   🔹 NAVIO → A2 e A4 | Valor: {nome_navio}")





            # ==========================
            # 💾 SALVAR
            # ==========================
            wb.save(destino)

            print("✅ Planilha CÁLCULO CONESUL gerada com sucesso!")

        except Exception as e:
            print(f"❌ Erro ao gerar planilha CÁLCULO CONESUL: {e}")
            raise




    def obter_valor_cargonave(self):
        """
        Retorna o valor do adiantamento CARGONAVE
        (lido direto do FRONT – célula H28)
        """
        valor = self.ws_front.range("H28").value
        try:
            return float(valor)
        except:
            return 0.0


# ===== DATAS / UTILITARIOS =====#


    def data_por_extenso(self, valor):
        if isinstance(valor, datetime):
            data = valor
        elif isinstance(valor, date):
            data = datetime(valor.year, valor.month, valor.day)
        elif isinstance(valor, str):
            try:
                data = datetime.strptime(valor, "%d/%m/%Y")
            except:
                return ""
        else:
            return ""
        return data.strftime("%d de %B de %Y")


    def obter_datas_extremos(self, ws_resumo):
        last_row = ws_resumo.used_range.last_cell.row
        valores = ws_resumo.range(f"B1:B{last_row}").value
        datas = []
        hoje = date.today()
        for v in valores:
            if v in (None, "", "Total"):
                continue
            if isinstance(v, datetime):
                d = v.date()
                if d == hoje:
                    continue
                datas.append(d)
                continue
            if isinstance(v, str):
                s = v.strip()
                try:
                    d = datetime.strptime(s, "%d/%m/%Y").date()
                    if d != hoje:
                        datas.append(d)
                    continue
                except:
                    pass
        if not datas:
            return None, None
        return min(datas), max(datas)


    def obter_periodos(self, ws_resumo):
        """
        Retorna o último valor válido da coluna AA como inteiro.
        """
        last_row = ws_resumo.used_range.last_cell.row
        # Lê toda a coluna AA
        valores = ws_resumo.range(f"AA1:AA{last_row}").value

        if not valores:
            return 1  # padrão

        # Garante que 'valores' seja uma lista
        if not isinstance(valores, list):
            valores = [valores]

        # Procura o último valor não vazio
        for v in reversed(valores):
            if v is not None and v != "":
                try:
                    return int(float(v))
                except:
                    continue

        return 1





class FaturamentoAtipico(FaturamentoCompleto):
    """
    Faturamento ATÍPICO:
    - NÃO gera ciclos por regra.
    - Lê linhas reais do RESUMO (NAVIO): B=data, C=periodo, Z=valor
    - Replica no REPORT VIGIA: C=data, E=periodo, G=valor
    - ✅ Corrige ordem: Data crescente + período (06x12,12x18,18x24,00x06)
    """

    # ordem oficial dos períodos no REPORT
    _RANK_PERIODO = {
        "06x12": 0,
        "12x18": 1,
        "18x24": 2,
        "00x06": 3,
    }

    # -------------------------
    # Normalização robusta do período vindo da coluna C
    # Aceita: "06h", "12h", "18h", "00h" e também "06x12"
    # -------------------------
    def normalizar_periodo_c(self, valor_c) -> str | None:
        if not valor_c:
            return None

        s = str(valor_c).strip().lower()
        s = s.replace(" ", "")

        # ✅ caso típico do teu atípico: "06h", "12h", "18h", "00h"
        m_h = re.match(r"^(\d{1,2})h$", s)
        if m_h:
            hh = int(m_h.group(1)) % 24
            mapa = {0: "00x06", 6: "06x12", 12: "12x18", 18: "18x24"}
            return mapa.get(hh)

        # ✅ aceita "06", "12", "18", "00" (às vezes vem sem 'h')
        if re.match(r"^\d{1,2}$", s):
            hh = int(s) % 24
            mapa = {0: "00x06", 6: "06x12", 12: "12x18", 18: "18x24"}
            return mapa.get(hh)

        # ✅ aceita formatos com x/h/-, etc: "06x12", "06-12", "06:12", "06h12"
        s = s.replace("h", "")
        s = s.replace(":", "x").replace("-", "x").replace("×", "x")
        s = re.sub(r"[^0-9x]", "", s)

        m = re.match(r"^(\d{1,2})x(\d{1,2})$", s)
        if not m:
            return None

        a = int(m.group(1)) % 24
        b = int(m.group(2)) % 24
        periodo = f"{a:02d}x{b:02d}"

        return periodo if periodo in self._RANK_PERIODO else None

    # -------------------------
    # Extrai linhas reais do RESUMO (B,C,Z) e já devolve ORDENADO
    # -------------------------
    def extrair_linhas_atipico_resumo(self, ws_resumo, linha_inicio=2):
        last_row = ws_resumo.used_range.last_cell.row

        col_b = ws_resumo.range(f"B{linha_inicio}:B{last_row}").value
        col_c = ws_resumo.range(f"C{linha_inicio}:C{last_row}").value
        col_z = ws_resumo.range(f"Z{linha_inicio}:Z{last_row}").value

        if not isinstance(col_b, list): col_b = [col_b]
        if not isinstance(col_c, list): col_c = [col_c]
        if not isinstance(col_z, list): col_z = [col_z]

        linhas = []
        data_atual = None

        for i in range(len(col_b)):
            b = col_b[i]
            c = col_c[i]
            z = col_z[i]

            # ignora linhas "Total"
            if isinstance(c, str) and c.strip().lower().startswith("total"):
                continue

            # atualiza data quando B vem preenchido
            if isinstance(b, datetime):
                data_atual = b.date()
            elif isinstance(b, date):
                data_atual = b
            elif isinstance(b, str) and b.strip():
                try:
                    data_atual = datetime.strptime(b.strip(), "%d/%m/%Y").date()
                except:
                    pass

            if not data_atual:
                continue

            periodo = self.normalizar_periodo_c(c)
            if not periodo:
                continue

            try:
                valor = self.extrair_numero_excel(z)
            except:
                continue

            # guarda também índice original pra desempate
            linhas.append((data_atual, periodo, float(valor), i))

        if not linhas:
            return []

        # ✅ AQUI está a correção da ordem:
        # 1) data crescente
        # 2) período na ordem fixa (06x12,12x18,18x24,00x06)
        # 3) desempate pelo índice original (mantém estabilidade)
        linhas.sort(key=lambda x: (x[0], self._RANK_PERIODO.get(x[1], 99), x[3]))

        # remove o índice antes de devolver
        return [(d, p, v) for (d, p, v, _) in linhas]

    # -------------------------
    # Monta o REPORT VIGIA baseado nas linhas extraídas
    # -------------------------
    def montar_report_atipico(self):
        ws_report = self.wb2.sheets["REPORT VIGIA"]
        ws_resumo = self.ws1

        linhas = self.extrair_linhas_atipico_resumo(ws_resumo, linha_inicio=2)
        if not linhas:
            raise RuntimeError("ATÍPICO: não encontrei linhas (B/C/Z) válidas no RESUMO do NAVIO.")

        linha_base = 22
        n = len(linhas)

        self.inserir_linhas_report(ws_report, linha_inicial=linha_base, periodos=n)

        for i, (d, p, v) in enumerate(linhas):
            linha = linha_base + i
            ws_report.range(f"C{linha}").value = d
            ws_report.range(f"E{linha}").value = p
            cell = ws_report.range(f"G{linha}")
            cell.value = v
            cell.api.NumberFormatLocal = "R$ #.##0,00"

        d_min = min(x[0] for x in linhas)
        d_max = max(x[0] for x in linhas)
        self.ws_front.range("D16").value = self.data_por_extenso(d_min)
        self.ws_front.range("D17").value = self.data_por_extenso(d_max)

        print(f"✅ ATÍPICO: Report montado e ORDENADO com {n} linhas.")
        return linhas

    def processar(self):
        self.preencher_front_vigia()
        self.processar_MMO(self.wb1, self.wb2)
        self.montar_report_atipico()

        self.arredondar_para_baixo_50_se_cargonave()
        self.gerar_recibo_cargonave_word()
        self.gerar_planilha_calculo_cargonave()
        self.gerar_planilha_calculo_conesul()

        print("✅ FATURAMENTO ATÍPICO finalizado com sucesso!")


# ==============================
# CLASSE 2: FATURAMENTO DE ACORDO
# ==============================

class FaturamentoDeAcordo:

    @staticmethod
    def limpar_celula_segura(ws, endereco):
        rng = ws.range(endereco)
        if rng.merge_cells:
            rng.merge_area.clear_contents()
        else:
            rng.clear_contents()

    @staticmethod
    def escrever_celula_segura(ws, endereco, valor):
        rng = ws.range(endereco)
        if rng.merge_cells:
            rng.merge_area.value = valor
        else:
            rng.value = valor


    @staticmethod
    def aplicar_regras(ws_front, regras):
        for celula, valor in regras.items():

            FaturamentoDeAcordo.limpar_celula_segura(ws_front, celula)

            if valor is not None:
                FaturamentoDeAcordo.escrever_celula_segura(ws_front, celula, valor)

        for extra in ("C27", "G27"):
            FaturamentoDeAcordo.limpar_celula_segura(ws_front, extra)


    
        # =========================
    # REGRAS POR CLIENTE
    # =========================
    REGRAS_CLIENTES = {
        "Unimar Agenciamentos": {
            "G26": 500,
            "C27": None,
            "C35": 25,
        },
        "A/C Delta Agenciamento Marítimo Ltda.": {
            "G26": 500,
            "C27": None,
        },
        "A/C NORTH STAR SUDESTE SERVIÇOS MARÍTIMOS LTDA.": {
            "G26": 500,
            "C27": None,
            "C28": None,
            "C29": None,
            "H28": None,
            "H29": None,
        },
    }


    # =========================
    # APLICA REGRAS
    # =========================
    @staticmethod
    def aplicar_regras_cliente(ws_front):
        cliente_c9 = str(ws_front.range("C9").value or "").strip()

        for nome_cliente, regras in FaturamentoDeAcordo.REGRAS_CLIENTES.items():
            if nome_cliente in cliente_c9:
                FaturamentoDeAcordo.aplicar_regras(ws_front, regras)
                print(f"🔧 Regras aplicadas para cliente: {nome_cliente}")
                return

        print("ℹ️ Nenhuma regra específica de cliente aplicada.")



    def executar(self):
        print("🚀 Iniciando execução (DE ACORDO)...")

        pasta_faturamentos = obter_pasta_faturamentos()
        pasta_navio = selecionar_pasta_navio()

        dn = obter_dn_da_pasta(pasta_navio)
        nome_navio = obter_nome_navio(pasta_navio)

        nome_base = montar_nome_faturamento(dn, nome_navio)

        app = wb = ws_front = None

        try:
            app, wb, ws_front = abrir_workbooks_de_acordo(
                pasta_faturamentos,
                pasta_navio
            )

            print(f"📋 DN: {dn}")
            print(f"🚢 Navio: {nome_navio}")
            escrever_de_acordo_nf(wb, nome_navio, dn, ano=datetime.now().year)

            hoje = datetime.now()
            meses = ["", "janeiro","fevereiro","março","abril","maio","junho",
                    "julho","agosto","setembro","outubro","novembro","dezembro"]
            data_extenso = f"{hoje.day} de {meses[hoje.month]} de {hoje.year}"

            # -------- PREENCHIMENTO FRONT --------
            ws_front.range("D15").value = nome_navio
            ws_front.range("C21").value = f"DN {str(dn).zfill(3)}/{hoje.strftime('%y')}"

            ws_front.range("D16").value = data_extenso
            ws_front.range("D17").value = data_extenso
            ws_front.range("D18").value = "-"
            ws_front.range("C26").value = f"DE ACORDO ( M/V {nome_navio} )"
            ws_front.range("C39").value = f" Santos, {data_extenso}"

            # 🔧 Regras por cliente
            self.aplicar_regras_cliente(ws_front)

            print("✅ Faturamento De Acordo concluído!")

            # ✅ SALVAR EXCEL (ainda dentro do try, com wb aberto)
            caminho_excel = salvar_excel_com_nome(
                wb=wb,
                pasta_saida=pasta_navio,
                nome_base=nome_base
            )
            print(f"💾 Excel salvo em: {caminho_excel}")

            # ✅ GERAR PDF (passando caminho_excel corretamente)
            caminho_pdf = gerar_pdf(
                caminho_excel=caminho_excel,
                pasta_saida=pasta_navio,
                nome_base=nome_base,
                ws=ws_front
            )
            print(f"📑 PDF FRONT salvo em: {caminho_pdf}")

        finally:
            fechar_workbooks(app=app, wb_cliente=wb)



# ==============================
# CLASSE 3: Fazer Ponto
# ==============================


class ProgramaCopiarPeriodo:
    def __init__(self, debug=False):
        self.debug = debug
        self.app = None
        self.wb = None
        self.wb_navio = None
        self.wb_cliente = None
        self.ws = None
        self.ws_front = None
        self.pasta_navio = None
        self.datas = []


    PERIODOS_MENU = {"1": "06h", "2": "12h", "3": "18h", "4": "00h"}
    MAPA_PERIODOS = {
        "06h": "06h", "6h": "06h", "06": "06h",
        "12h": "12h", "12": "12h",
        "18h": "18h", "18": "18h",
        "00h": "00h", "0h": "00h", "00": "00h", "24h": "00h"
    }
    EQUIVALENTES = {
        "06h": ["06h", "12h"],
        "12h": ["12h", "06h"],
        "18h": ["18h", "00h"],
        "00h": ["00h", "18h"]
    }
    BLOCOS = {"06h": 1, "12h": 1, "18h": 2, "00h": 2}


    # ---------------------------
    # Abrir arquivo NAVIO
    # ---------------------------



    def abrir_arquivo_navio(self):
        caminho = selecionar_arquivo_navio()
        if not caminho:
            raise FileNotFoundError("Arquivo do NAVIO não selecionado")

        self.app = xw.App(visible=False, add_book=False)
        self.wb_navio = self.app.books.open(caminho)
        self.wb = self.wb_navio
        self.ws = self.wb.sheets[0]



    # ---------------------------
    # Datas
    # ---------------------------
    def is_domingo(self, data_str):
        d = datetime.strptime(data_str, "%d/%m/%Y")
        return d.weekday() == 6

    def is_feriado(self, data_str):
        d = datetime.strptime(data_str, "%d/%m/%Y")
        return d in feriados_br

    def is_dia_bloqueado(self, data_str):
        """
        Retorna True se for domingo ou feriado nacional
        data_str no formato DD/MM/YYYY
        """
        data = datetime.strptime(data_str, "%d/%m/%Y").date()

        # Domingo
        if data.weekday() == 6:
            return True

        # Feriado nacional
        if data in feriados_br:
            return True

        return False

    def parse_data(self, data_str):
        return datetime.strptime(data_str, "%d/%m/%Y")

    def normalizar_texto(self, texto):
        return str(texto).lower().replace(" ", "")

    def normalizar_periodo(self, texto):
        t = self.normalizar_texto(texto)
        return self.MAPA_PERIODOS.get(t, None)




    # ---------------------------
    # Datas
    # ---------------------------
    def carregar_datas(self):
        ultima = self.ws.range("B" + str(self.ws.cells.last_cell.row)).end("up").row
        datas = []
        for i in range(1, ultima + 1):
            v = self.ws.range(f"B{i}").value
            if isinstance(v, (datetime, date)):
                datas.append(v.strftime("%d/%m/%Y"))
            elif isinstance(v, str) and "/" in v:
                datas.append(v.strip())
        self.datas = list(dict.fromkeys(datas))
        if not self.datas:
            raise Exception("Nenhuma data encontrada na coluna B.")

    def escolher_data(self):
        print("\nDatas disponíveis:")
        for i, d in enumerate(self.datas, 1):
            print(f"{i} - {d}")
        while True:
            try:
                return self.datas[int(input("Escolha a data: ")) - 1]
            except:
                print("Opção inválida.")

    def escolher_periodo(self):
        print("\nHorário:")
        print("1 = 06h | 2 = 12h | 3 = 18h | 4 = 00h")
        while True:
            op = input("Opção: ").strip()
            if op in self.PERIODOS_MENU:
                return self.PERIODOS_MENU[op]


    # ---------------------------
    # Localização
    # ---------------------------
    


    def encontrar_linha_data(self, data_str):
        ultima = self.ws.range("B" + str(self.ws.cells.last_cell.row)).end("up").row
        for i in range(1, ultima + 1):
            valor = self.ws.range(f"B{i}").value
            if isinstance(valor, (datetime, date)) and valor.strftime("%d/%m/%Y") == data_str:
                return i
            elif valor == data_str:
                return i
        raise Exception(f"Data {data_str} não encontrada.")

    def encontrar_total_data(self, linha_data):
        i = linha_data + 1
        while True:
            valor_c = self.ws.range(f"C{i}").value
            valor_a = self.ws.range(f"A{i}").value
            if isinstance(valor_a, str) and self.normalizar_texto(valor_a) == "totalgeral":
                raise Exception("❌ Total do dia não encontrado antes do Total Geral")
            if isinstance(valor_c, str) and self.normalizar_texto(valor_c) == "total":
                return i
            if i > self.ws.cells.last_cell.row:
                raise Exception("❌ Fim da planilha sem encontrar 'Total' do dia")
            i += 1

    # ---------------------------
    # Buscar modelo inteligente
    # ---------------------------

    def encontrar_modelo_periodo(self, data_destino, periodo):
        """
        Retorna: (linha_modelo, data_modelo)
        """

        datas_ordenadas = sorted(self.datas, key=lambda d: self.parse_data(d))
        if data_destino not in datas_ordenadas:
            raise Exception(f"Data base {data_destino} não está na lista de datas válidas")

        idx = datas_ordenadas.index(data_destino)

        def procurar_em_data(data, aceitar_equivalente):
            linha_data = self.encontrar_linha_data(data)
            i = linha_data + 1

            while True:
                valor_a = self.ws.range(f"A{i}").value
                valor_c = self.ws.range(f"C{i}").value

                if isinstance(valor_a, str) and self.normalizar_texto(valor_a) == "totalgeral":
                    return None

                if not isinstance(valor_c, str):
                    i += 1
                    continue

                texto = self.normalizar_texto(valor_c)

                if texto == "total":
                    return None

                p = self.normalizar_periodo(texto)
                if not p:
                    i += 1
                    continue

                if p == periodo:
                    if self.debug:
                        print(f"✔ Usando {p} da data {data}")
                    return i, data

                if aceitar_equivalente and p in self.EQUIVALENTES.get(periodo, []):
                    if self.debug:
                        print(f"⚠ Usando equivalente {p} da data {data}")
                    return i, data

                i += 1

        # 1️⃣ Mesmo dia
        resultado = procurar_em_data(data_destino, aceitar_equivalente=True)
        if resultado:
            return resultado

        # 2️⃣ Outros dias
        for offset in range(1, len(datas_ordenadas)):
            for novo_idx in (idx - offset, idx + offset):
                if 0 <= novo_idx < len(datas_ordenadas):
                    data = datas_ordenadas[novo_idx]

                    if self.is_dia_bloqueado(data):
                        if self.debug:
                            print(f"⛔ Pulando data bloqueada: {data}")
                        continue

                    resultado = procurar_em_data(data, aceitar_equivalente=False)
                    if resultado:
                        return resultado

        raise Exception(
            f"Nenhum modelo encontrado para o período '{periodo}' "
            f"a partir da data {data_destino}"
        )

    # ---------------------------
    # Copiar e colar
    # ---------------------------

        
    def copiar_colar(self, data, periodo):
        if self.is_dia_bloqueado(data):
            print(f"⛔ {data} é domingo ou feriado — período não será criado")
            return

        # ⚠️ CHAMAR APENAS UMA VEZ
        linha_modelo, data_modelo = self.encontrar_modelo_periodo(data, periodo)

        linha_data = self.encontrar_linha_data(data)
        linha_total_dia = self.encontrar_total_data(linha_data)

        print(
            f"\n✅ Executando FAZER PONTO no NAVIO - "
            f"Data: {data}, Período: {periodo} "
            f"(modelo: {data_modelo})"
        )  

        self.ws.api.Rows(linha_total_dia).Insert()

        if linha_modelo >= linha_total_dia:
            linha_modelo += 1

        self.ws.api.Rows(linha_modelo).Copy()
        self.ws.api.Rows(linha_total_dia).PasteSpecial(-4163)

        self.ws.api.Rows(linha_total_dia).Font.Bold = True
        self.ws.range((linha_total_dia, 3)).value = periodo

        linha_nova = linha_total_dia
        linha_total_dia += 1

        self.somar_linha_no_total_do_dia(linha_nova, linha_total_dia)
        self.somar_linha_no_total_geral(linha_nova)

        print("➕ Linha adicionada e somada ao TOTAL DO DIA e TOTAL GERAL")




    # ---------------------------
    # Soma totais
    # ---------------------------
    def somar_linha_no_total_do_dia(self, linha_origem, linha_total_dia):
        ultima_col = self.ws.range("A1").expand("right").last_cell.column
        for col in range(3, ultima_col + 1):
            v_origem = self.ws.range((linha_origem, col)).value
            v_total = self.ws.range((linha_total_dia, col)).value
            try:
                v_origem = float(v_origem)
            except:
                continue
            try:
                v_total = float(v_total or 0)
            except:
                v_total = 0
            self.ws.range((linha_total_dia, col)).value = v_total + v_origem
        if self.debug:
            print(f"➕ Linha {linha_origem} somada ao TOTAL DO DIA")

    def encontrar_linha_total_geral(self):
        ultima_linha = self.ws.cells.last_cell.row
        for i in range(1, ultima_linha + 1):
            valor_a = self.ws.range(f"A{i}").value
            if isinstance(valor_a, str) and self.normalizar_texto(valor_a) == "totalgeral":
                return i
        raise Exception("Total Geral não encontrado.")

    def somar_linha_no_total_geral(self, linha_origem):
        linha_total_geral = self.encontrar_linha_total_geral()
        ultima_col = self.ws.range("A1").expand("right").last_cell.column
        for col in range(4, ultima_col + 1):
            valor_origem = self.ws.range((linha_origem, col)).value
            if isinstance(valor_origem, (int, float)):
                celula_total = self.ws.range((linha_total_geral, col))
                celula_total.value = (celula_total.value or 0) + valor_origem
        if self.debug:
            print(f"➕ Linha {linha_origem} somada ao TOTAL GERAL")

    # ---------------------------
    # Executar
    # ---------------------------

    def executar(self, usar_arquivo_aberto=False):
        try:
            if not usar_arquivo_aberto or not self.ws:
                self.abrir_arquivo_navio()

            self.carregar_datas()
            data = self.escolher_data()
            periodo = self.escolher_periodo()
            self.copiar_colar(data, periodo)

            self.salvar()

        finally:
            if not usar_arquivo_aberto:
                fechar_workbooks(
                    app=self.app,
                    wb_navio=self.wb_navio,
                    wb_cliente=self.wb_cliente
                )



    def salvar(self):
        if not self.wb:
            raise Exception("Nenhum workbook aberto para salvar")

        self.wb.save()
        print("💾 Arquivo NAVIO salvo com sucesso")


# ==============================
# CLASSE 4: Remover Período 
# ==============================


class ProgramaRemoverPeriodo:
    def __init__(self, debug=False):
        self.debug = debug
        self.app = None
        self.wb = None
        self.wb_navio = None
        self.wb_cliente = None
        self.ws = None
        self.datas = []

    PERIODOS_MENU = {"1": "06h", "2": "12h", "3": "18h", "4": "00h"}
    MAPA_PERIODOS = {
        "06h": "06h", "6h": "06h", "06": "06h",
        "12h": "12h", "12": "12h",
        "18h": "18h", "18": "18h",
        "00h": "00h", "0h": "00h", "00": "00h", "24h": "00h"
    }

    # ---------------------------
    # Abrir arquivo NAVIO
    # ---------------------------

    def abrir_arquivo_navio(self):
        caminho = selecionar_arquivo_navio()
        if not caminho:
            return

        self.app = xw.App(visible=False, add_book=False)
        self.wb_navio = self.app.books.open(caminho)
        self.wb = self.wb_navio
        self.ws = self.wb.sheets[0]

    # ---------------------------
    # Utilidades
    # ---------------------------

    def normalizar_texto(self, texto):
        return str(texto).lower().replace(" ", "")

    def normalizar_periodo(self, texto):
        return self.MAPA_PERIODOS.get(self.normalizar_texto(texto))

    def parse_data(self, data_str):
        return datetime.strptime(data_str, "%d/%m/%Y")

    def is_dia_bloqueado(self, data_str):
        data = datetime.strptime(data_str, "%d/%m/%Y").date()
        if data.weekday() == 6:
            return True
        if data in feriados_br:
            return True
        return False

    def obter_nome_navio(self):
        return self.ws.range("A2").value
    
    



    # ---------------------------
    # Datas
    # ---------------------------

    def carregar_datas(self):
        ultima = self.ws.range("B" + str(self.ws.cells.last_cell.row)).end("up").row
        datas = []
        for i in range(1, ultima + 1):
            v = self.ws.range(f"B{i}").value
            if isinstance(v, (datetime, date)):
                datas.append(v.strftime("%d/%m/%Y"))
            elif isinstance(v, str) and "/" in v:
                datas.append(v.strip())

        self.datas = list(dict.fromkeys(datas))

    def escolher_data(self):
        print("\nDatas disponíveis:")
        for i, d in enumerate(self.datas, 1):
            print(f"{i} - {d}")
        while True:
            try:
                return self.datas[int(input("Escolha a data: ")) - 1]
            except:
                print("Opção inválida.")

    def escolher_periodo(self):
        print("\nHorário:")
        print("1 = 06h | 2 = 12h | 3 = 18h | 4 = 00h")
        while True:
            op = input("Opção: ").strip()
            if op in self.PERIODOS_MENU:
                return self.PERIODOS_MENU[op]

    # ---------------------------
    # Localização
    # ---------------------------

    def encontrar_linha_data(self, data_str):
        ultima = self.ws.range("B" + str(self.ws.cells.last_cell.row)).end("up").row
        for i in range(1, ultima + 1):
            v = self.ws.range(f"B{i}").value
            if isinstance(v, (datetime, date)) and v.strftime("%d/%m/%Y") == data_str:
                return i
            elif v == data_str:
                return i
        return None

    def encontrar_total_data(self, linha_data):
        i = linha_data + 1
        while True:
            valor_c = self.ws.range(f"C{i}").value
            valor_a = self.ws.range(f"A{i}").value

            if isinstance(valor_a, str) and self.normalizar_texto(valor_a) == "totalgeral":
                return None

            if isinstance(valor_c, str) and self.normalizar_texto(valor_c) == "total":
                return i

            i += 1

    def encontrar_linha_total_geral(self):
        ultima = self.ws.cells.last_cell.row
        for i in range(1, ultima + 1):
            v = self.ws.range(f"A{i}").value
            if isinstance(v, str) and self.normalizar_texto(v) == "totalgeral":
                return i
        return None

    # ---------------------------
    # Encontrar período EXATO
    # ---------------------------




    def encontrar_linha_periodo(self, data, periodo):
        linha_data = self.encontrar_linha_data(data)
        if not linha_data:
            return None

        # 🔴 REGRA ESPECIAL PARA 00h
        if periodo == "00h":
            linha_acima = linha_data - 1
            if linha_acima > 0:
                valor_c = self.ws.range(f"C{linha_acima}").value
                if isinstance(valor_c, str):
                    p = self.normalizar_periodo(valor_c)
                    if p == "00h":
                        return linha_acima

        # 🔽 Procura normal abaixo da data
        i = linha_data + 1
        while True:
            valor_a = self.ws.range(f"A{i}").value
            valor_c = self.ws.range(f"C{i}").value

            if isinstance(valor_a, str) and self.normalizar_texto(valor_a) == "totalgeral":
                return None

            if isinstance(valor_c, str):
                p = self.normalizar_periodo(valor_c)
                if p == periodo:
                    return i

                if self.normalizar_texto(valor_c) == "total":
                    return None

            i += 1


    # ---------------------------
    # Subtrações
    # ---------------------------

    def subtrair_total_dia(self, linha_origem, linha_total_dia):
        ultima_col = self.ws.range("A1").expand("right").last_cell.column
        for col in range(3, ultima_col + 1):
            v = self.ws.range((linha_origem, col)).value
            if isinstance(v, (int, float)):
                celula = self.ws.range((linha_total_dia, col))
                celula.value = (celula.value or 0) - v

    def subtrair_total_geral(self, linha_origem):
        linha_total_geral = self.encontrar_linha_total_geral()
        if not linha_total_geral:
            return

        ultima_col = self.ws.range("A1").expand("right").last_cell.column
        for col in range(4, ultima_col + 1):
            v = self.ws.range((linha_origem, col)).value
            if isinstance(v, (int, float)):
                celula = self.ws.range((linha_total_geral, col))
                celula.value = (celula.value or 0) - v

    # ---------------------------
    # Remover período
    # ---------------------------

    def remover_periodo(self, data, periodo):
        if self.is_dia_bloqueado(data):
            print(f"⛔ {data} é domingo ou feriado — nenhuma ação executada")
            return

        linha = self.encontrar_linha_periodo(data, periodo)
        if not linha:
            print(f"ℹ Período {periodo} não existe em {data} — nada a remover")
            return

        linha_data = self.encontrar_linha_data(data)
        linha_total_dia = self.encontrar_total_data(linha_data)

        print(f"\n🗑 Removendo período {periodo} — Data {data}")

        if linha_total_dia:
            self.subtrair_total_dia(linha, linha_total_dia)

        self.subtrair_total_geral(linha)

        self.ws.api.Rows(linha).Delete()

        print("➖ Linha removida e totais ajustados")

    # ---------------------------
    # Execução
    # ---------------------------

    def executar(self, usar_arquivo_aberto=False):
        try:
            if not usar_arquivo_aberto or not self.ws:
                self.abrir_arquivo_navio()

            self.carregar_datas()
            data = self.escolher_data()
            periodo = self.escolher_periodo()
            self.remover_periodo(data, periodo)

            self.salvar()

        finally:
            if not usar_arquivo_aberto:
                fechar_workbooks(
                    app=self.app,
                    wb_navio=self.wb_navio,
                    wb_cliente=self.wb_cliente
                )

    def salvar(self):
        if self.wb:
            self.wb.save()
            print("💾 Arquivo salvo com sucesso")


# ==============================
# CLASSE 5: FATURAMENTO SÃO SEBASTIÃO - OGMO
# ==============================


# -*- coding: utf-8 -*-



class FaturamentoSaoSebastiao:
    """
    ✅ Objetivo (organizado e estável):
    - Selecionar 1 ou MAIS PDFs (Sea Side geralmente vem com 2)
    - Ler TODOS os PDFs selecionados e manter quebras de linha
    - Identificar cliente/porto pela pasta do CLIENTE
    - Se for layout SS (Wilson SS / Sea Side PSS):
        - extrair valores somando entre PDFs (se tiver 2)
        - colar no REPORT VIGIA com o MAPA_FIXO (você já deixou as células)
        - preencher FRONT VIGIA
        - preencher CREDIT NOTE se existir
    - Se for cliente padrão (Aquarius e outros):
        - usar report padrão (datas e horários)
        - (extração financeira pode ser diferente: por enquanto fica como TODO)

    ⚠️ IMPORTANTE:
    - Eu NÃO removo '\n' na normalização, porque sua extração depende de splitlines().
    - A extração do layout SS soma automaticamente tudo que casar (ótimo pra Sea Side com 2 PDFs).
    """

    # ==================================================
    # INIT
    # ==================================================
    def __init__(self):
        self.caminhos_pdfs: list[Path] = []
        self.paginas_texto: list[dict] = []   # [{pdf, page, texto}]
        self.texto_pdf: str = ""
        self.dados: dict[str, float] = {}

    # ==================================================
    # UTIL: NORMALIZAÇÃO
    # ==================================================
    def _normalizar(self, s: str | None) -> str:
        if not s:
            return ""
        s = unicodedata.normalize("NFKD", str(s))
        s = s.encode("ASCII", "ignore").decode("ASCII")
        s = s.replace("-", " ")  # ajuda "sea-side" -> "sea side"
        return re.sub(r"\s+", " ", s).strip().lower()

    def _br_to_float(self, valor) -> float:
        """Converte '1.721,08' -> 1721.08 ; aceita float/int direto."""
        if valor in (None, "", "NÃO ENCONTRADO"):
            return 0.0
        if isinstance(valor, (int, float)):
            return float(valor)
        return float(str(valor).replace(".", "").replace(",", ".").strip())

    # Alias (você usava _to_float em alguns lugares)
    def _to_float(self, valor) -> float:
        return self._br_to_float(valor)

    # ==================================================
    # UTIL: EXCEL
    # ==================================================
    def _achar_aba(self, wb, nomes_possiveis: list[str]):
        for sheet in wb.sheets:
            nome = sheet.name.strip().lower()
            for n in nomes_possiveis:
                if nome == n.strip().lower():
                    return sheet
        raise RuntimeError(f"Aba não encontrada. Esperado uma de: {nomes_possiveis}")

    # ==================================================
    # IDENTIFICAÇÃO CLIENTE / PORTO
    # ==================================================
    def identificar_cliente_e_porto(self) -> tuple[str, str]:
        """
        Identifica cliente e porto pelo nome da pasta do CLIENTE (pai da pasta do navio).
        """
        if not self.caminhos_pdfs:
            raise RuntimeError("Nenhum PDF carregado para identificar cliente/porto.")

        pasta_navio = self.caminhos_pdfs[0].parent
        pasta_cliente = pasta_navio.parent

        nome_norm = self._normalizar(pasta_cliente.name)

        # WILSON SONS — SÃO SEBASTIÃO
        if "wilson" in nome_norm and "sebastiao" in nome_norm:
            return "WILSON SONS", "SAO SEBASTIAO"

        # SEA SIDE — PSS (mesmo layout de colagem do report)
        if "sea side" in nome_norm and "pss" in nome_norm:
            return "SEA SIDE", "PSS"

        # PADRÃO
        return pasta_cliente.name.strip().upper(), "PADRAO"

    def _usa_layout_ss(self, cliente: str, porto: str) -> bool:
        return (
            (cliente == "WILSON SONS" and porto == "SAO SEBASTIAO")
            or (cliente == "SEA SIDE" and porto == "PSS")
        )

    # ==================================================
    # PDF: SELEÇÃO E LEITURA (MULTI)
    # ==================================================
    def selecionar_pdfs_ogmo(self):
        root = Tk()
        root.withdraw()
        root.attributes("-topmost", True)

        caminhos = filedialog.askopenfilenames(
            title="Selecione 1 ou MAIS PDFs OGMO (Sea Side pode ter 2)",
            filetypes=[("PDF", "*.pdf")]
        )
        root.destroy()

        if not caminhos:
            raise RuntimeError("Nenhum PDF selecionado")

        self.caminhos_pdfs = [Path(c) for c in caminhos]
        print("📄 PDFs selecionados:")
        for p in self.caminhos_pdfs:
            print(f"   - {p.name}")

    def carregar_pdfs(self):
        self.paginas_texto.clear()

        for caminho in self.caminhos_pdfs:
            with pdfplumber.open(str(caminho)) as pdf:
                for i, page in enumerate(pdf.pages, start=1):
                    txt = page.extract_text() or ""
                    txt = txt.strip()

                    # se veio texto, guarda direto (SEM OCR)
                    if txt:
                        self.paginas_texto.append({"pdf": caminho.name, "page": i, "texto": txt, "src": "TXT"})
                    else:
                        ocr_txt = self._ocr_pagina(caminho, page_num=i)
                        if ocr_txt.strip():
                            self.paginas_texto.append({"pdf": caminho.name, "page": i, "texto": ocr_txt, "src": "OCR"})

        if not self.paginas_texto:
            raise RuntimeError("Nenhuma página com texto (nem pdfplumber nem OCR).")


        self.normalizar_texto_mantendo_linhas()




    
    def _money_to_float(self, s: str) -> float:
        if s is None:
            return 0.0
        s = str(s).strip()

        # remove espaços (OCR adora meter)
        s = s.replace(" ", "")

        # se tem vírgula e ponto, decide o decimal pelo ÚLTIMO separador
        if "," in s and "." in s:
            if s.rfind(",") > s.rfind("."):
                # 1.234,56  -> decimal = ,
                s = s.replace(".", "").replace(",", ".")
            else:
                # 1,234.56 -> decimal = .
                s = s.replace(",", "")
            return float(s)

        # só vírgula: 1234,56
        if "," in s:
            return float(s.replace(".", "").replace(",", "."))

        # só ponto: 1234.56
        return float(s)

    def _poppler_paths_candidatos(self) -> list[Path]:
        candidatos = []

        env_poppler = os.environ.get("POPPLER_PATH")
        if env_poppler:
            candidatos.append(Path(env_poppler))

        path_env = os.environ.get("PATH", "")
        for parte in path_env.split(os.pathsep):
            if parte and "poppler" in parte.lower():
                candidatos.append(Path(parte))

        if getattr(sys, "frozen", False):
            meipass = Path(getattr(sys, "_MEIPASS", Path(sys.executable).parent))
            exe_dir = Path(sys.executable).resolve().parent
            candidatos.extend(
                [
                    meipass / "poppler" / "Library" / "bin",
                    meipass / "poppler" / "bin",
                    exe_dir / "poppler" / "Library" / "bin",
                    exe_dir / "poppler" / "bin",
                ]
            )
        else:
            raiz_projeto = Path(__file__).resolve().parent
            candidatos.extend(
                [
                    raiz_projeto / "poppler" / "Library" / "bin",
                    raiz_projeto / "poppler" / "bin",
                ]
            )

        candidatos.extend(
            [
                Path(r"C:\poppler-25.12.0\Library\bin"),
                Path(r"C:\poppler\Library\bin"),
                Path(r"C:\Program Files\poppler\Library\bin"),
                Path(r"C:\Program Files (x86)\poppler\Library\bin"),
            ]
        )

        vistos = set()
        validos = []
        for pasta in candidatos:
            chave = str(pasta).lower().strip()
            if not chave or chave in vistos:
                continue
            vistos.add(chave)
            if pasta.exists() and (pasta / "pdfinfo.exe").exists():
                validos.append(pasta)
        return validos

    def _configurar_tesseract(self):
        candidatos = []

        env_tesseract = os.environ.get("TESSERACT_EXE")
        if env_tesseract:
            candidatos.append(Path(env_tesseract))

        if getattr(sys, "frozen", False):
            meipass = Path(getattr(sys, "_MEIPASS", Path(sys.executable).parent))
            exe_dir = Path(sys.executable).resolve().parent
            candidatos.extend(
                [
                    meipass / "tesseract" / "tesseract.exe",
                    exe_dir / "tesseract" / "tesseract.exe",
                ]
            )

        candidatos.extend(
            [
                Path(r"C:\Program Files\Tesseract-OCR\tesseract.exe"),
                Path(r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe"),
            ]
        )

        for exe in candidatos:
            if exe.exists():
                pytesseract.pytesseract.tesseract_cmd = str(exe)
                tessdata_dir = exe.parent / "tessdata"
                if tessdata_dir.exists():
                    os.environ["TESSDATA_PREFIX"] = str(tessdata_dir)
                return



    def _ocr_pagina(self, caminho_pdf: Path, page_num: int, dpi: int = 350, lang: str = "por") -> str:

        self._configurar_tesseract()

        imgs = []
        erros = []

        try:
            imgs = convert_from_path(
                str(caminho_pdf),
                dpi=dpi,
                grayscale=True,
                first_page=page_num,
                last_page=page_num,
            )
        except Exception as exc:
            erros.append(str(exc))

        if not imgs:
            for poppler_dir in self._poppler_paths_candidatos():
                try:
                    imgs = convert_from_path(
                        str(caminho_pdf),
                        dpi=dpi,
                        grayscale=True,
                        poppler_path=str(poppler_dir),
                        first_page=page_num,
                        last_page=page_num,
                    )
                    if imgs:
                        break
                except Exception as exc:
                    erros.append(str(exc))

        if not imgs:
            if erros:
                print(f"⚠️ OCR indisponível (Poppler/Tesseract): {erros[-1]}")
            return ""

        return pytesseract.image_to_string(imgs[0], lang=lang, config="--oem 3 --psm 6")



    def normalizar_texto_mantendo_linhas(self):
        """
        Normaliza espaços mas NÃO remove '\\n'.
        Isso mantém sua extração por linha estável.
        """
        blocos = []
        for item in self.paginas_texto:
            texto = item["texto"]
            texto = "\n".join(re.sub(r"[ \t]+", " ", ln).strip() for ln in texto.splitlines())
            item["texto"] = texto
            blocos.append(texto)

        self.texto_pdf = "\n\n".join(blocos)



    # ==================================================
    # PDF ORDER (OGMO 1..N)  -> agora retorna Path (não só nome)
    # ==================================================
    def _ordenar_pdfs_ogmo(self) -> list[Path]:
        """
        Retorna a lista de Paths ordenada pelo número do arquivo:
        FOLHAS OGMO 1.pdf, 2.pdf, 3.pdf ...
        Se não achar número, joga pro final mantendo ordem original.
        """
        def idx(p: Path) -> int:
            nome = p.name
            m = re.search(r"\bOGMO\s*(\d+)\b|\b(\d+)\b", nome, re.IGNORECASE)
            if not m:
                return 10_000
            g = m.group(1) or m.group(2)
            try:
                return int(g)
            except Exception:
                return 10_000

        return sorted(self.caminhos_pdfs, key=idx)


    def _pdfs_ordenados_nomes(self) -> list[str]:
        """Nomes ordenados (string) - útil se você quiser logar."""
        return [p.name for p in self._ordenar_pdfs_ogmo()]


    # ==================================================
    # EXTRAÇÃO - DATA (tolerante a OCR) por PDF (case-insensitive)
    # ==================================================
    def extrair_periodo_por_data(self, pdf_alvo: str | None = None) -> tuple[str, str]:
        if pdf_alvo:
            alvo_norm = pdf_alvo.strip().lower()
            textos = [it["texto"] for it in self.paginas_texto
                    if str(it.get("pdf","")).strip().lower() == alvo_norm]
            texto_busca = "\n".join(textos)
        else:
            texto_busca = self.texto_pdf or ""

        if not texto_busca.strip():
            raise RuntimeError("Período não encontrado no PDF (texto vazio).")

        # tolerância OCR
        rx_per = re.compile(r"per(?:[íi]|l|1|f|0)?odo", re.I)
        rx_ini = re.compile(r"inic(?:ial|iaI|ia1|lal)?", re.I)
        rx_fim = re.compile(r"fina(?:l|I|1)?", re.I)
        rx_data = re.compile(r"\b(\d{1,2}/\d{1,2}/\d{4})\b")

        linhas = texto_busca.splitlines()

        def achar_data(bloco_rx) -> str | None:
            for i, ln in enumerate(linhas):
                ln_norm = ln.replace("\u00ad", "")
                if rx_per.search(ln_norm) and bloco_rx.search(ln_norm):
                    # tenta na mesma linha
                    m = rx_data.search(ln_norm)
                    if m:
                        return m.group(1)
                    # tenta nas próximas 2 linhas (OCR às vezes joga a data abaixo)
                    for j in range(i+1, min(i+3, len(linhas))):
                        m2 = rx_data.search(linhas[j])
                        if m2:
                            return m2.group(1)
            return None

        data_ini = achar_data(rx_ini)
        data_fim = achar_data(rx_fim)

        if not data_ini or not data_fim:
            raise RuntimeError(f"Período (datas) não encontrado. ini={data_ini} fim={data_fim}")

        return data_ini, data_fim


    # ==================================================
    # EXTRAÇÃO - HORÁRIO (tolerante a OCR) por PDF (case-insensitive)
    # ==================================================
    def extrair_periodo_por_horario(self, pdf_alvo: str | None = None) -> tuple[str, str]:
        if pdf_alvo:
            alvo_norm = pdf_alvo.strip().lower()
            textos = [it["texto"] for it in self.paginas_texto
                    if str(it.get("pdf","")).strip().lower() == alvo_norm]
            texto_busca = "\n".join(textos)
        else:
            texto_busca = self.texto_pdf or ""

        if not texto_busca.strip():
            raise RuntimeError("Horários não encontrados (texto vazio).")

        rx_per = re.compile(r"per(?:[íi]|l|1|f|0)?odo", re.I)
        rx_ini = re.compile(r"inic(?:ial|iaI|ia1|lal)?", re.I)
        rx_fim = re.compile(r"fina(?:l|I|1)?", re.I)

        # aceita 07x13, 07×13, 07-13, 07h13
        rx_h = re.compile(r"\b(\d{1,2})\s*[x×h\-]\s*(\d{1,2})\b", re.I)

        linhas = texto_busca.splitlines()

        def achar_horario(bloco_rx) -> str | None:
            for i, ln in enumerate(linhas):
                if rx_per.search(ln) and bloco_rx.search(ln):
                    m = rx_h.search(ln)
                    if m:
                        a, b = int(m.group(1)) % 24, int(m.group(2)) % 24
                        return f"{a:02d}x{b:02d}"
                    for j in range(i+1, min(i+3, len(linhas))):
                        m2 = rx_h.search(linhas[j])
                        if m2:
                            a, b = int(m2.group(1)) % 24, int(m2.group(2)) % 24
                            return f"{a:02d}x{b:02d}"
            return None

        p_ini = achar_horario(rx_ini)
        p_fim = achar_horario(rx_fim)

        if not p_ini or not p_fim:
            raise RuntimeError(f"Período (horários) não encontrado. ini={p_ini} fim={p_fim}")

        # validação
        ordem = {"07x13", "13x19", "19x01", "01x07"}
        if p_ini not in ordem or p_fim not in ordem:
            raise RuntimeError(f"Horários inválidos: ini={p_ini} fim={p_fim}")

        return p_ini, p_fim

    # ==================================================
    # PERÍODO MESCLADO N PDFs (primeiro que tem INI, último que tem FIM)
    # ==================================================


    
    def extrair_datas_mescladas(self) -> tuple[str, str]:
        pdfs = self._ordenar_pdfs_ogmo()
        if not pdfs:
            raise RuntimeError("Nenhum PDF selecionado.")

        # ✅ início = menor OGMO (normalmente 1)
        p_ini = self._achar_pdf_menor_numero() or pdfs[0]

        # ✅ fim = maior OGMO (último: 2, 3, 4...)
        p_fim = self._achar_pdf_maior_numero() or pdfs[-1]

        try:
            di, _ = self.extrair_periodo_por_data(p_ini.name)
        except Exception as e:
            raise RuntimeError(
                f"Não consegui extrair a DATA INICIAL do OGMO {self._numero_ogmo(p_ini.name)} ({p_ini.name}). Erro: {e}"
            ) from e

        try:
            _, df = self.extrair_periodo_por_data(p_fim.name)
        except Exception as e:
            raise RuntimeError(
                f"Não consegui extrair a DATA FINAL do OGMO {self._numero_ogmo(p_fim.name)} ({p_fim.name}). Erro: {e}"
            ) from e

        print(f"✔ Data inicial de: {p_ini.name} -> {di}")
        print(f"✔ Data final de:   {p_fim.name} -> {df}")

        return di, df






    # ==================================================
    # EXTRAÇÃO: LAYOUT SS (WILSON SS / SEA SIDE PSS)
    # ==================================================
    def _somar_valor_item(self, regex_nome: str, paginas_validas: set[int] | None = None, pick: str = "last") -> float:
        total = 0.0

        # ✅ BR ou US "limpo", e evita pegar pedaços quando tem "1.229.35"
        padrao_valor = r"\d{1,3}(?:\.\d{3})*,\d{2}|\b\d+\.\d{2}\b(?!\.)"

        for item in self.paginas_texto:
            if paginas_validas is not None and item.get("page") not in paginas_validas:
                continue

            linhas = item["texto"].splitlines()

            for i, linha in enumerate(linhas):
                if re.search(regex_nome, linha, re.IGNORECASE):
                    vals = re.findall(padrao_valor, linha)
                    if vals:
                        escolhido = vals[0] if pick == "first" else vals[-1]
                        total += self._br_or_us_to_float(escolhido)
                        continue

                    if i + 1 < len(linhas):
                        prox = linhas[i + 1]
                        vals = re.findall(padrao_valor, prox)
                        if vals:
                            escolhido = vals[0] if pick == "first" else vals[-1]
                            total += self._br_or_us_to_float(escolhido)

        return total

    def _debug_match_valores(self, regex_nome: str, paginas_validas=None):
        padrao_valor = r"\d{1,3}(?:\.\d{3})*,\d{2}|\b\d+\.\d{2}\b(?!\.)"
        print(f"\n=== DEBUG MATCHES: {regex_nome} ===")
        for it in self.paginas_texto:
            if paginas_validas is not None and it.get("page") not in paginas_validas:
                continue
            for linha in it["texto"].splitlines():
                if re.search(regex_nome, linha, re.IGNORECASE):
                    vals = re.findall(padrao_valor, linha)
                    print(f"[{it['pdf']} pág {it['page']}] {linha}")
                    print(f"   -> valores: {vals}")
        print("=== FIM DEBUG ===\n")



    def _br_or_us_to_float(self, valor) -> float:
        if valor in (None, "", "NÃO ENCONTRADO"):
            return 0.0
        if isinstance(valor, (int, float)):
            return float(valor)

        s = str(valor).strip()

        # remove espaços dentro do número: "742 266.46" -> "742266.46"
        s = re.sub(r"(?<=\d)\s+(?=\d)", "", s)

        # pt-BR: 1.234,56
        if re.match(r"^\d{1,3}([.\s]\d{3})*,\d{2}$", s):
            s = s.replace(" ", "").replace(".", "").replace(",", ".")
            return float(s)

        # US com milhar: 1,234.56
        if re.match(r"^\d{1,3}([,\s]\d{3})*\.\d{2}$", s):
            s = s.replace(" ", "").replace(",", "")
            return float(s)

        # simples "1234,56"
        if re.match(r"^\d+,\d{2}$", s):
            return float(s.replace(",", "."))

        # simples "1234.56"
        if re.match(r"^\d+\.\d{2}$", s):
            return float(s)

        # fallback: tenta limpar tudo menos dígito , .
        s2 = re.sub(r"[^0-9.,]", "", s)
        if "," in s2 and "." in s2:
            # assume pt-BR (.) milhar e (, ) decimal
            s2 = s2.replace(".", "").replace(",", ".")
        elif "," in s2:
            s2 = s2.replace(",", ".")
        return float(s2)



    def _somar_rat_ajustado(self, paginas_validas: set[int] | None = None, lookahead: int = 6) -> float:
        """
        Pega o VALOR do INSS (RAT Ajustado) ignorando percentual (1,5000%)
        e sem cair no INSS (Terceiros/Previdência).
        Aceita BR (53,24) e US (53.24).
        """
        total = 0.0

        padrao_br = r"\d{1,3}(?:\.\d{3})*,\d{2}(?!\d)"
        padrao_us = r"\d+\.\d{2}(?!\d)"

        for item in self.paginas_texto:
            if paginas_validas is not None and item.get("page") not in paginas_validas:
                continue

            linhas = item["texto"].splitlines()

            for i, linha in enumerate(linhas):
                if re.search(r"INSS\s*\(\s*RAT", linha, re.IGNORECASE):

                    trecho = [linha]
                    for j in range(i + 1, min(len(linhas), i + 1 + lookahead)):
                        ln = linhas[j]

                        # para no próximo INSS que não seja RAT (pra não cair no Terceiros)
                        if re.search(r"INSS\s*\(", ln, re.IGNORECASE) and not re.search(r"INSS\s*\(\s*RAT", ln, re.IGNORECASE):
                            break

                        trecho.append(ln)

                    bloco = " ".join(trecho)

                    # remove percentuais tipo 1,5000%
                    bloco = re.sub(r"\d+(?:[.,]\d+)?\s*%", " ", bloco)

                    # 1) tenta BR
                    vals = re.findall(padrao_br, bloco)
                    if vals:
                        total += self._br_or_us_to_float(vals[-1])
                        continue

                    # 2) tenta US
                    vals = re.findall(padrao_us, bloco)
                    if vals:
                        total += self._br_or_us_to_float(vals[-1])
                        continue

        return total


    def _valor_apos_rs(self, linha: str) -> float | None:
        # pega números logo depois de "R$"
        m = re.search(r"R\$\s*([0-9][0-9\.\,\s]*[0-9][\.,][0-9]{2})", linha, re.IGNORECASE)
        if not m:
            return None
        return self._br_or_us_to_float(m.group(1))
    

    def _somar_seguranca_trabalhador_avulso(self, paginas_validas: set[int] | None = None) -> float:
        total = 0.0

        # dinheiro BR ou US
        padrao_valor = r"\d{1,3}(?:\.\d{3})*,\d{2}|\b\d+\.\d{2}\b(?!\.)"

        for item in self.paginas_texto:
            if paginas_validas is not None and item.get("page") not in paginas_validas:
                continue

            for linha in item["texto"].splitlines():
                if re.search(r"Seguran[cç]a\s+do\s+Trabalhador\s+Portu[aá]rio\s+Avulso", linha, re.IGNORECASE):
                    # ✅ pega só valores monetários e usa o ÚLTIMO (que é o valor)
                    vals = re.findall(padrao_valor, linha)
                    if vals:
                        total += self._br_or_us_to_float(vals[-1])

        return total



    def _pegar_valor_monetario_da_linha(self, linha: str) -> float | None:
        ln = str(linha)

        # BR: 3.483,17 ou 3483,17 ou 3 483,17
        br = re.findall(r"\d{1,3}(?:[.\s]\d{3})*,\d{2}", ln)

        # US: 1,229.35 ou 1229.35 ou 1 229.35
        us = re.findall(r"\d{1,3}(?:[,\s]\d{3})*\.\d{2}", ln)

        # junta e pega o último valor monetário real da linha
        vals = br + us
        if not vals:
            return None

        return self._money_to_float(vals[-1])



    def _texto_pdf_pagina(self, pdf_nome: str, page_num: int) -> str:
        blocos = [it["texto"] for it in self.paginas_texto
                if it.get("pdf") == pdf_nome and it.get("page") == page_num]
        return "\n".join(blocos)
        
    def _somar_rotulo_em_pagina(self, pdf_nome: str, page_num: int, rotulo_regex: str) -> float:
        texto = self._texto_pdf_pagina(pdf_nome, page_num)
        if not texto:
            return 0.0

        total = 0.0
        for ln in texto.splitlines():
            if re.search(rotulo_regex, self._normalizar(ln), re.IGNORECASE):
                v = self._valor_apos_rs(ln)  # ✅ sempre após R$
                if v is not None:
                    total += v
        return total



    def extrair_dados_layout_sea_side_wilson(self):
        print("🔍 Extraindo dados – layout SEA SIDE")

        PAG_FIN = {1}
        PAG_HE  = {2}

        self.dados = {
            "Salário Bruto (MMO)": self._somar_valor_item(r"Sal[aá]rio\s+Bruto\s*\(MMO\)", paginas_validas=PAG_FIN, pick="last"),
            "Vale Refeição": self._somar_valor_item(r"Vale\s+Refei", paginas_validas=PAG_FIN, pick="last"),

            # ✅ NOVO
            "Segurança do Trabalhador Portuário Avulso": self._somar_seguranca_trabalhador_avulso(paginas_validas=PAG_FIN),

            "Encargos Administrativos": self._somar_encargos_adm(paginas_validas=PAG_FIN),
            "INSS (RAT Ajustado)": self._somar_rat_ajustado(paginas_validas=PAG_FIN, lookahead=8),
            "Taxas Bancárias": self._somar_valor_item(r"Taxas\s+Banc", paginas_validas=PAG_FIN, pick="last"),
            "Horas Extras": self._somar_valor_item(r"Horas?\s+Extras?", paginas_validas=PAG_HE, pick="last"),
        }

        for k, v in self.dados.items():
            print(f"✔ {k}: {float(v or 0.0):.2f}")






    def _somar_ultimo_valor_por_linha_por_pdf(self, regex_nome: str, paginas_validas: set[int] | None = None) -> dict[str, float]:
        totais = {}
        for item in self.paginas_texto:
            if paginas_validas is not None and item.get("page") not in paginas_validas:
                continue

            pdf = item.get("pdf", "DESCONHECIDO")
            linhas = item["texto"].splitlines()

            for linha in linhas:
                if re.search(regex_nome, linha, re.IGNORECASE):
                    # pega BR e US e também casos com espaço no milhar
                    valores = re.findall(r"\d[\d\.\s]*,\d{2}|\d[\d\.\s]*\.\d{2}", linha)
                    if valores:
                        val = self._br_or_us_to_float(valores[-1].replace(" ", ""))
                        totais[pdf] = totais.get(pdf, 0.0) + val
        return totais


    def _somar_encargos_adm(self, paginas_validas: set[int] | None = None) -> float:
        total = 0.0

        # dinheiro BR ou US
        padrao_valor = r"\d{1,3}(?:\.\d{3})*,\d{2}|\b\d+\.\d{2}\b(?!\.)"

        for item in self.paginas_texto:
            if paginas_validas is not None and item.get("page") not in paginas_validas:
                continue

            for linha in item["texto"].splitlines():
                if re.search(r"Encargos\s+Administrativos?", linha, re.IGNORECASE):
                    # ✅ remove o bloco "TPAS 5,28155" ou "TPAS 5.91828"
                    linha_limpa = re.sub(r"\bTPAS\b\s*\d+(?:[.,]\d+)?", " ", linha, flags=re.IGNORECASE)

                    vals = re.findall(padrao_valor, linha_limpa)
                    if vals:
                        # ✅ aqui queremos o valor final da linha (ex: 68,66 / 23,67)
                        total += self._br_or_us_to_float(vals[-1])

        return total



    def _somar_ultimo_valor_por_linha(self, regex_nome: str, paginas_validas: set[int] | None = None) -> float:
        total = 0.0

        # valor BR ou US, aceitando espaços
        padrao_valor = r"\d{1,3}(?:\.\d{3})*,\d{2}|\b\d+\.\d{2}\b(?!\.)"


        # compila regex uma vez
        rx = re.compile(regex_nome, re.IGNORECASE)

        for item in self.paginas_texto:
            if paginas_validas is not None and item.get("page") not in paginas_validas:
                continue

            for linha in item["texto"].splitlines():
                ln = self._normalizar(linha)  # <<< AQUI É O PULO DO GATO
                if rx.search(ln):
                    vals = re.findall(padrao_valor, linha)  # pega do original pra manter número certo
                    if vals:
                        s = vals[-1].replace(" ", "")
                        total += self._br_or_us_to_float(s)

        return total



    def _somar_valor_apos_rotulo(self, regex_nome: str, paginas_validas: set[int] | None = None, lookahead: int = 12) -> float:
        """
        Acha o rótulo e busca o primeiro valor numérico nas próximas N linhas.
        Resolve:
        - valores em outra linha (Taxas Bancárias)
        - tabelas onde os rótulos vem e os números aparecem abaixo (Horas Extras)
        - número BR e US
        """
        total = 0.0
        padrao_valor = r"\d{1,3}(?:\.\d{3})*,\d{2}|\d+\.\d{2}"

        for item in self.paginas_texto:
            if paginas_validas is not None and item.get("page") not in paginas_validas:
                continue

            linhas = item["texto"].splitlines()

            for i, linha in enumerate(linhas):
                if re.search(regex_nome, linha, re.IGNORECASE):

                    # procura valor na mesma linha + próximas linhas
                    fim = min(len(linhas), i + 1 + lookahead)
                    bloco = " ".join(linhas[i:fim])

                    vals = re.findall(padrao_valor, bloco)
                    if vals:
                        total += self._br_or_us_to_float(vals[0])  # primeiro valor após o rótulo
        return total


    def _numero_ogmo(self, nome: str) -> int | None:
        """
        Extrai o número do OGMO do nome do arquivo.
        Aceita:
        - 'FOLHAS OGMO 1.pdf'
        - 'FOLHAS OGMO (2).pdf'
        - 'OGMO 3.pdf'
        """
        m = re.search(r"\bOGMO\s*\(?\s*(\d+)\s*\)?\b", nome, re.IGNORECASE)
        if m:
            return int(m.group(1))

        # fallback: tenta achar "(n)" no final
        m = re.search(r"\(\s*(\d+)\s*\)", nome)
        if m:
            return int(m.group(1))

        return None



    def _achar_pdf_menor_numero(self) -> Path | None:
        candidatos = []
        for p in self.caminhos_pdfs:
            n = self._numero_ogmo(p.name)
            if n is not None:
                candidatos.append((n, p))
        if not candidatos:
            return None
        return min(candidatos, key=lambda x: x[0])[1]


    def _achar_pdf_maior_numero(self) -> Path | None:
        candidatos = []
        for p in self.caminhos_pdfs:
            n = self._numero_ogmo(p.name)
            if n is not None:
                candidatos.append((n, p))
        if not candidatos:
            return None
        return max(candidatos, key=lambda x: x[0])[1]



    # ==================================================
    # REPORT VIGIA - LAYOUT SS (Wilson SS / Sea Side PSS)
    # ==================================================



    def colar_report_layout_ss(self, wb):
        aba = next(s for s in wb.sheets if s.name.strip().lower() == "report vigia")
        print("📌 Report (layout SS) – colando valores fixos")

        MAPA_FIXO = {
            "Salário Bruto (MMO)": "G22",
            "Vale Refeição": "G25",
            "Segurança do Trabalhador Portuário Avulso": "G26",
            "Encargos Administrativos": "G27",


            "INSS (RAT Ajustado)": "G30",  

            "Taxas Bancárias": "G32",
            "Horas Extras": "G35",
        }

        for chave, celula in MAPA_FIXO.items():
            aba.range(celula).value = float(self.dados.get(chave, 0.0) or 0.0)


    def _garantir_linhas_report(self, aba, linha_base: int, total_linhas: int):
        """
        Garante que existam `total_linhas` linhas disponíveis a partir de `linha_base`,
        inserindo linhas abaixo e herdando formatação da linha de cima (sem Copy/PasteSpecial).

        Isso evita:
        - erro PasteSpecial
        - conflito com clipboard
        - bug com células mescladas
        """
        if total_linhas <= 1:
            return

        # Constantes do Excel
        xlShiftDown = -4121
        xlFormatFromLeftOrAbove = 0

        # Precisamos criar (total_linhas - 1) linhas abaixo da base
        qtd_inserir = total_linhas - 1

        # Insere em bloco (mais rápido e mais estável)
        # Ex: base=22, inserir 5 => insere linhas 23..27
        r = aba.api.Rows(linha_base + 1)
        for _ in range(qtd_inserir):
            r.Insert(Shift=xlShiftDown, CopyOrigin=xlFormatFromLeftOrAbove)



    # ==================================================
    # CONFIGURAÇÃO DE MODELO POR CLIENTE
    # ==================================================
    def obter_configuracao_cliente(self, cliente: str, porto: str) -> dict:
        """
        ✅ Aqui fica o coração do “qual modelo usar” e “qual colagem fazer”.
        Você falou:
        - Sea Side tem modelo DIFERENTE de Wilson
        - mas o REPORT (células) é o mesmo modo.
        """
        if self._usa_layout_ss(cliente, porto):
            if cliente == "WILSON SONS":
                modelo = "WILSON SONS - SÃO SEBASTIÃO.xlsx"
            elif cliente == "SEA SIDE":
                modelo = "SEA SIDE - PSS.xlsx"
            else:
                modelo = f"{cliente} - {porto}.xlsx"

            return {
                "modelo": modelo,
                "colar_report": self.colar_report_layout_ss
            }

        # Padrão (Aquarius e outros clientes São Sebastião)
        return {
            "modelo": f"{cliente}.xlsx",
            "colar_report": self.colar_report_padrao
        }


    def _escolher_pdf_inicio_fim(self) -> tuple[str, str]:
        """
        Decide qual PDF é o início e qual é o fim.
        - tenta identificar OGMO 1 e OGMO 2 pelo nome
        - fallback: primeiro selecionado = início, último = fim
        Retorna (nome_pdf_inicio, nome_pdf_fim)
        """
        nomes = [p.name for p in self.caminhos_pdfs]

        if len(nomes) == 1:
            return nomes[0], nomes[0]

        # tenta achar "1" e "2" pelo nome do arquivo
        n1 = next((n for n in nomes if re.search(r"(ogmo\s*1|folhas\s*ogmo\s*1|\b1\b)", n, re.I)), None)
        n2 = next((n for n in nomes if re.search(r"(ogmo\s*2|folhas\s*ogmo\s*2|\b2\b)", n, re.I)), None)

        if n1 and n2:
            return n1, n2

        return nomes[0], nomes[-1]



    # ==================================================
    # FRONT VIGIA
    # ==================================================
    def preencher_front_vigia(self, wb):
        try:
            aba = next(s for s in wb.sheets if s.name.strip().lower() == "front vigia")

            pasta = self.caminhos_pdfs[0].parent
            navio = obter_nome_navio(pasta, None)
            nd = obter_dn_da_pasta(pasta)
            
            # ✅ aqui é o pulo do gato
            if len(self.caminhos_pdfs) >= 2:
                data_ini, data_fim = self.extrair_datas_mescladas()
            else:
                data_ini, data_fim = self.extrair_periodo_por_data()


            def fmt(data_str: str) -> str:
                d = datetime.strptime(data_str, "%d/%m/%Y")
                return f"{calendar.month_name[d.month]} {d.day}, {d.year}"

            aba.range("D15").merge_area.value = navio
            aba.range("D16").merge_area.value = fmt(data_ini)
            aba.range("D17").merge_area.value = fmt(data_fim)

            ano = datetime.now().year % 100
            aba.range("C21").merge_area.value = f"DN {nd}/{ano:02d}"

            hoje = datetime.now()
            meses = ["", "janeiro", "fevereiro", "março", "abril", "maio", "junho",
                    "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"]
            aba.range("C39").merge_area.value = f"  Santos, {hoje.day} de {meses[hoje.month]} de {hoje.year}"

            print("✅ FRONT VIGIA preenchido")

        except StopIteration:
            print("⚠️ Aba FRONT VIGIA não encontrada")

    # ==================================================
    # CREDIT NOTE
    # ==================================================
    def escrever_cn_credit_note(self, wb, nd: str):
        ws_credit = None
        for sheet in wb.sheets:
            if sheet.name.strip().lower() == "credit note":
                ws_credit = sheet
                break

        if ws_credit is None:
            print("ℹ️ Aba Credit Note não existe — seguindo fluxo.")
            return

        ano = datetime.now().year % 100
        ws_credit.range("C21").merge_area.value = f"CN {nd}/{ano:02d}"
        print("✅ Credit Note preenchida (C21)")

    # ==================================================
    # REPORT VIGIA - PADRÃO (Aquarius e outros)
    # ==================================================

    def _tarifa_por_status(self, ws_report, d: date, periodo: str, status: str) -> float:
        dom_fer = self._is_domingo_ou_feriado(d)
        noite = self._is_noite_por_periodo(periodo)

        # ✅ ATRACADO usa linha 9, FUNDEIO usa linha 16
        linha_ref = {
            "ATRACADO": 9,
            "FUNDEIO": 16,
        }.get(status)

        if linha_ref is None:
            return 0.0  # AO_LARGO ou desconhecido -> por enquanto não calcula

        # escolhe coluna base
        if not dom_fer and not noite:
            col = "N"
        elif not dom_fer and noite:
            col = "O"
        elif dom_fer and not noite:
            col = "P"
        else:
            col = "Q"

        cell = f"{col}{linha_ref}"
        val = ws_report.range(cell).value
        return float(val or 0.0)



    def preencher_tarifa_por_linha(self, ws_report, linha_base: int, n: int, status: str, coluna_saida: str = "G"):
        """
        Lê data em C{linha} e período em E{linha}.
        Se status == ATRACADO ou FUNDEIO: escreve tarifa na coluna_saida.
        """
        if status not in ("ATRACADO", "FUNDEIO"):
            return

        for i in range(n):
            linha = linha_base + i
            d = ws_report.range(f"C{linha}").value
            p = ws_report.range(f"E{linha}").value

            if isinstance(d, datetime):
                d = d.date()
            if not isinstance(d, date):
                continue

            tarifa = self._tarifa_por_status(ws_report, d, str(p or ""), status=status)
            ws_report.range(f"{coluna_saida}{linha}").value = tarifa


    def gerar_horarios(self, periodo_inicial: str, periodo_final: str) -> list[str]:
        """
        Gera sequência entre início e fim, respeitando final diferente.
        """
        seq = ["01x07", "07x13", "13x19", "19x01"]
        if periodo_inicial not in seq or periodo_final not in seq:
            # fallback: devolve só inicial se algo vier fora do padrão
            return [periodo_inicial]

        horarios = []
        idx = seq.index(periodo_inicial)

        while True:
            atual = seq[idx]
            horarios.append(atual)
            if atual == periodo_final:
                break
            idx = (idx + 1) % len(seq)

        return horarios

    def preencher_coluna_horarios(self, ws_report, horarios_ogmo: list[str], linha_inicial: int = 22):
        for i, horario in enumerate(horarios_ogmo):
            ws_report.range(f"E{linha_inicial + i}").value = horario


    # ==================================================
    # REPORT VIGIA - PADRÃO (Aquarius e outros)
    # ==================================================
    def colar_report_padrao(self, wb):
        aba = self._achar_aba(wb, ["report vigia"])
        print("📌 Report PADRÃO – Outros Clientes")

        if len(self.caminhos_pdfs) >= 2:
            data_ini, data_fim, periodo_inicial, periodo_final = self.extrair_periodo_mesclado_n()
        else:
            data_ini, data_fim = self.extrair_periodo_por_data()
            periodo_inicial, periodo_final = self.extrair_periodo_por_horario()


        print("DEBUG extração:",
                "data_ini=", data_ini,
                "data_fim=", data_fim,
                "p_ini=", periodo_inicial,
                "p_fim=", periodo_final)




        periodos_com_data = self.gerar_periodos_report_padrao_ssz_por_dia(
            data_ini=data_ini,
            data_fim=data_fim,
            periodo_inicial=periodo_inicial,
            periodo_final=periodo_final,
        )

        linha_base = 22
        n = len(periodos_com_data)

        self._garantir_linhas_report(aba, linha_base, n)

        for i, (d, p) in enumerate(periodos_com_data):
            linha = linha_base + i
            aba.range(f"C{linha}").value = self._fmt_data_excel(d)
            aba.range(f"E{linha}").value = p

        # ✅ status pelo nome do navio (o "nome" com (ATRACADO)/(AO LARGO))
        pasta = self.caminhos_pdfs[0].parent
        navio = obter_nome_navio(pasta, None)  # você já tem
        status = self._status_atracacao(navio)

        # ✅ preenche tarifa por linha usando C e E como base
        self.preencher_tarifa_por_linha(aba, linha_base, n, status=status, coluna_saida="G")

        print(f"✔ Colado {n} períodos + tarifa (status={status}) a partir de C{linha_base}/E{linha_base}")


    def gerar_periodos_report_padrao_ssz_por_dia(self, data_ini, data_fim, periodo_inicial, periodo_final):
        ordem = ["07x13", "13x19", "19x01", "01x07"]

        def norm_periodo(p: str) -> str:
            p = (p or "").strip().lower().replace(" ", "")
            p = p.replace("h", "")
            p = p.replace("-", "x").replace("×", "x")
            p = p.replace(".", "")
            try:
                a, b = p.split("x")
                return f"{int(a):02d}x{int(b):02d}"
            except Exception:
                return (p or "").upper()

        def to_date(d):
            if isinstance(d, datetime):
                return d.date()
            if isinstance(d, date):
                return d
            s = str(d).strip()
            for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
                try:
                    return datetime.strptime(s, fmt).date()
                except Exception:
                    pass
            raise ValueError(f"Data inválida: {d!r}")

        def seq_entre(inicio: str, fim: str) -> list[str]:
            i = ordem.index(inicio)
            out = []
            while True:
                out.append(ordem[i])
                if ordem[i] == fim:
                    break
                i = (i + 1) % 4
                if len(out) > 4:
                    break
            return out

        p_ini = norm_periodo(periodo_inicial)
        p_fim = norm_periodo(periodo_final)

        if p_ini not in ordem:
            raise ValueError(f"Período inicial inválido: {periodo_inicial!r} -> {p_ini!r}")
        if p_fim not in ordem:
            raise ValueError(f"Período final inválido: {periodo_final!r} -> {p_fim!r}")

        d_ini = to_date(data_ini)
        d_fim = to_date(data_fim)
        if d_fim < d_ini:
            raise ValueError(f"Data final menor que inicial: {d_ini} > {d_fim}")

        out = []
        dia = d_ini

        while dia <= d_fim:
            # Mantém sua regra: em dias “do meio”, começa sempre em 07x13
            inicio = p_ini if dia == d_ini else "07x13"

            # No último dia, termina no período final; caso contrário, vai até 01x07
            fim = p_fim if dia == d_fim else "01x07"

            for p in seq_entre(inicio, fim):
                out.append((dia, p))  # mantém 01x07 no mesmo dia (como você já faz)

            dia += timedelta(days=1)

            if len(out) > 400:
                raise RuntimeError("Proteção: períodos demais gerados. Verifique datas/períodos extraídos.")

        return out

        


    def _fmt_data_excel(self, d):
        if isinstance(d, datetime):
            return d.date()
        if isinstance(d, date):
            return d
        raise ValueError(f"Data inválida para Excel: {d!r}")


    def extrair_periodo_mesclado_n(self) -> tuple[str, str, str, str]:
        """
        Retorna (data_ini, data_fim, periodo_ini, periodo_fim)
        usando:
        - OGMO menor número = inicio
        - OGMO maior número = fim
        Funciona com 1 ou N PDFs.
        """
        pdfs = self._ordenar_pdfs_ogmo()
        if not pdfs:
            raise RuntimeError("Nenhum PDF selecionado.")

        p_ini = self._achar_pdf_menor_numero() or pdfs[0]
        p_fim = self._achar_pdf_maior_numero() or pdfs[-1]

        try:
            di, _ = self.extrair_periodo_por_data(p_ini.name)
        except Exception as e:
            raise RuntimeError(
                f"Não consegui extrair DATA INICIAL do OGMO {self._numero_ogmo(p_ini.name)} ({p_ini.name}). Erro: {e}"
            ) from e

        try:
            _, df = self.extrair_periodo_por_data(p_fim.name)
        except Exception as e:
            raise RuntimeError(
                f"Não consegui extrair DATA FINAL do OGMO {self._numero_ogmo(p_fim.name)} ({p_fim.name}). Erro: {e}"
            ) from e

        try:
            pi, _ = self.extrair_periodo_por_horario(p_ini.name)
        except Exception as e:
            raise RuntimeError(
                f"Não consegui extrair PERÍODO INICIAL do OGMO {self._numero_ogmo(p_ini.name)} ({p_ini.name}). Erro: {e}"
            ) from e

        try:
            _, pf = self.extrair_periodo_por_horario(p_fim.name)
        except Exception as e:
            raise RuntimeError(
                f"Não consegui extrair PERÍODO FINAL do OGMO {self._numero_ogmo(p_fim.name)} ({p_fim.name}). Erro: {e}"
            ) from e

        print(f"✔ Data inicial de: {p_ini.name} -> {di} ({pi})")
        print(f"✔ Data final de:   {p_fim.name} -> {df} ({pf})")

        return di, df, pi, pf



    # --------------------------------------------------
    # 1) status ATRACADO / AO LARGO pelo nome
    # --------------------------------------------------
    def _status_atracacao(self, nome: str) -> str | None:
        if not nome:
            return None

        s = str(nome).upper()

        # se tiver parênteses, pega dentro; se não, usa tudo
        m = re.search(r"\((.*?)\)", s)
        dentro = m.group(1).strip() if m else s

        dentro = dentro.replace("-", " ").replace("_", " ")
        dentro = re.sub(r"\s+", " ", dentro)

        if "ATRAC" in dentro:
            return "ATRACADO"
        if "FUNDE" in dentro:   # ✅ FUNDEIO
            return "FUNDEIO"
        if "AO LARGO" in dentro or "A LARGO" in dentro or "LARGO" in dentro:
            return "AO_LARGO"

        return None

    # --------------------------------------------------
    # 2) dia/noite pelo período OGMO (coluna E)
    # --------------------------------------------------
    def _is_noite_por_periodo(self, periodo: str) -> bool:
        p = (periodo or "").strip().upper().replace(" ", "")
        # noite: 19x01 e 01x07
        return p in ("19X01", "01X07", "19x01", "01x07")


    # --------------------------------------------------
    # 3) domingo/feriado (mínimo viável)
    #    (se você já tiver função de feriado no projeto, plugue aqui)
    # --------------------------------------------------
    def _is_domingo_ou_feriado(self, d: date) -> bool:
        if isinstance(d, datetime):
            d = d.date()
        # domingo
        if d.weekday() == 6:
            return True

        # ✅ feriados nacionais fixos (mínimo)
        fixos = {
            (1, 1),    # Confraternização Universal
            (4, 21),   # Tiradentes
            (5, 1),    # Dia do Trabalho
            (9, 7),    # Independência
            (10, 12),  # Nossa Sra Aparecida
            (11, 2),   # Finados
            (11, 15),  # Proclamação da República
            (12, 25),  # Natal
        }
        if (d.month, d.day) in fixos:
            return True

        # Se você quiser incluir feriados móveis (Carnaval/Paixão/Corpus Christi),
        # eu adiciono um cálculo de Páscoa e derivados aqui.
        return False


    # --------------------------------------------------
    # 4) pega a tarifa ATRACADO pela regra:
    #    - Seg-Sáb dia:   N9
    #    - Seg-Sáb noite: O9
    #    - Dom/Feriado dia:   P9
    #    - Dom/Feriado noite: Q9
    # --------------------------------------------------
    def _tarifa_atracado(self, ws_report, d: date, periodo: str) -> float:
        dom_fer = self._is_domingo_ou_feriado(d)
        noite = self._is_noite_por_periodo(periodo)

        if not dom_fer and not noite:
            cell = "N9"  # seg-sab dia
        elif not dom_fer and noite:
            cell = "O9"  # seg-sab noite
        elif dom_fer and not noite:
            cell = "P9"  # dom/fer dia
        else:
            cell = "Q9"  # dom/fer noite

        val = ws_report.range(cell).value
        return float(val or 0.0)


    # --------------------------------------------------
    # 5) aplica tarifa linha a linha (baseado em C=data e E=periodo)
    # --------------------------------------------------
    def preencher_tarifa_por_linha(self, ws_report, linha_base: int, n: int, status: str, coluna_saida: str = "G"):
        if status not in ("ATRACADO", "FUNDEIO"):
            return

        # ATRACADO usa linha 9, FUNDEIO usa linha 16
        linha_ref = 9 if status == "ATRACADO" else 16

        for i in range(n):
            linha = linha_base + i
            d = ws_report.range(f"C{linha}").value
            p = ws_report.range(f"E{linha}").value

            if isinstance(d, datetime):
                d = d.date()
            if not isinstance(d, date):
                continue

            dom_fer = self._is_domingo_ou_feriado(d)
            noite = self._is_noite_por_periodo(str(p or ""))

            if not dom_fer and not noite:
                cell = f"N{linha_ref}"
            elif not dom_fer and noite:
                cell = f"O{linha_ref}"
            elif dom_fer and not noite:
                cell = f"P{linha_ref}"
            else:
                cell = f"Q{linha_ref}"

            val = ws_report.range(cell).value
            ws_report.range(f"{coluna_saida}{linha}").value = float(val or 0.0)


        print("DEBUG status:", status)

    # ==================================================
    # EXECUÇÃO PRINCIPAL
    # ==================================================

    def executar(self):
        self.selecionar_pdfs_ogmo()
        self.carregar_pdfs()   # já faz pdfplumber e OCR só se precisar
        self.normalizar_texto_mantendo_linhas()




        cliente, porto = self.identificar_cliente_e_porto()
        print(f"\n🚢 FATURAMENTO OGMO – {cliente} / {porto}")

        if cliente == "WILSON SONS":
            self.extrair_dados_layout_sea_side_wilson()
        elif cliente == "SEA SIDE":
            self.extrair_dados_layout_sea_side_wilson()

        else:
            self.dados = {}

        config = self.obter_configuracao_cliente(cliente, porto)

        modelo = obter_pasta_faturamentos() / config["modelo"]
        caminho_local = copiar_para_temp_xlwings(modelo)

        app = xw.App(visible=False, add_book=False)
        wb = app.books.open(str(caminho_local))


        try:
            pasta = self.caminhos_pdfs[0].parent
            navio = obter_nome_navio(pasta, None)
            nd = obter_dn_da_pasta(pasta)

            # FRONT
            self.preencher_front_vigia(wb)

            # CREDIT NOTE
            self.escrever_cn_credit_note(wb, nd)

            # REPORT
            colar_report = config.get("colar_report")
            if colar_report:
                colar_report(wb)

            # NF
            escrever_nf_faturamento_completo(wb, navio, nd)

            # ✅ SALVAR EXCEL (com wb aberto)
            nome_base = f"FATURAMENTO - ND {nd} - MV {navio}"
            caminho_excel = salvar_excel_com_nome(wb, pasta, nome_base)
            print(f"💾 Excel salvo em: {caminho_excel}")

            # ✅ GERAR PDF SEM REABRIR O EXCEL (evita erro COM)
            gerar_pdf_do_wb_aberto(wb, pasta, nome_base, ignorar_abas=("NF",))



            print("✅ FATURAMENTO FINALIZADO")

        finally:
            wb.close()
            app.quit()


# ==================================================
# ✅ PLACEHOLDERS (você DISSE que já tem no seu projeto)
# Se você já tiver, APAGA essa parte daqui.



# ==============================
# CLASSE 6: GERAR RELATÓRIO - X
# ==============================


class GerarRelatorio:
    pass


# ==============================
# MENU PRINCIPAL
# ==============================


class CentralSanport:
    def __init__(self):
        self.opcoes = [
            "FATURAMENTO",
            "FATURAMENTO SÃO SEBASTIÃO",
            "DE ACORDO",
            "FAZER PONTO",
            "DESFAZER PONTO - X",
            "RELATÓRIO - X",
            "SAIR DO PROGRAMA"
        ]

        # 🔹 instâncias (recomendo instanciar sob demanda p/ não carregar Excel antes)
        self.de_acordo = FaturamentoDeAcordo()
        self.relatorio = GerarRelatorio()

    # =========================
    # UTILITÁRIOS
    # =========================
    def limpar_tela(self):
        os.system("cls" if os.name == "nt" else "clear")

    def limpar_buffer_teclado(self):
        while msvcrt.kbhit():
            msvcrt.getch()

    def pausar_e_voltar(self, selecionado):
        print("\n🔁 Pressione ENTER para voltar ao menu...")
        while True:
            key = msvcrt.getch()
            if key in (b"\r", b"\n"):
                self.limpar_buffer_teclado()
                self.mostrar_menu(selecionado)
                return

    # =========================
    # MENU PRINCIPAL
    # =========================
    def mostrar_menu(self, selecionado):
        self.limpar_tela()

        print("╔" + "═" * 62 + "╗")
        print(f"║{' 🚢 CENTRAL DE PROCESSOS - SANPORT 🚢 '.center(60)}║")
        print("╚" + "═" * 62 + "╝\n")

        for i, opcao in enumerate(self.opcoes):
            if i == selecionado:
                print(f"          ►► {opcao} ◄◄")
            else:
                print(f"              {opcao}")

        print("\n" + "═" * 64)
        print("   ↑ ↓ = Navegar     ENTER = Selecionar")
        print("═" * 64)

    # =========================
    # SUBMENU FATURAMENTO
    # =========================
    def menu_faturamento(self):
        opcoes = [
            "Faturamento (Normal)",
            "Faturamento Atípico",
            "Voltar"
        ]
        selecionado = 0

        while True:
            self.limpar_tela()
            print("╔" + "═" * 62 + "╗")
            print(f"║{' 💰 MENU FATURAMENTO 💰 '.center(60)}║")
            print("╚" + "═" * 62 + "╝\n")

            for i, opcao in enumerate(opcoes):
                if i == selecionado:
                    print(f"          ►► {opcao} ◄◄")
                else:
                    print(f"              {opcao}")

            print("\n" + "═" * 64)
            print("   ↑ ↓ = Navegar     ENTER = Selecionar")
            print("═" * 64)

            key = msvcrt.getch()

            # setas
            if key in (b"\xe0", b"\x00"):
                key = msvcrt.getch()
                if key == b"H":
                    selecionado = max(0, selecionado - 1)
                elif key == b"P":
                    selecionado = min(len(opcoes) - 1, selecionado + 1)
                continue

            # enter
            if key in (b"\r", b"\n"):
                self.limpar_tela()

                # NORMAL
                if selecionado == 0:
                    print("╔" + "═" * 62 + "╗")
                    print("║" + " INICIANDO FATURAMENTO (NORMAL)... ".center(60) + "║")
                    print("╚" + "═" * 62 + "╝\n")

                    try:
                        FaturamentoCompleto().executar()
                    except Exception as e:
                        print(f"\n❌ ERRO NO FATURAMENTO: {e}")

                    print("\n🔁 Pressione ENTER para voltar...")
                    while msvcrt.getch() not in (b"\r", b"\n"):
                        pass

                # ATÍPICO
                elif selecionado == 1:
                    print("╔" + "═" * 62 + "╗")
                    print("║" + " INICIANDO FATURAMENTO (ATÍPICO)... ".center(60) + "║")
                    print("╚" + "═" * 62 + "╝\n")

                    try:
                        FaturamentoAtipico().executar()
                    except Exception as e:
                        print(f"\n❌ ERRO NO FATURAMENTO ATÍPICO: {e}")

                    print("\n🔁 Pressione ENTER para voltar...")
                    while msvcrt.getch() not in (b"\r", b"\n"):
                        pass

                # VOLTAR
                else:
                    return

    # =========================
    # EXECUÇÃO PRINCIPAL
    # =========================
    def rodar(self):
        selecionado = 0
        self.mostrar_menu(selecionado)

        while True:
            key = msvcrt.getch()

            # SETAS
            if key in (b"\xe0", b"\x00"):
                key = msvcrt.getch()

                if key == b"H":  # ↑
                    selecionado = max(0, selecionado - 1)
                    self.mostrar_menu(selecionado)

                elif key == b"P":  # ↓
                    selecionado = min(len(self.opcoes) - 1, selecionado + 1)
                    self.mostrar_menu(selecionado)

                continue

            # ENTER → EXECUTA A OPÇÃO
            if key in (b"\r", b"\n"):
                self.limpar_tela()

                # ----------------------------
                # FATURAMENTO (SUBMENU)
                # ----------------------------
                if selecionado == 0:
                    self.menu_faturamento()
                    self.mostrar_menu(selecionado)

                # ----------------------------
                # FATURAMENTO SÃO SEBASTIÃO
                # ----------------------------
                elif selecionado == 1:
                    print("╔" + "═" * 62 + "╗")
                    print("║" + " INICIANDO FATURAMENTO SÃO SEBASTIÃO... ".center(60) + "║")
                    print("╚" + "═" * 62 + "╝\n")

                    try:
                        programa = FaturamentoSaoSebastiao()
                        programa.executar()
                    except Exception as e:
                        print(f"\n❌ ERRO NO FATURAMENTO SSZ: {e}")

                    self.pausar_e_voltar(selecionado)

                # ----------------------------
                # DE ACORDO
                # ----------------------------
                elif selecionado == 2:
                    print("╔" + "═" * 62 + "╗")
                    print("║" + " INICIANDO DE ACORDO... ".center(60) + "║")
                    print("╚" + "═" * 62 + "╝\n")

                    try:
                        self.de_acordo.executar()
                    except Exception as e:
                        print(f"\n❌ ERRO: {e}")

                    self.pausar_e_voltar(selecionado)

                # ----------------------------
                # FAZER PONTO
                # ----------------------------
                elif selecionado == 3:
                    programa = ProgramaCopiarPeriodo(debug=True)

                    try:
                        programa.executar()
                    except Exception as e:
                        print(f"\n❌ ERRO NO FAZER PONTO: {e}")

                    self.pausar_e_voltar(selecionado)

                # ----------------------------
                # DESFAZER PONTO
                # ----------------------------
                elif selecionado == 4:
                    programa = ProgramaRemoverPeriodo(debug=True)

                    try:
                        programa.executar()
                    except Exception as e:
                        print(f"\n❌ ERRO NO DESFAZER PONTO: {e}")

                    self.pausar_e_voltar(selecionado)

                # ----------------------------
                # RELATÓRIO
                # ----------------------------
                elif selecionado == 5:
                    print("╔" + "═" * 62 + "╗")
                    print("║" + " INICIANDO RELATÓRIO... ".center(60) + "║")
                    print("╚" + "═" * 62 + "╝\n")

                    try:
                        self.relatorio.executar()
                        print("\n✅ RELATÓRIO GERADO COM SUCESSO")
                    except Exception as e:
                        print(f"\n❌ ERRO NO RELATÓRIO: {e}")

                    self.pausar_e_voltar(selecionado)

                # ----------------------------
                # SAIR
                # ----------------------------
                elif selecionado == 6:
                    self.limpar_tela()
                    print("\n👋 Saindo do programa...")
                    break

if __name__ == "__main__":
    CentralSanport().rodar()

