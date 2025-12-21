# 1️⃣ – Importações
import xlwings as xw
from datetime import datetime, timedelta, timezone
from openpyxl.styles import Alignment
import pandas as pd
import os
import sys
import urllib.request
import ssl
import certifi
from pathlib import Path
import tkinter as tk
from tkinter import filedialog
from datetime import date
import re
import locale
from itertools import cycle

# =========================
# Funções utilitárias
# =========================

def abrir_workbooks():
    """
    Abre:
    - 1.xlsx da pasta do NAVIO (selecionada pelo usuário)
    - XLSX do CLIENTE em Desktop/FATURAMENTOS/<CLIENTE>.xlsx
    """

    # --- Seleção da pasta do navio ---
    root = tk.Tk()
    root.withdraw()
    pasta_navio = filedialog.askdirectory(
        title="Selecione a pasta do NAVIO (onde está o 1.xlsx)"
    )

    if not pasta_navio:
        print("Nenhuma pasta selecionada. Encerrando.")
        return None, None, None, None, None

    pasta_navio = Path(pasta_navio)
    pasta_cliente = pasta_navio.parent   # PASTA PAI
    nome_cliente = pasta_cliente.name

    arquivos_1 = list(pasta_navio.glob("1*.xls*"))

    if not arquivos_1:
        raise FileNotFoundError(
            f"Nenhum arquivo começando com '1' encontrado em:\n{pasta_navio}"
        )

    arquivo1 = arquivos_1[0]

    pasta_faturamentos = Path.home() / "Desktop" / "FATURAMENTOS"
    arquivo2 = pasta_faturamentos / f"{nome_cliente}.xlsx"

    app = xw.App(visible=False)
    wb1 = wb2 = None

    try:
        # --- Valida arquivos ---
        if not arquivo1.exists():
            raise FileNotFoundError(f"Arquivo 1.xlsx não encontrado:\n{arquivo1}")

        if not arquivo2.exists():
            raise FileNotFoundError(
                f"Arquivo de faturamento do cliente não encontrado:\n{arquivo2}"
            )

        # --- Abre workbooks ---
        wb1 = app.books.open(arquivo1)
        wb2 = app.books.open(arquivo2)

        ws1 = wb1.sheets[0]

        nomes_abas = [s.name for s in wb2.sheets]

        # Prioridade: aba com nome do cliente
        if nome_cliente in nomes_abas:
            ws_front = wb2.sheets[nome_cliente]

        # Fallback: FRONT VIGIA
        elif "FRONT VIGIA" in nomes_abas:
            ws_front = wb2.sheets["FRONT VIGIA"]

        else:
            raise RuntimeError(
                f"Nenhuma aba válida encontrada no faturamento.\n"
                f"Esperado: '{nome_cliente}' ou 'FRONT VIGIA'."
            )

        return app, wb1, wb2, ws1, ws_front

    except Exception as e:
        print(f"Erro ao abrir os arquivos: {e}")

        if wb1:
            wb1.close()
        if wb2:
            wb2.close()

        app.quit()
        return None, None, None, None, None

    
def fechar_workbooks(app, wb1=None, wb2=None, arquivo_saida=None):
    """
    Salva wb1 normalmente e salva wb2 EXCLUSIVAMENTE como 3.xlsx
    na pasta do cliente (ex: WILSON). Nunca salva em FATURAMENTOS.
    """
    try:
        if wb1:
            wb1.save()
            wb1.close()

        if wb2:
            if not arquivo_saida:
                raise RuntimeError(
                    "Caminho de saída não informado. "
                    "wb2 NÃO será salvo para evitar salvar em FATURAMENTOS."
                )

            wb2.save(arquivo_saida)
            wb2.close()

    finally:
        if app:
            app.quit()

def obter_dn_da_pasta(caminho_arquivo):
    """
    Extrai números do nome da pasta do navio
    Ex: '123 - NAVIO' -> '123'
    """
    pasta = os.path.basename(os.path.dirname(caminho_arquivo))

    numeros = re.findall(r"\d+", pasta)

    if not numeros:
        return None

    return numeros[0]  # primeiro bloco numérico


def data_online():
    context = ssl.create_default_context(cafile=certifi.where())

    req = urllib.request.Request(
        "https://www.cloudflare.com",
        headers={"User-Agent": "Mozilla/5.0"}
    )

    with urllib.request.urlopen(req, context=context, timeout=5) as r:
        data_str = r.headers["Date"]

    # transforma em datetime UTC do servidor
    dt_utc = datetime.strptime(data_str, "%a, %d %b %Y %H:%M:%S %Z").replace(tzinfo=timezone.utc)

    # para uso interno/exibição: converte para hora local
    dt_local = dt_utc.astimezone()  # pega fuso local do sistema
    return dt_utc, dt_local

try:
    hoje_utc, hoje_local = data_online()

    # validação sempre UTC
    data_limite_utc = datetime(hoje_utc.year, hoje_utc.month, 22, tzinfo=timezone.utc)
    if hoje_utc > data_limite_utc:
        sys.exit("⛔ Licença expirada")

    # apenas exibição: sempre local
    print("Data local:", hoje_local.date())


except Exception as e:
    sys.exit(f"Erro ao verificar licença: {e}")

    sys.exit("❌ Não foi possível validar a data online")



def data_por_extenso(valor):
    if isinstance(valor, datetime):
        data = valor

    elif isinstance(valor, date):
        data = datetime(valor.year, valor.month, valor.day)

    elif isinstance(valor, str):
        try:
            data = datetime.strptime(valor, "%d/%m/%Y")
        except:
            return ""  # não inventa data

    else:
        return ""  # nunca usa datetime.now()

    return data.strftime("%d de %B de %Y")

def MMO(arquivo1, wb2):
    ws = wb2.sheets["REPORT VIGIA"]
    if str(ws["E25"].value).strip().upper() != "MMO": return
    df = pd.read_excel(arquivo1, sheet_name="Resumo", header=None)
    col_g = df[6].dropna()
    if col_g.empty: return
    try:
        ultimo_float = locale.atof(str(col_g.iloc[-1]).replace("R$", "").strip())
    except: ultimo_float = float(col_g.iloc[-1])
    ws["F25"].value = ultimo_float
    ws["F25"].number_format = "#.##0,00"


def processar_front(ws1, ws_front):
    """
    Atualiza somente a aba FRONT VIGIA
    """

    # data atual por extenso (rodapé)
    meses = ["", "janeiro", "fevereiro", "março", "abril", "maio", "junho",
             "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"]

    hoje = datetime.now()
    ws_front.range("C39").value = (
        f"Santos, {hoje.day} de {meses[hoje.month]} de {hoje.year}"
    )

    # pega datas extremas do RESUMO
    data_min, data_max = obter_datas_extremos(ws1)

    # mostra no FRONT
    if data_min:
        ws_front.range("D16").value = data_por_extenso(data_min)

    if data_max:
        ws_front.range("D17").value = data_por_extenso(data_max)

    # 👉 retorna as datas para o main
    return data_min, data_max


# ===== Funções REPORT =====#

def inserir_linhas_report(ws_report, linha_inicial, periodos):
    """
    Insere linhas copiando a linha inicial para acomodar periodos > 1
    """
    if periodos <= 1:
        return

    row_height = ws_report.api.Rows(linha_inicial).RowHeight

    for i in range(periodos - 1):
        destino = linha_inicial + 1 + i
        ws_report.api.Rows(destino).Insert()
        ws_report.api.Rows(linha_inicial).Copy(ws_report.api.Rows(destino))
        ws_report.api.Rows(destino).RowHeight = row_height


# ===== COLUNA E ===== #

def obter_periodos(arquivo1_path):
    df = pd.read_excel(arquivo1_path, sheet_name="Resumo", header=None)
    col_aa = df[26].dropna()
    try:
        periodos = int(float(str(col_aa.iloc[-1]).replace("R$", "").replace(",", ".").replace(" ", "")))
    except:
        periodos = 1
    return periodos


def gerar_coluna_E_ajustada(ws1, periodos, coluna_horario="C"):
    """
    Gera a lista de ciclos para preencher a coluna E do REPORT VIGIA.
    
    - Se C3 for 06h, 12h, 18h ou 00h, começa a lista por este ciclo.
    - Se C3 for "Total" ou vazio, assume primeiro ciclo 00x06 e continua a sequência normal.
    - Repete a sequência até completar 'periodos'.
    """
    # Mapear horários para ciclos
    horario_para_ciclo = {
        "06h": "06x12", "06H": "06x12",
        "12h": "12x18", "12H": "12x18",
        "18h": "18x24", "18H": "18x24",
        "00h": "00x06", "00H": "00x06"
    }

    # Sequência padrão completa
    sequencia_padrao = ["06x12", "12x18", "18x24", "00x06"]

    # Ler primeira e segunda células da coluna
    primeiro_horario = str(ws1.range(f"{coluna_horario}2").value).strip()
    segundo_valor = ws1.range(f"{coluna_horario}3").value
    segundo_valor = str(segundo_valor).strip() if segundo_valor is not None else ""

    # Determinar primeiro ciclo
    if segundo_valor.lower() == "total" or segundo_valor not in horario_para_ciclo:
        primeiro_ciclo = "00x06"  # primeiro ciclo é sempre 00x06
    else:
        primeiro_ciclo = horario_para_ciclo[segundo_valor]

    # Rotacionar sequência padrão para iniciar pelo primeiro ciclo
    idx_inicio = sequencia_padrao.index(primeiro_ciclo)
    sequencia = sequencia_padrao[idx_inicio:] + sequencia_padrao[:idx_inicio]

    # Gerar lista final até completar periodos
    ciclos_linha = []
    for c in cycle(sequencia):
        if len(ciclos_linha) >= periodos:
            break
        ciclos_linha.append(c)

    return ciclos_linha




def preencher_coluna_E_por_ciclos(ws_report, ciclos_linha, linha_inicial=22):
    for i, ciclo in enumerate(ciclos_linha):
        ws_report.range(f"E{linha_inicial + i}").value = ciclo




# ===== COLUNA G ===== #


def mapear_valores_por_ciclo(ws1, coluna_horario="C", coluna_valor="Z"):
    """
    Lê os valores do wb1 e agrupa por ciclo.
    ws1 : planilha do arquivo 1
    coluna_horario : coluna que contém os horários (06h, 12h, etc.)
    coluna_valor : coluna que contém os valores a preencher
    """
    horario_para_ciclo = {"06h":"06x12", "12h":"12x18", "18h":"18x24", "00h":"00x06"}
    sequencia_ciclos = ["06x12", "12x18", "18x24", "00x06"]

    last_row = ws1.used_range.last_cell.row
    horarios = ws1.range(f"{coluna_horario}1:{coluna_horario}{last_row}").value
    valores = ws1.range(f"{coluna_valor}1:{coluna_valor}{last_row}").value

    # Normaliza horários para minúsculo
    horarios = [str(h).strip().lower() if h is not None else None for h in horarios]

    valores_por_ciclo = {c: [] for c in sequencia_ciclos}

    for h, v in zip(horarios, valores):
        if h in horario_para_ciclo:
            ciclo = horario_para_ciclo[h]
            valores_por_ciclo[ciclo].append(v)

    return valores_por_ciclo

def preencher_coluna_G_por_ciclo(ws_report, ciclos_linha, valores_por_ciclo, coluna="G", linha_inicial=22):
    """
    Preenche a coluna G do REPORT VIGIA alinhando os valores da coluna Z
    à sequência de ciclos já definida na coluna E.
    """
    indices_ciclo = {c: 0 for c in valores_por_ciclo}

    for i, ciclo_val in enumerate(ciclos_linha):
        linha = linha_inicial + i
        lista_valores = valores_por_ciclo.get(ciclo_val, [])
        idx = indices_ciclo[ciclo_val]

        valor = lista_valores[idx] if idx < len(lista_valores) else None
        indices_ciclo[ciclo_val] += 1

        cel = ws_report.range(f"{coluna}{linha}")
        cel.value = valor

        # Formatação
        try:
            cel.api.NumberFormat = 'R$ #.##0,00'
            cel.api.HorizontalAlignment = xw.constants.HAlign.xlHAlignRight
            cel.api.VerticalAlignment = xw.constants.VAlign.xlVAlignCenter
            cel.api.Font.Name = "Calibri"
            cel.api.Font.Size = 18
        except:
            pass

    return len(ciclos_linha)

# ===== COLUNA C ===== #


def montar_datas_report_vigia(ws_report, ws_resumo, linha_inicial=22, periodos=None):
    """
    Preenche a coluna C (DATE) do REPORT VIGIA.
    - O dia só avança quando o ciclo da coluna E for 00x06.
    - Mantém a sequência correta independentemente do primeiro horário.
    """
    if periodos is None:
        raise ValueError("É necessário informar 'periodos' para preencher as datas")

    data_inicio, data_fim = obter_datas_extremos(ws_resumo)
    if not data_inicio or not data_fim:
        raise ValueError("Não foi possível determinar as datas extremas na aba RESUMO")

    data_atual = data_inicio

    for i in range(periodos):
        linha = linha_inicial + i
        ciclo = ws_report.range(f"E{linha}").value

        if ciclo in (None, ""):
            break

        # Coloca a data atual na coluna C
        ws_report.range(f"C{linha}").value = data_atual

        # Se o ciclo for 00x06, incrementa o dia para a próxima linha
        if isinstance(ciclo, str) and ciclo.strip().lower() == "00x06":
            data_atual += timedelta(days=1)

    return periodos


# ===== DATA INICIAL E FINAL DO FRONT =====#

MESES_EN = {
    1: "JAN", 2: "FEB", 3: "MAR", 4: "APR",
    5: "MAY", 6: "JUN", 7: "JUL", 8: "AUG",
    9: "SEP", 10: "OCT", 11: "NOV", 12: "DEC"
}

def obter_datas_extremos(ws_resumo):
    last_row = ws_resumo.used_range.last_cell.row
    valores = ws_resumo.range(f"B1:B{last_row}").value

    datas = []
    hoje = date.today()

    for v in valores:
        if v in (None, "", "Total"):
            continue

        # datetime vindo do Excel
        if isinstance(v, datetime):
            d = v.date()

            # 🚫 ignora fórmulas HOJE()
            if d == hoje:
                continue

            datas.append(d)
            continue

        # string
        if isinstance(v, str):
            v = v.strip().lower()

            # 19/10/2025
            try:
                datas.append(datetime.strptime(v, "%d/%m/%Y").date())
                continue
            except:
                pass

            # 19/out/25
            try:
                dia, mes_txt, ano = v.split("/")
                mes = MESES_EN.get(int(mes_txt))
                if mes:
                    ano = int(ano)
                    if ano < 100:
                        ano += 2000
                    datas.append(date(ano, mes, int(dia)))
            except:
                pass

    if not datas:
        return None, None

    return min(datas), max(datas)


# ===== ABAS ESPECIFICAS =====#


def OC(arquivo1, wb2):
    ws = wb2.sheets["FRONT VIGIA"]
    if str(ws["G16"].value).strip().upper() == "O.C.:":
        ws["H16"].value = input("OC: ")

def credit_note(wb, valor_c21):
    if "Credit Note" in [s.name for s in wb.sheets]:
        wb.sheets["Credit Note"]["C21"].value = valor_c21

def quitacao(wb, valor_c21):
    if "Quitação" not in [s.name for s in wb.sheets]: return
    ws = wb.sheets["Quitação"]
    ws["C22"].value = valor_c21
    pasta_pdfs = os.path.join(os.path.expanduser("~"), "Desktop", "JANEIRO")
    pdfs = [f for f in os.listdir(pasta_pdfs) if f.lower().endswith(".pdf")]
    pdfs.sort(key=lambda x: int(os.path.splitext(x)[0]))
    ws["H22"].value = f"NF.: {len(pdfs)+1}"

def cargonave(ws):
    valor_c9 = ws.range("C9").value
    return str(valor_c9).strip().upper() == "A/C AGÊNCIA MARÍTIMA CARGONAVE LTDA."

def arredondar_para_baixo_50(ws_front_vigia):
    if not cargonave(ws_front_vigia): return
    valor = ws_front_vigia.range("E37").value
    if valor is None: return
    try: resultado = (int(valor)//50)*50
    except: return
    ws_front_vigia.range("H28").value = resultado

def obter_nome_navio_da_pasta(caminho_arquivo):
    """
    Ex: '123 - NAVIO' -> 'NAVIO'
    """
    pasta = os.path.basename(os.path.dirname(caminho_arquivo))

    if "-" in pasta:
        return pasta.split("-", 1)[1].strip()

    return pasta.strip()

def obter_aba_nf_opcional(wb):
    for sheet in wb.sheets:
        nome = sheet.name.strip().lower()
        if nome == "nf" or nome.startswith("nf") or "nota" in nome:
            return sheet
    return None

def escrever_nf(wb_faturamento, nome_navio, dn):
    # tenta localizar aba NF
    ws_nf = None
    for sheet in wb_faturamento.sheets:
        if sheet.name.strip().lower() == "nf":
            ws_nf = sheet
            break

    if ws_nf is None:
        print("⚠️ Aba NF não encontrada — seguindo sem escrever NF")
        return  # NÃO quebra o programa

    ano = datetime.now().year

    texto = (
        f"SERVIÇO PRESTADO DE ATENDIMENTO/APOIO AO M/V {nome_navio}\n"
        f"DN {dn}/{ano}"
    )

    # escreve na primeira célula
    cel = ws_nf.range("A1")
    cel.value = texto

    # mescla para ficar bonito
    ws_nf.range("A1:E2").merge()

    # formatação
    cel.api.HorizontalAlignment = -4108  # center
    cel.api.VerticalAlignment = -4108
    cel.api.WrapText = True
    cel.api.Font.Name = "Calibri"
    cel.api.Font.Size = 14
    cel.api.Font.Bold = True

    print("✅ Texto da NF escrito com sucesso")

def main():
    print("🚀 Iniciando execução...")

    # ========= 1 – Licença =========
    hoje_utc, hoje_local = data_online()
    limite = datetime(hoje_utc.year, hoje_utc.month, 22, tzinfo=timezone.utc)

    if hoje_utc > limite:
        sys.exit("⛔ Licença expirada")

    print(f"📅 Data local: {hoje_local.date()}")

    # ========= 2 – Abrir arquivos =========
    app, wb1, wb2, ws1, ws_front = abrir_workbooks()
    if not all([app, wb1, wb2, ws1, ws_front]):
        sys.exit("❌ Erro ao abrir workbooks")

    print("📂 Workbooks abertos")

    # ========= 3 – DN e Navio =========
    dn = obter_dn_da_pasta(wb1.fullname)
    if not dn:
        sys.exit("❌ DN não identificada pela pasta")

    nome_navio = obter_nome_navio_da_pasta(wb1.fullname)
    ano_atual = datetime.now().year
    texto_dn = f"DN: {dn}/{ano_atual}"

    # FRONT VIGIA
    ws_front.range("D15").value = nome_navio
    ws_front.range("C21").value = texto_dn

#    berco = input("WAREHOUSE / BERÇO: ").strip().upper()
#    ws_front["D18"].value = berco

    # ========= 4 – FRONT (OBRIGATÓRIO PRIMEIRO) =========
    print("⚙️ Processando FRONT VIGIA...")
    data_inicio, data_fim = processar_front(ws1, ws_front)
    if not data_inicio or not data_fim:
        sys.exit("❌ Datas extremas inválidas no RESUMO")
    print(f"📆 Datas extremas: {data_inicio} → {data_fim}")

    # ========= 5 – MMO =========
    print("⚙️ Processando MMO...")
    MMO(wb1.fullname, wb2)

    # ========= 6 – NF =========
    escrever_nf(wb2, nome_navio, dn)

    # ===== 7 – REPORT VIGIA =====
    ws_resumo = wb1.sheets["Resumo"]
    ws_report = wb2.sheets["REPORT VIGIA"]

    # 1️⃣ Obter quantidade de períodos
    periodos = obter_periodos(wb1.fullname)

    # 2️⃣ Inserir linhas extras se necessário
    inserir_linhas_report(ws_report, linha_inicial=22, periodos=periodos)

    # 3️⃣ Gerar lista de ciclos confiável (coluna E)
    ciclos_linha = gerar_coluna_E_ajustada(ws1, periodos, coluna_horario="C")

    # 4️⃣ Preencher coluna E do REPORT VIGIA
    preencher_coluna_E_por_ciclos(ws_report, ciclos_linha, linha_inicial=22)


    # 5️⃣ Mapear valores do wb1 por ciclo (coluna G)
    valores_por_ciclo = mapear_valores_por_ciclo(ws1, coluna_horario="C", coluna_valor="Z")

    # 6️⃣ Preencher coluna G do REPORT VIGIA
    preencher_coluna_G_por_ciclo(ws_report, ciclos_linha, valores_por_ciclo, coluna="G", linha_inicial=22)

    # 7️⃣ Preencher coluna C (datas) respeitando ciclos 00x06
    montar_datas_report_vigia(
        ws_report=ws_report,
        ws_resumo=ws_resumo,
    linha_inicial=22,
    periodos=periodos
)



    # 7️⃣ Garantir que data_inicio seja datetime
    if not isinstance(data_inicio, datetime):
        try:
            data_inicio = pd.to_datetime(data_inicio)
        except Exception as e:
            raise ValueError(f"data_inicio inválida: {data_inicio}") from e


    # ========= 8 – Financeiro =========
    OC(str(wb1.fullname), wb2)
    credit_note(wb2, texto_dn)
    quitacao(wb2, texto_dn)

    # ========= 10 – Ajustes finais =========
    arredondar_para_baixo_50(ws_front)
    cargonave(ws_front)

    pasta_saida = Path(wb1.fullname).parent
    arquivo_saida = pasta_saida / "3.xlsx"

    fechar_workbooks(app, wb1, wb2, arquivo_saida)

    print(f"✅ Processo finalizado: {arquivo_saida}")


if __name__ == "__main__":
    
    main()

# Fim do código
