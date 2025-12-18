# 1️⃣ – Importações
import xlwings as xw
from datetime import datetime, timedelta, timezone
from openpyxl.styles import Alignment
import pandas as pd
import locale
import os
import sys
import urllib.request
import ssl
import certifi
from pathlib import Path
import tkinter as tk
from tkinter import filedialog



# =========================
# Funções utilitárias
# =========================

def abrir_workbooks():
    """Abre os arquivos 1.xlsx e 2.xlsx de uma pasta escolhida pelo usuário."""
    root = tk.Tk()
    root.withdraw()
    pasta = filedialog.askdirectory(title="Selecione a pasta com os arquivos")
    if not pasta:
        print("Nenhuma pasta selecionada. Encerrando.")
        return None, None, None, None, None

    pasta = Path(pasta)
    arquivo1 = pasta / "1.xlsx"
    arquivo2 = pasta / "2.xlsx"

    app = xw.App(visible=False)
    wb1 = wb2 = None

    try:
        if not arquivo1.exists():
            arquivo1.touch()
        if not arquivo2.exists():
            arquivo2.touch()

        wb1 = app.books.open(arquivo1)
        wb2 = app.books.open(arquivo2)

        ws1 = wb1.sheets[0]
        ws_front = wb2.sheets["FRONT VIGIA"]

        return app, wb1, wb2, ws1, ws_front

    except Exception as e:
        print(f"Erro ao abrir os arquivos: {e}")
        if wb1: wb1.close()
        if wb2: wb2.close()
        app.quit()
        return None, None, None, None, None
    
def fechar_workbooks(app, wb1=None, wb2=None):
    """Salva e fecha os workbooks e encerra o Excel corretamente."""
    try:
        if wb1:
            wb1.save()
            wb1.close()

        if wb2:
            wb2.save()
            wb2.close()

    finally:
        if app:
            app.quit()




def data_online():
    context = ssl.create_default_context(cafile=certifi.where())

    req = urllib.request.Request(
        "https://www.cloudflare.com",
        headers={"User-Agent": "Mozilla/5.0"}
    )

    with urllib.request.urlopen(req, context=context, timeout=5) as r:
        data_str = r.headers["Date"]

    # transforma em datetime UTC do servidor
    dt_utc = datetime.strptime(data_str, "%a, %d %b %Y %H:%M:%S %Z").replace(tzinfo=timezone.utc)

    # para uso interno/exibição: converte para hora local
    dt_local = dt_utc.astimezone()  # pega fuso local do sistema
    return dt_utc, dt_local

try:
    hoje_utc, hoje_local = data_online()

    # validação sempre UTC
    data_limite_utc = datetime(hoje_utc.year, hoje_utc.month, 22, tzinfo=timezone.utc)
    if hoje_utc > data_limite_utc:
        sys.exit("⛔ Licença expirada")

    # apenas exibição: sempre local
    print("Data local:", hoje_local.date())


except Exception as e:
    sys.exit(f"Erro ao verificar licença: {e}")

    sys.exit("❌ Não foi possível validar a data online")


def data_por_extenso(valor):
    if isinstance(valor, datetime):
        data = valor
    else:
        try:
            data = datetime.strptime(str(valor), "%d/%m/%Y")
        except:
            data = datetime.now()
    return data.strftime("%B %d, %Y")

def aplicar_fonte_xw(ws, celula, tamanho=15):
    """
    Aplica fonte Arial tamanho definido (padrão = 15)
    e centraliza a célula horizontal e verticalmente usando xlwings.
    """
    rng = ws.range(celula)
    rng.api.Font.Name = "Calibri"
    rng.api.Font.Size = tamanho
    rng.api.Font.Bold = False
    rng.api.Font.Italic = False
    rng.api.Font.Underline = False
    rng.api.HorizontalAlignment = xw.constants.HAlign.xlHAlignCenter
    rng.api.VerticalAlignment = xw.constants.VAlign.xlVAlignCenter

# =========================
# Funções específicas do relatório
# =========================

def processar_report(wb1, wb2, ws1, ws_front, arquivo1_path):
    # Data por extenso em FRONT VIGIA
    meses = ["", "janeiro", "fevereiro", "março", "abril", "maio", "junho",
             "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"]
    hoje = datetime.now()
    ws_front.range("C39").value = f" Santos, {hoje.day} de {meses[hoje.month]} de {hoje.year}"

    # REPORT VIGIA
    ws_report = wb2.sheets["REPORT VIGIA"]
    linha_origem = 22
    row_height = ws_report.api.Rows(linha_origem).RowHeight

    df = pd.read_excel(arquivo1_path, sheet_name="Resumo", header=None)
    col_aa = df[26].dropna()
    periodos = int(float(str(col_aa.iloc[-1]).replace("R$", "").replace(",", ".").replace(" ", "")))

    for i in range(periodos):
        destino = linha_origem + 1 + i
        ws_report.api.Rows(destino).Insert()
        ws_report.api.Rows(linha_origem).Copy(ws_report.api.Rows(destino))
        ws_report.api.Rows(destino).RowHeight = row_height

    # Preencher coluna E com ciclo
    df2 = pd.read_excel(arquivo1_path, sheet_name="Resumo")
    coluna_c = df2.iloc[:, 2].astype(str).str.strip()
    lista_reduzida = [v for v in coluna_c if v.lower() != "total"]
    horarios = [h.lower() for h in lista_reduzida if h.lower() in ["00h","06h","12h","18h"]]
    ordem = ["06h","12h","18h","00h"]
    menor_horario = horarios[0] if horarios else "06h"
    ciclo = ["06x12","12x18","18x00","00x06"]
    mapa = {"06h":0,"12h":1,"18h":2,"00h":3}
    indice = mapa[menor_horario]

    for i in range(periodos):
        linha = linha_origem + i
        rng = ws_report.range(f"E{linha}")
        rng.value = ciclo[(indice+i)%4]
        try:
            if rng.api.MergeCells:
                rng.api.UnMerge()
        except: pass
        aplicar_fonte_xw(ws_report, f"E{linha}", tamanho=18)

    # Preencher coluna G com valores da coluna Z do wb1
    valores_validos = []
    dadosC = ws1.range("C2").options(expand='down').value
    dadosZ = ws1.range("Z2").options(expand='down').value
    if not isinstance(dadosC,list): dadosC = [dadosC] if dadosC else []
    if not isinstance(dadosZ,list): dadosZ = [dadosZ] if dadosZ else []
    for c_val,z_val in zip(dadosC,dadosZ):
        if c_val is None: break
        if str(c_val).strip().lower() in ["00h","06h","12h","18h"]:
            valores_validos.append(z_val)
    valores_validos = valores_validos[:periodos]

    linha_destino = linha_origem
    for valor in valores_validos:
        cel = ws_report.range(f"G{linha_destino}")
        cel.value = valor
        cel.api.HorizontalAlignment = xw.constants.HAlign.xlHAlignRight
        cel.api.VerticalAlignment = xw.constants.VAlign.xlVAlignCenter
        cel.api.NumberFormat = 'R$ #.##0,00'
        cel.api.Font.Name = "Calibri"
        cel.api.Font.Size = 18
        linha_destino += 1

        # =========================
    # Preencher coluna C com datas (incrementa quando ciclo for 00x06)
    # =========================

    # Data inicial (vem do FRONT VIGIA D16)
    valor_data = ws_front.range("D16").value

    if isinstance(valor_data, datetime):
        data_atual = valor_data
    else:
        try:
            data_atual = datetime.strptime(str(valor_data), "%d/%m/%Y")
        except:
            data_atual = datetime.now()

    for i in range(periodos):
        linha = linha_origem + i

        # escreve a data por extenso na coluna C
        ws_report.range(f"C{linha}").value = data_por_extenso(data_atual)
        aplicar_fonte_xw(ws_report, f"C{linha}", tamanho=20)

        # se o ciclo for 00x06, avança 1 dia
        ciclo_atual = ws_report.range(f"E{linha}").value
        if ciclo_atual == "00x06":
            data_atual += timedelta(days=1)


# 6️⃣ – Funções auxiliares
def obter_data_mais_distante(ws_origem, ws_front_vigia, berco):
    """Pega a maior data da coluna B e coloca em D17, atualiza D18 com berco"""
    last_row = ws_origem.used_range.last_cell.row
    coluna_b = ws_origem.range(f"B1:B{last_row}").value
    datas_validas = []
    for v in coluna_b:
        if v in (None, "", "Total"): continue
        if isinstance(v, datetime):
            datas_validas.append(v)
        elif isinstance(v, str):
            try: datas_validas.append(datetime.strptime(v.strip(), "%d/%m/%Y"))
            except: pass
    if not datas_validas: return
    ws_front_vigia["D17"].value = data_por_extenso(max(datas_validas))
    ws_front_vigia["D18"].value = berco

def atualizar_report_vigia(arquivo1, wb2):
    ws = wb2.sheets["REPORT VIGIA"]
    if str(ws["E25"].value).strip().upper() != "MMO": return
    df = pd.read_excel(arquivo1, sheet_name="Resumo", header=None)
    col_g = df[6].dropna()
    if col_g.empty: return
    try:
        ultimo_float = locale.atof(str(col_g.iloc[-1]).replace("R$", "").strip())
    except: ultimo_float = float(col_g.iloc[-1])
    ws["F25"].value = ultimo_float
    ws["F25"].number_format = "#.##0,00"

def OC(arquivo1, wb2):
    ws = wb2.sheets["FRONT VIGIA"]
    if str(ws["G16"].value).strip().upper() == "O.C.:":
        ws["H16"].value = input("OC: ")

def credit_note(wb, valor_c21):
    if "Credit Note" in [s.name for s in wb.sheets]:
        wb.sheets["Credit Note"]["C21"].value = valor_c21

def quitacao(wb, valor_c21):
    if "Quitação" not in [s.name for s in wb.sheets]: return
    ws = wb.sheets["Quitação"]
    ws["C22"].value = valor_c21
    pasta_pdfs = os.path.join(os.path.expanduser("~"), "Desktop", "JANEIRO")
    pdfs = [f for f in os.listdir(pasta_pdfs) if f.lower().endswith(".pdf")]
    pdfs.sort(key=lambda x: int(os.path.splitext(x)[0]))
    ws["H22"].value = f"NF.: {len(pdfs)+1}"

def validar_c9(ws):
    valor_c9 = ws.range("C9").value
    return str(valor_c9).strip().upper() == "A/C AGÊNCIA MARÍTIMA CARGONAVE LTDA."

def arredondar_para_baixo_50(ws_front_vigia):
    if not validar_c9(ws_front_vigia): return
    valor = ws_front_vigia.range("E37").value
    if valor is None: return
    try: resultado = (int(valor)//50)*50
    except: return
    ws_front_vigia.range("H28").value = resultado

# =========================
# Programa principal
# =========================

# 7️⃣ – Main
def main():
    # 1 – Validar licença
    try:
        hoje_utc, hoje_local = data_online()
        if hoje_utc > datetime(hoje_utc.year, hoje_utc.month, 22, tzinfo=timezone.utc):
            sys.exit("⛔ Licença expirada")
    except Exception as e:
        sys.exit(f"Erro ao verificar licença: {e}")

    # 2 – Abrir arquivos
    app, wb1, wb2, ws1, ws_front = abrir_workbooks()
    if not all([app, wb1, wb2, ws1, ws_front]):
        return

    # 3 – Inputs do usuário
    numero_dn = input("DN: ").strip()
    berco = input("WAREHOUSE / BERÇO: ").strip().upper()
    ws_front["D18"].value = berco

    # DN SOMENTE NO FRONT VIGIA
    valor_c21 = f"DN: {numero_dn}/25"
    ws_front.range("C21").value = valor_c21

    # 4 – Transferências iniciais
    ws_front.range("D15").value = ws1.range("A2").value
    ws_front.range("D16").value = data_por_extenso(ws1.range("B2").value)

    # 5 – Processamento
    processar_report(wb1, wb2, ws1, ws_front, str(wb1.fullname))
    OC(str(wb1.fullname), wb2)
    obter_data_mais_distante(ws1, ws_front, berco)
    atualizar_report_vigia(str(wb1.fullname), wb2)

    # ❌ NÃO colar DN fora do FRONT VIGIA
    # credit_note(wb2, valor_c21)
    # quitacao(wb2, valor_c21)

    credit_note(wb2, valor_c21)
    quitacao(wb2, valor_c21)

    # 6 – Ajustes finais
    arredondar_para_baixo_50(ws_front)
    validar_c9(ws_front)
    # Após todos os ajustes no wb2
    # 7 – Salvar e fechar
    wb1.save()  # mantém 1.xlsx
    # Salva wb2 como um novo arquivo na mesma pasta dos originais
    arquivo_saida = Path(wb2.fullname).parent / "3.xlsx"
    wb2.save(arquivo_saida)

    fechar_workbooks(app, wb1, wb2)

    print(f"Processo finalizado. Arquivo final salvo em: {arquivo_saida}")



if __name__ == "__main__":
    main()
# Fim do código
