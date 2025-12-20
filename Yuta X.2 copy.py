# 1️⃣ – Importações
import xlwings as xw
from datetime import datetime, timedelta, timezone
from openpyxl.styles import Alignment
import pandas as pd
import os
import sys
import urllib.request
import ssl
import certifi
from pathlib import Path
import tkinter as tk
from tkinter import filedialog
from datetime import date
import re
import locale


# =========================
# Funções utilitárias
# =========================

def abrir_workbooks():
    """
    Abre:
    - 1.xlsx da pasta do NAVIO (selecionada pelo usuário)
    - XLSX do CLIENTE em Desktop/FATURAMENTOS/<CLIENTE>.xlsx
    """

    # --- Seleção da pasta do navio ---
    root = tk.Tk()
    root.withdraw()
    pasta_navio = filedialog.askdirectory(
        title="Selecione a pasta do NAVIO (onde está o 1.xlsx)"
    )

    if not pasta_navio:
        print("Nenhuma pasta selecionada. Encerrando.")
        return None, None, None, None, None

    pasta_navio = Path(pasta_navio)
    pasta_cliente = pasta_navio.parent   # PASTA PAI
    nome_cliente = pasta_cliente.name

    arquivo1 = pasta_navio / "1.xlsx"

    pasta_faturamentos = Path.home() / "Desktop" / "FATURAMENTOS"
    arquivo2 = pasta_faturamentos / f"{nome_cliente}.xlsx"

    app = xw.App(visible=False)
    wb1 = wb2 = None

    try:
        # --- Valida arquivos ---
        if not arquivo1.exists():
            raise FileNotFoundError(f"Arquivo 1.xlsx não encontrado:\n{arquivo1}")

        if not arquivo2.exists():
            raise FileNotFoundError(
                f"Arquivo de faturamento do cliente não encontrado:\n{arquivo2}"
            )

        # --- Abre workbooks ---
        wb1 = app.books.open(arquivo1)
        wb2 = app.books.open(arquivo2)

        ws1 = wb1.sheets[0]

        nomes_abas = [s.name for s in wb2.sheets]

        # Prioridade: aba com nome do cliente
        if nome_cliente in nomes_abas:
            ws_front = wb2.sheets[nome_cliente]

        # Fallback: FRONT VIGIA
        elif "FRONT VIGIA" in nomes_abas:
            ws_front = wb2.sheets["FRONT VIGIA"]

        else:
            raise RuntimeError(
                f"Nenhuma aba válida encontrada no faturamento.\n"
                f"Esperado: '{nome_cliente}' ou 'FRONT VIGIA'."
            )

        return app, wb1, wb2, ws1, ws_front

    except Exception as e:
        print(f"Erro ao abrir os arquivos: {e}")

        if wb1:
            wb1.close()
        if wb2:
            wb2.close()

        app.quit()
        return None, None, None, None, None

    
def fechar_workbooks(app, wb1=None, wb2=None, arquivo_saida=None):
    """
    Salva wb1 normalmente e salva wb2 EXCLUSIVAMENTE como 3.xlsx
    na pasta do cliente (ex: WILSON). Nunca salva em FATURAMENTOS.
    """
    try:
        if wb1:
            wb1.save()
            wb1.close()

        if wb2:
            if not arquivo_saida:
                raise RuntimeError(
                    "Caminho de saída não informado. "
                    "wb2 NÃO será salvo para evitar salvar em FATURAMENTOS."
                )

            wb2.save(arquivo_saida)
            wb2.close()

    finally:
        if app:
            app.quit()

def obter_dn_da_pasta(caminho_arquivo):
    """
    Extrai números do nome da pasta do navio
    Ex: '123 - NAVIO' -> '123'
    """
    pasta = os.path.basename(os.path.dirname(caminho_arquivo))

    numeros = re.findall(r"\d+", pasta)

    if not numeros:
        return None

    return numeros[0]  # primeiro bloco numérico


def data_online():
    context = ssl.create_default_context(cafile=certifi.where())

    req = urllib.request.Request(
        "https://www.cloudflare.com",
        headers={"User-Agent": "Mozilla/5.0"}
    )

    with urllib.request.urlopen(req, context=context, timeout=5) as r:
        data_str = r.headers["Date"]

    # transforma em datetime UTC do servidor
    dt_utc = datetime.strptime(data_str, "%a, %d %b %Y %H:%M:%S %Z").replace(tzinfo=timezone.utc)

    # para uso interno/exibição: converte para hora local
    dt_local = dt_utc.astimezone()  # pega fuso local do sistema
    return dt_utc, dt_local

try:
    hoje_utc, hoje_local = data_online()

    # validação sempre UTC
    data_limite_utc = datetime(hoje_utc.year, hoje_utc.month, 22, tzinfo=timezone.utc)
    if hoje_utc > data_limite_utc:
        sys.exit("⛔ Licença expirada")

    # apenas exibição: sempre local
    print("Data local:", hoje_local.date())


except Exception as e:
    sys.exit(f"Erro ao verificar licença: {e}")

    sys.exit("❌ Não foi possível validar a data online")



def data_por_extenso(valor):
    if isinstance(valor, datetime):
        data = valor

    elif isinstance(valor, date):
        data = datetime(valor.year, valor.month, valor.day)

    elif isinstance(valor, str):
        try:
            data = datetime.strptime(valor, "%d/%m/%Y")
        except:
            return ""  # não inventa data

    else:
        return ""  # nunca usa datetime.now()

    return data.strftime("%d de %B de %Y")

def MMO(arquivo1, wb2):
    ws = wb2.sheets["REPORT VIGIA"]
    if str(ws["E25"].value).strip().upper() != "MMO": return
    df = pd.read_excel(arquivo1, sheet_name="Resumo", header=None)
    col_g = df[6].dropna()
    if col_g.empty: return
    try:
        ultimo_float = locale.atof(str(col_g.iloc[-1]).replace("R$", "").strip())
    except: ultimo_float = float(col_g.iloc[-1])
    ws["F25"].value = ultimo_float
    ws["F25"].number_format = "#.##0,00"


def processar_front(ws1, ws_front):
    """
    Atualiza somente a aba FRONT VIGIA
    """

    # data atual por extenso (rodapé)
    meses = ["", "janeiro", "fevereiro", "março", "abril", "maio", "junho",
             "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"]

    hoje = datetime.now()
    ws_front.range("C39").value = (
        f"Santos, {hoje.day} de {meses[hoje.month]} de {hoje.year}"
    )

    # pega datas extremas do RESUMO
    data_min, data_max = obter_datas_extremos(ws1)

    # mostra no FRONT
    if data_min:
        ws_front.range("D16").value = data_por_extenso(data_min)

    if data_max:
        ws_front.range("D17").value = data_por_extenso(data_max)

    # 👉 retorna as datas para o main
    return data_min, data_max


def processar_report_completo(
    wb2,
    ws_origem,
    ws_front,
    arquivo1_path,
    linha_inicial=22
):
    """
    Atualiza a aba REPORT VIGIA:
    1) Cria linhas conforme o número de períodos (Resumo AA)
    2) Preenche coluna E com turnos alinhados ao Resumo
    3) Preenche coluna G com valores da coluna Z
    """

    ws_report = wb2.sheets["REPORT VIGIA"]
    row_height = ws_report.api.Rows(linha_inicial).RowHeight

    # ======================================================
    # 1) obter quantidade de períodos (coluna AA)
    # ======================================================
    df_resumo = pd.read_excel(arquivo1_path, sheet_name="Resumo", header=None)
    col_aa = df_resumo[26].dropna()

    try:
        periodos = int(float(
            str(col_aa.iloc[-1])
            .replace("R$", "")
            .replace(",", ".")
            .replace(" ", "")
        ))
    except:
        periodos = 1

    # ======================================================
    # 2) inserir linhas no REPORT
    # ======================================================
    for i in range(periodos - 1):
        destino = linha_inicial + 1 + i
        ws_report.api.Rows(destino).Insert()
        ws_report.api.Rows(linha_inicial).Copy(ws_report.api.Rows(destino))
        ws_report.api.Rows(destino).RowHeight = row_height

    # ======================================================
    # 3) ler horários (col C) e valores (col Z)
    # ======================================================
    df2 = pd.read_excel(arquivo1_path, sheet_name="Resumo")
    col_c = df2.iloc[:, 2].astype(str)
    col_z = df2.iloc[:, 25]

    mapa_horario = {
        "00": "00x06",
        "06": "06x12",
        "12": "12x18",
        "18": "18x00"
    }

    # --- descobrir o primeiro turno real ---
    primeiro_turno = None
    for h in col_c:
        h_norm = h.lower().replace("h", "").replace(":", "").strip()
        if h_norm in mapa_horario:
            primeiro_turno = mapa_horario[h_norm]
            break

    if not primeiro_turno:
        primeiro_turno = "06x12"  # fallback seguro

    # ======================================================
    # 4) separar valores por ciclo
    # ======================================================
    valores_por_ciclo = {
        "06x12": [],
        "12x18": [],
        "18x00": [],
        "00x06": []
    }

    for h, z in zip(col_c, col_z):
        h_norm = h.lower().replace("h", "").replace(":", "").strip()
        if h_norm in mapa_horario:
            ciclo = mapa_horario[h_norm]
            valores_por_ciclo[ciclo].append(z)

    # ======================================================
    # 5) preencher coluna E (turnos alinhados)
    # ======================================================
    ciclo_base = ["06x12", "12x18", "18x00", "00x06"]
    indice_inicio = ciclo_base.index(primeiro_turno)

    ciclos_linha = []

    for i in range(periodos):
        linha = linha_inicial + i
        ciclo = ciclo_base[(indice_inicio + i) % 4]

        cel = ws_report.range(f"E{linha}")
        try:
            if cel.api.MergeCells:
                cel.api.UnMerge()
        except:
            pass

        cel.value = ciclo
        ciclos_linha.append(ciclo)

    # ======================================================
    # 6) preencher coluna G (valores)
    # ======================================================
    indices_ciclo = {c: 0 for c in valores_por_ciclo}

    for i, ciclo in enumerate(ciclos_linha):
        linha = linha_inicial + i
        idx = indices_ciclo[ciclo]
        lista = valores_por_ciclo[ciclo]

        valor = lista[idx] if idx < len(lista) else None
        indices_ciclo[ciclo] += 1

        cel = ws_report.range(f"G{linha}")
        cel.value = valor
        cel.api.NumberFormat = 'R$ #.##0,00'
        cel.api.HorizontalAlignment = -4152
        cel.api.VerticalAlignment = -4108
        cel.api.Font.Name = "Calibri"
        cel.api.Font.Size = 18

    return periodos






MESES_EN = {
    1: "JAN", 2: "FEB", 3: "MAR", 4: "APR",
    5: "MAY", 6: "JUN", 7: "JUL", 8: "AUG",
    9: "SEP", 10: "OCT", 11: "NOV", 12: "DEC"
}

def obter_datas_extremos(ws_resumo):
    last_row = ws_resumo.used_range.last_cell.row
    valores = ws_resumo.range(f"B1:B{last_row}").value

    datas = []
    hoje = date.today()

    for v in valores:
        if v in (None, "", "Total"):
            continue

        # datetime vindo do Excel
        if isinstance(v, datetime):
            d = v.date()

            # 🚫 ignora fórmulas HOJE()
            if d == hoje:
                continue

            datas.append(d)
            continue

        # string
        if isinstance(v, str):
            v = v.strip().lower()

            # 19/10/2025
            try:
                datas.append(datetime.strptime(v, "%d/%m/%Y").date())
                continue
            except:
                pass

            # 19/out/25
            try:
                dia, mes_txt, ano = v.split("/")
                mes = MESES_EN.get(int(mes_txt))
                if mes:
                    ano = int(ano)
                    if ano < 100:
                        ano += 2000
                    datas.append(date(ano, mes, int(dia)))
            except:
                pass

    if not datas:
        return None, None

    return min(datas), max(datas)


def preencher_datas_e_turnos_report_vigia(periodos, linha_inicial,ws_report, data_inicio):
        data_atual = data_inicio

        for i in range(periodos):
                linha = linha_inicial + i
                turno = ws_report.range(f"E{linha}").value

                ws_report.range(f"C{linha}").value = data_atual

                if isinstance(turno, str) and turno.strip() == "00x06":
                    data_atual += timedelta(days=1)








def OC(arquivo1, wb2):
    ws = wb2.sheets["FRONT VIGIA"]
    if str(ws["G16"].value).strip().upper() == "O.C.:":
        ws["H16"].value = input("OC: ")

def credit_note(wb, valor_c21):
    if "Credit Note" in [s.name for s in wb.sheets]:
        wb.sheets["Credit Note"]["C21"].value = valor_c21

def quitacao(wb, valor_c21):
    if "Quitação" not in [s.name for s in wb.sheets]: return
    ws = wb.sheets["Quitação"]
    ws["C22"].value = valor_c21
    pasta_pdfs = os.path.join(os.path.expanduser("~"), "Desktop", "JANEIRO")
    pdfs = [f for f in os.listdir(pasta_pdfs) if f.lower().endswith(".pdf")]
    pdfs.sort(key=lambda x: int(os.path.splitext(x)[0]))
    ws["H22"].value = f"NF.: {len(pdfs)+1}"

def cargonave(ws):
    valor_c9 = ws.range("C9").value
    return str(valor_c9).strip().upper() == "A/C AGÊNCIA MARÍTIMA CARGONAVE LTDA."

def arredondar_para_baixo_50(ws_front_vigia):
    if not cargonave(ws_front_vigia): return
    valor = ws_front_vigia.range("E37").value
    if valor is None: return
    try: resultado = (int(valor)//50)*50
    except: return
    ws_front_vigia.range("H28").value = resultado

# =========================
# Programa principal
# =========================

def obter_nome_navio_da_pasta(caminho_arquivo):
    """
    Ex: '123 - NAVIO' -> 'NAVIO'
    """
    pasta = os.path.basename(os.path.dirname(caminho_arquivo))

    if "-" in pasta:
        return pasta.split("-", 1)[1].strip()

    return pasta.strip()


def obter_aba_nf_opcional(wb):
    for sheet in wb.sheets:
        nome = sheet.name.strip().lower()
        if nome == "nf" or nome.startswith("nf") or "nota" in nome:
            return sheet
    return None

def escrever_nf(wb_faturamento, nome_navio, dn):
    # tenta localizar aba NF
    ws_nf = None
    for sheet in wb_faturamento.sheets:
        if sheet.name.strip().lower() == "nf":
            ws_nf = sheet
            break

    if ws_nf is None:
        print("⚠️ Aba NF não encontrada — seguindo sem escrever NF")
        return  # NÃO quebra o programa

    ano = datetime.now().year

    texto = (
        f"SERVIÇO PRESTADO DE ATENDIMENTO/APOIO AO M/V {nome_navio}\n"
        f"DN {dn}/{ano}"
    )

    # escreve na primeira célula
    cel = ws_nf.range("A1")
    cel.value = texto

    # mescla para ficar bonito
    ws_nf.range("A1:E2").merge()

    # formatação
    cel.api.HorizontalAlignment = -4108  # center
    cel.api.VerticalAlignment = -4108
    cel.api.WrapText = True
    cel.api.Font.Name = "Calibri"
    cel.api.Font.Size = 14
    cel.api.Font.Bold = True

    print("✅ Texto da NF escrito com sucesso")



def main():
    print("🚀 Iniciando execução...")

    # ========= 1 – Licença =========
    hoje_utc, hoje_local = data_online()
    limite = datetime(hoje_utc.year, hoje_utc.month, 22, tzinfo=timezone.utc)

    if hoje_utc > limite:
        sys.exit("⛔ Licença expirada")

    print(f"📅 Data local: {hoje_local.date()}")

    # ========= 2 – Abrir arquivos =========
    app, wb1, wb2, ws1, ws_front = abrir_workbooks()
    if not all([app, wb1, wb2, ws1, ws_front]):
        sys.exit("❌ Erro ao abrir workbooks")

    print("📂 Workbooks abertos")

    # ========= 3 – DN e Navio pela pasta =========
    dn = obter_dn_da_pasta(wb1.fullname)
    if not dn:
        sys.exit("❌ DN não identificada pela pasta")

    nome_navio = obter_nome_navio_da_pasta(wb1.fullname)
    ano_atual = datetime.now().year
    texto_dn = f"DN: {dn}/{ano_atual}"

    # FRONT VIGIA
    ws_front.range("D15").value = nome_navio
    ws_front.range("C21").value = texto_dn

    # ========= 4 – Berço =========
#    berco = input("WAREHOUSE / BERÇO: ").strip().upper()
#    ws_front.range("D18").value = berco

    # ========= 5 – FRONT (OBRIGATÓRIO PRIMEIRO) =========
    print("⚙️ Processando FRONT VIGIA...")
    data_inicio, data_fim = processar_front(ws1, ws_front)

    if not data_inicio or not data_fim:
        sys.exit("❌ Datas extremas inválidas no RESUMO")

    print(f"📆 Datas extremas: {data_inicio} → {data_fim}")

    # ========= 5.1 – MMO =========
    print("⚙️ Processando MMO...")
    MMO(wb1.fullname, wb2)

    # ========= 6 – NF =========
    escrever_nf(wb2, nome_navio, dn)

    # ========= 7 – REPORT VIGIA =========



    ws_report = wb2.sheets["REPORT VIGIA"]

    print("⚙️ Processando REPORT VIGIA...")
    periodos = processar_report_completo(
        wb2,
        ws1,
        ws_front,
        wb1.fullname,
        linha_inicial=22
    )

    # ========= 8 – Datas no REPORT =========
    print("🖊️ Preenchendo datas no REPORT VIGIA...")

    if not data_inicio:
        sys.exit("❌ Data inicial inválida para o REPORT VIGIA")

    preencher_datas_e_turnos_report_vigia(
        ws_report=ws_report,
        data_inicio=data_inicio,
        periodos=periodos,
        linha_inicial=22
    )


    # ========= 9 – Financeiro =========
    OC(str(wb1.fullname), wb2)
    credit_note(wb2, texto_dn)
    quitacao(wb2, texto_dn)

 

    # ========= 10 – Ajustes finais =========
    arredondar_para_baixo_50(ws_front)
    cargonave(ws_front)

    pasta_saida = Path(wb1.fullname).parent
    arquivo_saida = pasta_saida / "3.xlsx"

    fechar_workbooks(app, wb1, wb2, arquivo_saida)

    print(f"✅ Processo finalizado: {arquivo_saida}")


if __name__ == "__main__":
    
    main()
    
# Fim do código
