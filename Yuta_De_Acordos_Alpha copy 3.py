# ==============================
# IMPORTS (no topo do arquivo)
# ==============================
import sys
import time
from datetime import datetime
from pathlib import Path
import xlwings as xw
from datetime import datetime, timedelta, timezone
from openpyxl.styles import Alignment
import pandas as pd
import os
import sys
import urllib.request
import ssl
import certifi
from pathlib import Path
import tkinter as tk
from datetime import date
import re
from itertools import cycle
from tkinter import Tk, filedialog
import shutil
import tempfile
import time


# ==============================
# FUN√á√ïES AUXILIARES GLOBAIS (fora de qualquer classe)
# ==============================

def obter_pasta_faturamentos() -> Path:
    # (cola aqui a fun√ß√£o completa que voc√™ tem, a que busca na rede/OneDrive)
    # ... (todo o c√≥digo dela)

def obter_dn_da_pasta(pasta: Path) -> str:
    numeros = re.findall(r"\d+", pasta.name)
    if not numeros:
        raise ValueError("N√£o foi poss√≠vel identificar o DN no nome da pasta.")
    return numeros[0]

def obter_nome_navio_da_pasta(pasta: Path) -> str:
    nome_limpo = re.sub(r"^\d+[\s\-_]*", "", pasta.name, flags=re.IGNORECASE).strip()
    return nome_limpo if nome_limpo else "NAVIO N√ÉO IDENTIFICADO"


def copiar_para_temp_e_ler_excel(caminho_original: Path | str) -> pd.DataFrame:
    """
    Copia o arquivo para pasta tempor√°ria local e l√™ com pandas.
    Resolve PermissionError em OneDrive/rede.
    """
    caminho_original = Path(caminho_original)
    if not caminho_original.exists():
        raise FileNotFoundError(f"Arquivo n√£o encontrado: {caminho_original}")

    with tempfile.TemporaryDirectory() as temp_dir:
        caminho_temp = Path(temp_dir) / caminho_original.name
        print(f"Copiando {caminho_original.name} temporariamente...")
        shutil.copy2(caminho_original, caminho_temp)
        print(f"Lendo {caminho_temp} com pandas...")
        df = pd.read_excel(caminho_temp, engine="openpyxl")
        return df  # ‚Üê ESSA LINHA √â OBRIGAT√ìRIA




def abrir_workbooks(pasta_faturamentos: Path):
            """
            Abre:
            - 1.xlsx (ou similar) da pasta do NAVIO selecionada pelo usu√°rio
            - XLSX do cliente na pasta FATURAMENTOS
            """
            root = tk.Tk()
            root.withdraw()

            # 1Ô∏è‚É£ Usu√°rio seleciona a pasta do NAVIO
            pasta_navio_str = filedialog.askdirectory(
                title="Selecione a pasta do NAVIO (onde est√° o 1.xlsx)"
            )

            if not pasta_navio_str:
                print("Sele√ß√£o cancelada pelo usu√°rio.")
                return None, None, None, None, None

            pasta_navio = Path(pasta_navio_str)
            pasta_cliente = pasta_navio.parent
            nome_cliente = pasta_cliente.name.strip()  # remove espa√ßos acidentais

            # Procura arquivo que come√ßa com "1" (1.xlsx, 1.xls, etc.)
            arquivos_1 = list(pasta_navio.glob("1*.xls*"))
            if not arquivos_1:
                raise FileNotFoundError(
                    f"Nenhum arquivo iniciando com '1' encontrado em:\n{pasta_navio}"
                )

            arquivo1 = arquivos_1[
                0
            ]  # pega o primeiro (se tiver mais de um, pode melhorar depois)

            # 2Ô∏è‚É£ Arquivo de faturamento do cliente (caminho fixo)
            arquivo2 = pasta_faturamentos / f"{nome_cliente}.xlsx"

            if not arquivo2.exists():
                raise FileNotFoundError(
                    f"Arquivo de faturamento n√£o encontrado:\n{arquivo2}"
                )

            app.api.Calculate()
            time.sleep(0.5)

            app = xw.App(visible=False)
            wb1 = wb2 = None

            try:
                # Aqui voc√™ pode usar xlwings direto, mas se quiser evitar lock, pode copiar tamb√©m:
                wb1 = app.books.open(str(arquivo1))
                wb2 = app.books.open(str(arquivo2))

                ws1 = wb1.sheets[0]

                nomes_abas = [s.name for s in wb2.sheets]

                if nome_cliente in nomes_abas:
                    ws_front = wb2.sheets[nome_cliente]
                elif "FRONT VIGIA" in nomes_abas:
                    ws_front = wb2.sheets["FRONT VIGIA"]
                else:
                    raise RuntimeError(
                        f"Nenhuma aba v√°lida encontrada.\n"
                        f"Esperado: '{nome_cliente}' ou 'FRONT VIGIA'"
                    )

                return app, wb1, wb2, ws1, ws_front

            except Exception as e:
                if wb1:
                    wb1.close()
                if wb2:
                    wb2.close()
                if app:
                    app.quit()
                raise e


def fechar_workbooks(app, wb1=None, wb2=None, arquivo_saida=None):
            """
            Salva wb1 normalmente e salva wb2 EXCLUSIVAMENTE como 3.xlsx
            na pasta do cliente (ex: WILSON). Nunca salva em FATURAMENTOS.
            """
            try:
                if wb1:
                    wb1.save()
                    wb1.close()

                if wb2:
                    if not arquivo_saida:
                        raise RuntimeError(
                            "Caminho de sa√≠da n√£o informado. "
                            "wb2 N√ÉO ser√° salvo para evitar salvar em FATURAMENTOS."
                        )

                    wb2.save(arquivo_saida)
                    wb2.close()

            finally:
                if app:
                    app.quit()

        # ===== Licen√ßa =====#

def data_online():
            context = ssl.create_default_context(cafile=certifi.where())

            req = urllib.request.Request(
                "https://www.cloudflare.com", headers={"User-Agent": "Mozilla/5.0"}
            )

            with urllib.request.urlopen(req, context=context, timeout=5) as r:
                data_str = r.headers["Date"]

            dt_utc = datetime.strptime(data_str, "%a, %d %b %Y %H:%M:%S %Z").replace(
                tzinfo=timezone.utc
            )

            dt_local = dt_utc.astimezone()
            return dt_utc, dt_local

def validar_licenca():
            hoje_utc, hoje_local = data_online()

            limite = datetime(hoje_utc.year, hoje_utc.month, 30, tzinfo=timezone.utc)
            if hoje_utc > limite:
                sys.exit("‚õî Licen√ßa expirada")

            print(f"üìÖ Data local: {hoje_local.date()}")

def data_por_extenso(valor):
            if isinstance(valor, datetime):
                data = valor

            elif isinstance(valor, date):
                data = datetime(valor.year, valor.month, valor.day)

            elif isinstance(valor, str):
                try:
                    data = datetime.strptime(valor, "%d/%m/%Y")
                except:
                    return ""  # n√£o inventa data

            else:
                return ""  # nunca usa datetime.now()

            return data.strftime("%d de %B de %Y")

# ==============================
# CLASSE 1: FATURAMENTO COMPLETO
# ==============================

class FaturamentoCompleto:

    def executar(self):
        print("üöÄ Faturamento Completo iniciado")
        pasta_faturamentos = obter_pasta_faturamentos()
        resultado = abrir_workbooks(pasta_faturamentos)

        def processar_front(ws1, ws_front):
            """
            Atualiza somente a aba FRONT VIGIA
            """

            # data atual por extenso (rodap√©)
            meses = [
                "",
                "janeiro",
                "fevereiro",
                "mar√ßo",
                "abril",
                "maio",
                "junho",
                "julho",
                "agosto",
                "setembro",
                "outubro",
                "novembro",
                "dezembro",
            ]

            hoje = datetime.now()
            ws_front.range("C39").value = (
                f"Santos, {hoje.day} de {meses[hoje.month]} de {hoje.year}"
            )

            # pega datas extremas do RESUMO
            data_min, data_max = obter_datas_extremos(ws1)

            # mostra no FRONT
            if data_min:
                ws_front.range("D16").value = data_por_extenso(data_min)

            if data_max:
                ws_front.range("D17").value = data_por_extenso(data_max)

            # üëâ retorna as datas para o main
            return data_min, data_max

        # ===== Fun√ß√µes REPORT =====#

        def inserir_linhas_report(ws_report, linha_inicial, periodos):
            """
            Insere linhas copiando a linha inicial para acomodar periodos > 1
            """
            if periodos <= 1:
                return

            row_height = ws_report.api.Rows(linha_inicial).RowHeight

            for i in range(periodos - 1):
                destino = linha_inicial + 1 + i
                ws_report.api.Rows(destino).Insert()
                ws_report.api.Rows(linha_inicial).Copy(ws_report.api.Rows(destino))
                ws_report.api.Rows(destino).RowHeight = row_height

        # ===== COLUNA E ===== #

        def obter_periodos(ws_resumo):
            """
            L√™ a coluna AA da aba Resumo usando xlwings
            (sem pandas, sem conflito de arquivo)
            """
            valores = ws_resumo.range("AA:AA").value

            # Remove None
            valores = [v for v in valores if v is not None]

            try:
                ultimo = str(valores[-1]).replace("R$", "").replace(",", ".").strip()
                return int(float(ultimo))
            except:
                return 1

        def gerar_coluna_E_ajustada(ws1, periodos, coluna_horario="C"):
            """
            Gera a lista de ciclos para preencher a coluna E do REPORT VIGIA.

            - Se C3 for 06h, 12h, 18h ou 00h, come√ßa a lista por este ciclo.
            - Se C3 for "Total" ou vazio, assume primeiro ciclo 00x06 e continua a sequ√™ncia normal.
            - Repete a sequ√™ncia at√© completar 'periodos'.
            """
            # Mapear hor√°rios para ciclos
            horario_para_ciclo = {
                "06h": "06x12",
                "06H": "06x12",
                "12h": "12x18",
                "12H": "12x18",
                "18h": "18x24",
                "18H": "18x24",
                "00h": "00x06",
                "00H": "00x06",
            }

            # Sequ√™ncia padr√£o completa
            sequencia_padrao = ["06x12", "12x18", "18x24", "00x06"]

            # Ler primeira e segunda c√©lulas da coluna
            primeiro_horario = str(ws1.range(f"{coluna_horario}2").value).strip()
            segundo_valor = ws1.range(f"{coluna_horario}3").value
            segundo_valor = (
                str(segundo_valor).strip() if segundo_valor is not None else ""
            )

            # Determinar primeiro ciclo
            if (
                segundo_valor.lower() == "total"
                or segundo_valor not in horario_para_ciclo
            ):
                primeiro_ciclo = "00x06"  # primeiro ciclo √© sempre 00x06
            else:
                primeiro_ciclo = horario_para_ciclo[segundo_valor]

            # Rotacionar sequ√™ncia padr√£o para iniciar pelo primeiro ciclo
            idx_inicio = sequencia_padrao.index(primeiro_ciclo)
            sequencia = sequencia_padrao[idx_inicio:] + sequencia_padrao[:idx_inicio]

            # Gerar lista final at√© completar periodos
            ciclos_linha = []
            for c in cycle(sequencia):
                if len(ciclos_linha) >= periodos:
                    break
                ciclos_linha.append(c)

            return ciclos_linha

        def preencher_coluna_E_por_ciclos(ws_report, ciclos_linha, linha_inicial=22):
            for i, ciclo in enumerate(ciclos_linha):
                ws_report.range(f"E{linha_inicial + i}").value = ciclo

        # ===== COLUNA G ===== #

        def mapear_valores_por_ciclo(ws1, coluna_horario="C", coluna_valor="Z"):
            """
            L√™ os valores do wb1 e agrupa por ciclo.
            ws1 : planilha do arquivo 1
            coluna_horario : coluna que cont√©m os hor√°rios (06h, 12h, etc.)
            coluna_valor : coluna que cont√©m os valores a preencher
            """
            horario_para_ciclo = {
                "06h": "06x12",
                "12h": "12x18",
                "18h": "18x24",
                "00h": "00x06",
            }
            sequencia_ciclos = ["06x12", "12x18", "18x24", "00x06"]

            last_row = ws1.used_range.last_cell.row
            horarios = ws1.range(f"{coluna_horario}1:{coluna_horario}{last_row}").value
            valores = ws1.range(f"{coluna_valor}1:{coluna_valor}{last_row}").value

            # Normaliza hor√°rios para min√∫sculo
            horarios = [
                str(h).strip().lower() if h is not None else None for h in horarios
            ]

            valores_por_ciclo = {c: [] for c in sequencia_ciclos}

            for h, v in zip(horarios, valores):
                if h in horario_para_ciclo:
                    ciclo = horario_para_ciclo[h]
                    valores_por_ciclo[ciclo].append(v)

            return valores_por_ciclo

        def preencher_coluna_G_por_ciclo(
            ws_report, ciclos_linha, valores_por_ciclo, coluna="G", linha_inicial=22
        ):
            """
            Preenche a coluna G do REPORT VIGIA alinhando os valores da coluna Z
            √† sequ√™ncia de ciclos j√° definida na coluna E.
            """
            indices_ciclo = {c: 0 for c in valores_por_ciclo}

            for i, ciclo_val in enumerate(ciclos_linha):
                linha = linha_inicial + i
                lista_valores = valores_por_ciclo.get(ciclo_val, [])
                idx = indices_ciclo[ciclo_val]

                valor = lista_valores[idx] if idx < len(lista_valores) else None
                indices_ciclo[ciclo_val] += 1

                cel = ws_report.range(f"{coluna}{linha}")
                cel.value = valor

                # Formata√ß√£o
                try:
                    cel.api.NumberFormat = "R$ #.##0,00"
                    cel.api.HorizontalAlignment = xw.constants.HAlign.xlHAlignRight
                    cel.api.VerticalAlignment = xw.constants.VAlign.xlVAlignCenter
                    cel.api.Font.Name = "Calibri"
                    cel.api.Font.Size = 18
                except:
                    pass

            return len(ciclos_linha)

        # ===== COLUNA C ===== #

        def montar_datas_report_vigia(
            ws_report, ws_resumo, linha_inicial=22, periodos=None
        ):
            """
            Preenche a coluna C (DATE) do REPORT VIGIA.
            - O dia s√≥ avan√ßa quando o ciclo da coluna E for 00x06.
            - Mant√©m a sequ√™ncia correta independentemente do primeiro hor√°rio.
            """
            if periodos is None:
                raise ValueError(
                    "√â necess√°rio informar 'periodos' para preencher as datas"
                )

            data_inicio, data_fim = obter_datas_extremos(ws_resumo)
            if not data_inicio or not data_fim:
                raise ValueError(
                    "N√£o foi poss√≠vel determinar as datas extremas na aba RESUMO"
                )

            data_atual = data_inicio

            for i in range(periodos):
                linha = linha_inicial + i
                ciclo = ws_report.range(f"E{linha}").value

                if ciclo in (None, ""):
                    break

                # Coloca a data atual na coluna C
                ws_report.range(f"C{linha}").value = data_atual

                # Se o ciclo for 00x06, incrementa o dia para a pr√≥xima linha
                if isinstance(ciclo, str) and ciclo.strip().lower() == "00x06":
                    data_atual += timedelta(days=1)

            return periodos

        # ===== DATA INICIAL E FINAL DO FRONT =====#

        MESES_EN = {
            1: "JAN",
            2: "FEB",
            3: "MAR",
            4: "APR",
            5: "MAY",
            6: "JUN",
            7: "JUL",
            8: "AUG",
            9: "SEP",
            10: "OCT",
            11: "NOV",
            12: "DEC",
        }

        def obter_datas_extremos(ws_resumo):
            last_row = ws_resumo.used_range.last_cell.row
            valores = ws_resumo.range(f"B1:B{last_row}").value

            datas = []
            hoje = date.today()

            for v in valores:
                if v in (None, "", "Total"):
                    continue

                # datetime vindo do Excel
                if isinstance(v, datetime):
                    d = v.date()

                    # üö´ ignora f√≥rmulas HOJE()
                    if d == hoje:
                        continue

                    datas.append(d)
                    continue

                # string
                if isinstance(v, str):
                    v = v.strip().lower()

                    # 19/10/2025
                    try:
                        datas.append(datetime.strptime(v, "%d/%m/%Y").date())
                        continue
                    except:
                        pass

                    # 19/out/25
                    try:
                        dia, mes_txt, ano = v.split("/")
                        mes = MESES_EN.get(int(mes_txt))
                        if mes:
                            ano = int(ano)
                            if ano < 100:
                                ano += 2000
                            datas.append(date(ano, mes, int(dia)))
                    except:
                        pass

            if not datas:
                return None, None

            return min(datas), max(datas)

        # ===== ABAS ESPECIFICAS =====#

        def OC(wb1, wb2):
            ws = wb2.sheets["FRONT VIGIA"]
            if str(ws["G16"].value).strip().upper() == "O.C.:":
                ws["H16"].value = input("OC: ")

        def credit_note(wb, valor_c21):
            if "Credit Note" in [s.name for s in wb.sheets]:
                wb.sheets["Credit Note"]["C21"].value = valor_c21

        def obter_proxima_nf(pasta_nfs):
            if not os.path.exists(pasta_nfs):
                print(f"‚ö†Ô∏è Pasta de NFs n√£o encontrada:\n{pasta_nfs}")
                return 1

            numeros = []

            for arquivo in os.listdir(pasta_nfs):
                nome, ext = os.path.splitext(arquivo)

                # ignora arquivos que n√£o sejam PDF (se quiser)
                if ext.lower() != ".pdf":
                    continue

                # captura apenas o n√∫mero inicial do nome
                match = re.match(r"(\d+)", nome)
                if match:
                    numeros.append(int(match.group(1)))

            if not numeros:
                return 1

            return max(numeros) + 1

        def colar_nf(ws, celula, numero_nf):
            ws[celula].value = f"NF.: {numero_nf}"

        def MMO(wb1, wb2):
            """
            wb1 = NAVIO (tem 'Resumo')
            wb2 = CLIENTE (tem 'REPORT VIGIA')
            """

            print("   Iniciando MMO...")

            # --- REPORT VIGIA (destino) ---
            try:
                ws_report = wb2.sheets["REPORT VIGIA"]
            except:
                print("   ‚ö†Ô∏è Aba 'REPORT VIGIA' n√£o encontrada. Pulando MMO.")
                return

            if str(ws_report["E25"].value).strip().upper() != "MMO":
                print("   MMO n√£o necess√°rio (E25 != 'MMO').")
                return

            # --- RESUMO (origem) ---
            try:
                ws_resumo = wb1.sheets["Resumo"]
            except:
                print("   ‚ö†Ô∏è Aba 'Resumo' n√£o encontrada no NAVIO. Pulando MMO.")
                return

            print("   Lendo coluna G do Resumo...")

            valores_g = ws_resumo.range("G1:G1000").value
            valores_limpos = [v for v in valores_g if v not in (None, "")]

            if not valores_limpos:
                print("   Coluna G vazia. Pulando MMO.")
                return

            ultimo_valor = valores_limpos[-1]

            try:
                texto = str(ultimo_valor)
                texto = texto.replace("R$", "").replace(" ", "")
                texto = texto.replace(".", "").replace(",", ".")
                ultimo_float = float(texto)
            except Exception as e:
                print(f"   Erro ao converter '{ultimo_valor}': {e}")
                ultimo_float = 0.0

            ws_report["F25"].value = ultimo_float
            ws_report["F25"].number_format = "#,##0.00"

            print(f"   ‚úÖ MMO conclu√≠do: R$ {ultimo_float:,.2f} escrito em F25")

        def cargonave(ws):
            valor_c9 = ws.range("C9").value
            return (
                str(valor_c9).strip().upper() == "A/C AG√äNCIA MAR√çTIMA CARGONAVE LTDA."
            )

        def arredondar_para_baixo_50(ws_front_vigia):
            if not cargonave(ws_front_vigia):
                return
            valor = ws_front_vigia.range("E37").value
            if valor is None:
                return
            try:
                resultado = (int(valor) // 50) * 50
            except:
                return
            ws_front_vigia.range("H28").value = resultado

        def obter_nome_navio_da_pasta(caminho_arquivo):
            """
            Ex: '123 - NAVIO' -> 'NAVIO'
            """
            pasta = os.path.basename(os.path.dirname(caminho_arquivo))

            if "-" in pasta:
                return pasta.split("-", 1)[1].strip()

            return pasta.strip()

        def obter_aba_nf_opcional(wb):
            for sheet in wb.sheets:
                nome = sheet.name.strip().lower()
                if nome == "nf" or nome.startswith("nf") or "nota" in nome:
                    return sheet
            return None

        def escrever_nf(wb_faturamento, nome_navio, dn):
            # tenta localizar aba NF
            ws_nf = None
            for sheet in wb_faturamento.sheets:
                if sheet.name.strip().lower() == "nf":
                    ws_nf = sheet
                    break

            if ws_nf is None:
                print("‚ö†Ô∏è Aba NF n√£o encontrada ‚Äî seguindo sem escrever NF")
                return  # N√ÉO quebra o programa

            ano = datetime.now().year

            texto = (
                f"SERVI√áO PRESTADO DE ATENDIMENTO/APOIO AO M/V {nome_navio}\n"
                f"DN {dn}/{ano}"
            )

            # escreve na primeira c√©lula
            cel = ws_nf.range("A1")
            cel.value = texto

            # mescla para ficar bonito
            ws_nf.range("A1:E2").merge()

            # formata√ß√£o
            cel.api.HorizontalAlignment = -4108  # center
            cel.api.VerticalAlignment = -4108
            cel.api.WrapText = True
            cel.api.Font.Name = "Calibri"
            cel.api.Font.Size = 14
            cel.api.Font.Bold = True

            print("‚úÖ Texto da NF escrito com sucesso")

        def quitacao(wb, valor_c21):
            if "Quita√ß√£o" not in [s.name for s in wb.sheets]:
                return

            ws = wb.sheets["Quita√ß√£o"]

            # Colar valor financeiro
            ws["C22"].value = valor_c21

            # Caminho real das NFs
            pasta_nfs = r"C:\Users\Carol\SANPORT LOG√çSTICA PORTU√ÅRIA LTDA\Central de Documentos - Documentos\2.2 CONTABILIDADE 2025\12 - DEZEMBRO"

            proxima_nf = obter_proxima_nf(pasta_nfs)

            # Colar NF
            colar_nf(ws, "H22", proxima_nf)

        def main():
            print("üöÄ Iniciando execu√ß√£o...")

            # ========= 1 ‚Äì Licen√ßa =========
            validar_licenca()

            # ========= 2 ‚Äì Localizar FATURAMENTOS =========
            pasta_faturamentos = obter_pasta_faturamentos()

            # ========= 3 ‚Äì Abrir arquivos =========
            resultado = abrir_workbooks(pasta_faturamentos)
            if not resultado:
                sys.exit("‚ùå Usu√°rio cancelou ou erro ao abrir workbooks")

            app, wb1, wb2, ws1, ws_front = resultado

            try:
                # ========= 4 ‚Äì DN e Navio =========
                dn = obter_dn_da_pasta(wb1.fullname)
                if not dn:
                    sys.exit("‚ùå DN n√£o identificada pela pasta")

                nome_navio = obter_nome_navio_da_pasta(wb1.fullname)
                ano_atual = datetime.now().year
                texto_dn = f"DN: {dn}/{ano_atual}"

                # Preenchimento FRONT VIGIA
                ws_front.range("D15").value = nome_navio
                ws_front.range("C21").value = texto_dn

                berco = input("WAREHOUSE / BER√áO: ").strip().upper()
                ws_front["D18"].value = berco

                # ========= 4 ‚Äì Processar FRONT VIGIA =========
                print("‚öôÔ∏è Processando FRONT VIGIA...")
                data_inicio, data_fim = processar_front(ws1, ws_front)

                if not data_inicio or not data_fim:
                    sys.exit("‚ùå Datas extremas inv√°lidas no RESUMO")

                ws_resumo = wb1.sheets["Resumo"]
                # ========= 5 ‚Äì MMO =========
                print("‚öôÔ∏è Processando MMO...")

                app.api.Calculate()
                time.sleep(0.5)

                MMO(wb1, wb2)  # ou se voc√™ mudou para MMO(wb1, wb2), deixa assim
                # ========= 6 ‚Äì NF =========
                escrever_nf(wb2, nome_navio, dn)

                # ========= 7 ‚Äì REPORT VIGIA =========
                print("‚öôÔ∏è Processando REPORT VIGIA...")
                periodos = obter_periodos(ws_resumo)

                ws_report = wb2.sheets["REPORT VIGIA"]

                # Inserir linhas extras se necess√°rio
                inserir_linhas_report(ws_report, linha_inicial=22, periodos=periodos)

                # Coluna E - Ciclos
                ciclos_linha = gerar_coluna_E_ajustada(
                    ws1, periodos, coluna_horario="C"
                )
                preencher_coluna_E_por_ciclos(ws_report, ciclos_linha, linha_inicial=22)

                # Coluna G - Valores
                valores_por_ciclo = mapear_valores_por_ciclo(
                    ws1, coluna_horario="C", coluna_valor="Z"
                )
                preencher_coluna_G_por_ciclo(
                    ws_report,
                    ciclos_linha,
                    valores_por_ciclo,
                    coluna="G",
                    linha_inicial=22,
                )

                # Coluna C - Datas (respeitando ciclos 00x06)
                montar_datas_report_vigia(
                    ws_report=ws_report,
                    ws_resumo=ws_resumo,
                    linha_inicial=22,
                    periodos=periodos,
                )

                # ========= 8 ‚Äì Financeiro =========
                print("‚öôÔ∏è Processando Financeiro...")
                OC(str(wb1.fullname), wb2)
                credit_note(wb2, texto_dn)
                quitacao(wb2, texto_dn)  # descomentado se precisar

                # ========= 9 ‚Äì Ajustes finais =========
                print("‚öôÔ∏è Aplicando ajustes finais...")
                arredondar_para_baixo_50(ws_front)
                cargonave(ws_front)

                # ========= 10 ‚Äì Salvar e fechar =========
                pasta_saida = Path(wb1.fullname).parent
                arquivo_saida = pasta_saida / "3.xlsx"

                fechar_workbooks(app, wb1, wb2, arquivo_saida)

                print(f"‚úÖ Processo conclu√≠do com sucesso!")
                print(f"   Arquivo salvo em: {arquivo_saida}")

            except Exception as e:
                print(f"‚ùå Erro durante o processamento: {e}")
                # Garante que o Excel feche mesmo em caso de erro
                try:
                    if wb1:
                        wb1.close()
                    if wb2:
                        wb2.close()
                    if app:
                        app.quit()
                except:
                    pass
                sys.exit(1)

# ==============================
# CLASSE 2: FATURAMENTO DE ACORDO (simples)
# ==============================

class FaturamentoCompleto:
    
    def executar(self):
        print("üöÄ Faturamento Completo iniciado")
        pasta_faturamentos = obter_pasta_faturamentos()
        resultado = abrir_workbooks(pasta_faturamentos)

        try:
            # Usa exatamente a mesma abertura da op√ß√£o completa
            resultado = abrir_workbooks(obter_pasta_faturamentos())

            if not resultado:
                print("‚ùå Opera√ß√£o cancelada.")
                return

            app, wb1, wb2, ws1, ws_front = resultado

            try:
                # Extrai DN e navio
                pasta_navio = Path(wb1.fullname).parent
                dn = obter_dn_da_pasta(pasta_navio)
                nome_navio = obter_nome_navio_da_pasta(pasta_navio)
                print(f"üìã DN: {dn}")
                print(f"üö¢ Navio: {nome_navio}")

                # Preenche FRONT VIGIA
                self._preencher_front_vigia(ws_front, dn, nome_navio)

                # Remove outras abas
                nome_aba = ws_front.name
                for sheet in list(wb2.sheets):
                    if sheet.name != nome_aba:
                        sheet.delete()

                # Salva na √Årea de Trabalho
                desktop = Path.home() / "Desktop"
                arquivo_excel = desktop / f"3 - DN_{dn}.xlsx"
                arquivo_pdf = desktop / f"3 - DN_{dn}.pdf"

                wb2.save(str(arquivo_excel))
                print(f"üìÑ Excel salvo: {arquivo_excel.name}")

                print("üìÑ Gerando PDF...")
                with ws_front.api.PageSetup:
                    ws_front.api.PageSetup.Zoom = False
                    ws_front.api.PageSetup.FitToPagesWide = 1
                    ws_front.api.PageSetup.FitToPagesTall = 1

                ws_front.api.ExportAsFixedFormat(
                    Type=0,
                    Filename=str(arquivo_pdf),
                    Quality=0,
                    IncludeDocProperties=True,
                    IgnorePrintAreas=False,
                    OpenAfterPublish=False
                )
                print(f"üìÑ PDF salvo: {arquivo_pdf.name}")

                print(f"\n‚úÖ CONCLU√çDO COM SUCESSO!")
                print(f"Arquivos na √Årea de Trabalho:")
                print(f"   üìä {arquivo_excel.name}")
                print(f"   üìÑ {arquivo_pdf.name}")

            finally:
                try:
                    wb1.close()
                    wb2.close()
                    app.quit()
                except:
                    pass

        except Exception as e:
            print(f"\n‚ùå ERRO: {e}")

        def _preencher_front_vigia(self, ws_front, dn, nome_navio):
            hoje = datetime.now()
            meses = ["", "janeiro", "fevereiro", "mar√ßo", "abril", "maio", "junho",
                    "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"]
            data_extenso = f"{hoje.day} de {meses[hoje.month]} de {hoje.year}"
            texto_dn = f"DN: {dn}/{hoje.year}"

            ws_front.range("D15").value = nome_navio
            ws_front.range("C21").value = texto_dn
            ws_front.range("D16").value = data_extenso
            ws_front.range("D17").value = data_extenso
            ws_front.range("D18").value = "-"
            ws_front.range("C26").value = f'  DE ACORDO ( M/V "{nome_navio}" )'
            ws_front.range("C27").value = '  VOY SA02325'
            ws_front.range("G27").value = None

            cliente_c9 = str(ws_front.range("C9").value or "").strip()
            if "Unimar Agenciamentos" in cliente_c9:
                ws_front.range("G26").value = 400

            try:
                valor = float(ws_front.range("C35").value or 0)
                ws_front.range("C35").value = valor + 20
            except:
                ws_front.range("C35").value = 20

            ws_front.range("C39").value = f"Santos, {data_extenso}"

            print("‚úÖ FRONT VIGIA preenchida!")

# ==============================
# CLASSE 3: MENU PRINCIPAL
# ==============================

class CentralSanport:
   
    def __init__(self):
        self.de_acordo = FaturamentoDeAcordo()
        self.completo = FaturamentoCompleto()

    @staticmethod
    def mostrar_menu():
        print("=" * 60)
        print("     üö¢ CENTRAL DE FATURAMENTO - SANPORT")
        print("     Bem-vindo(a)!")
        print("=" * 60)
        print("1 - Faturamento Completo")
        print("2 - Faturamento De Acordo")
        print("0 - Sair")
        print("-" * 60)

    def rodar(self):
        while True:
            self.mostrar_menu()
            opcao = input("\nDigite a op√ß√£o (1, 2 ou 0): ").strip()

            if opcao == "1":
                print("\nüöÄ Iniciando Faturamento Completo...\n")
                try:
                    self.completo.executar()
                except Exception as e:
                    print(f"\n‚ùå Erro: {e}")
                finally:
                    print("\n‚èé Voltando ao menu...")
                    input("Pressione ENTER para continuar...")

            elif opcao == "2":
                print("\nüöÄ Iniciando Faturamento De Acordo...\n")
                try:
                    self.de_acordo.executar()
                except Exception as e:
                    print(f"\n‚ùå Erro: {e}")
                finally:
                    print("\n‚èé Voltando ao menu...")
                    input("Pressione ENTER para continuar...")

            elif opcao == "0":
                print("\nüëã Saindo. At√© logo!\n")
                break
            else:
                print("‚ùå Op√ß√£o inv√°lida.\n")

# ==============================
# EXECU√á√ÉO
# ==============================

if __name__ == "__main__":
    CentralSanport().rodar()