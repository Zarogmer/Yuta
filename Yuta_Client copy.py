# 1️⃣ – Importações
import xlwings as xw
from datetime import datetime, timedelta, timezone
from openpyxl.styles import Alignment
import pandas as pd
import os
import sys
import urllib.request
import ssl
import certifi
from pathlib import Path
import tkinter as tk
from datetime import date
import re
import locale
from itertools import cycle
from tkinter import Tk, filedialog
import shutil
import tempfile

# =========================
# Funções utilitárias
# =========================


def copiar_para_temp_e_ler_excel(caminho_original: Path | str) -> pd.DataFrame:
    """
    Copia o arquivo para pasta temporária local e lê com pandas.
    Resolve a maioria dos PermissionError em pastas OneDrive/rede.
    """
    caminho_original = Path(caminho_original)
    if not caminho_original.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {caminho_original}")

    with tempfile.TemporaryDirectory() as temp_dir:
        temp_dir_path = Path(temp_dir)
        caminho_temp = temp_dir_path / caminho_original.name

        print(f"Copiando {caminho_original.name} para pasta temporária local...")
        shutil.copy2(caminho_original, caminho_temp)

        print(f"Lendo arquivo temporário: {caminho_temp}")
        df = pd.read_excel(caminho_temp, engine="openpyxl")  # engine explícito ajuda

    return df

def abrir_workbooks():
    """
    Abre:
    - wb_navio   -> arquivo 1.xlsx (fonte de dados)
    - wb_cliente -> arquivo 2.xlsx (template - somente leitura)
    - wb_3       -> arquivo 3.xlsx (cópia do 2, onde tudo será escrito)
    """

    root = tk.Tk()
    root.withdraw()

    # ========= NAVIO =========
    pasta_navio_str = filedialog.askdirectory(
        title="Selecione a pasta do NAVIO (onde está o 1.xlsx)"
    )
    if not pasta_navio_str:
        return None

    pasta_navio = Path(pasta_navio_str)
    pasta_cliente = pasta_navio.parent
    nome_cliente = pasta_cliente.name.strip()

    arquivos_navio = list(pasta_navio.glob("1*.xls*"))
    if not arquivos_navio:
        raise FileNotFoundError("Arquivo 1.xlsx não encontrado.")

    arquivo_1 = arquivos_navio[0]

    # ========= FATURAMENTOS =========
    atual = pasta_cliente
    pasta_faturamentos = None

    while atual.parent != atual:
        if atual.name == "01. FATURAMENTOS" and (atual / "FATURAMENTOS").exists():
            pasta_faturamentos = atual / "FATURAMENTOS"
            break
        atual = atual.parent

    if not pasta_faturamentos:
        raise FileNotFoundError("Pasta FATURAMENTOS não encontrada.")

    arquivo_2 = pasta_faturamentos / f"{nome_cliente}.xlsx"
    if not arquivo_2.exists():
        raise FileNotFoundError(f"Arquivo cliente não encontrado: {arquivo_2}")

    # ========= TEMP =========
    temp_dir = tempfile.TemporaryDirectory()
    temp_path = Path(temp_dir.name)

    arquivo_1_temp = temp_path / arquivo_1.name
    arquivo_2_temp = temp_path / arquivo_2.name
    arquivo_3_temp = temp_path / "3.xlsx"

    shutil.copy2(arquivo_1, arquivo_1_temp)
    shutil.copy2(arquivo_2, arquivo_2_temp)
    shutil.copy2(arquivo_2, arquivo_3_temp)  # 3 = cópia do 2

    # ========= XLWINGS =========
    app = xw.App(visible=False)

    try:
        wb_navio   = app.books.open(str(arquivo_1_temp))
        wb_cliente = app.books.open(str(arquivo_2_temp))  # leitura
        wb_3       = app.books.open(str(arquivo_3_temp))  # escrita

        # sheets principais
        ws_navio = wb_navio.sheets[0]

        if nome_cliente in [s.name for s in wb_3.sheets]:
            ws_front_3 = wb_3.sheets[nome_cliente]
        elif "FRONT VIGIA" in [s.name for s in wb_3.sheets]:
            ws_front_3 = wb_3.sheets["FRONT VIGIA"]
        else:
            ws_front_3 = wb_3.sheets.add("FRONT VIGIA")

        # metadados
        wb_3.pasta_navio = pasta_navio
        wb_3.temp_dir = temp_dir

        return app, wb_navio, wb_cliente, wb_3, ws_navio, ws_front_3

    except Exception as e:
        app.quit()
        temp_dir.cleanup()
        raise e


def fechar_workbooks(app, wb_navio, wb_cliente, wb_3):
    try:
        if wb_3:
            arquivo_saida = wb_3.pasta_navio / "3.xlsx"
            wb_3.save(str(arquivo_saida))
            wb_3.close()

        if wb_navio:
            wb_navio.close()

        if wb_cliente:
            wb_cliente.close()

    finally:
        if app:
            app.quit()










def obter_pasta_faturamentos() -> Path:
    """
    Localiza automaticamente a pasta FATURAMENTOS dentro da estrutura OneDrive da SANPORT.
    Funciona mesmo com espaços e hífens no nome da pasta.
    """
    print("\n=== BUSCANDO PASTA FATURAMENTOS AUTOMATICAMENTE ===")

    # Possíveis locais base onde o OneDrive sincroniza a pasta da empresa
    possiveis_bases = [
        Path(r"C:\Users\Carol\SANPORT LOGÍSTICA PORTUÁRIA LTDA"),
        Path(r"C:\Users\Carol\OneDrive - SANPORT LOGÍSTICA PORTUÁRIA LTDA"),  # caso seja OneDrive pessoal
        Path.home() / "SANPORT LOGÍSTICA PORTUÁRIA LTDA",
        Path.home() / "OneDrive" / "SANPORT LOGÍSTICA PORTUÁRIA LTDA",
    ]

    caminho_alvo = None

    for base in possiveis_bases:
        if base.exists():
            print(f"✅ Encontrada pasta base: {base}")

            # Procurar recursivamente por uma pasta chamada "FATURAMENTOS" dentro de "01. FATURAMENTOS"
            candidatos = list(base.rglob("FATURAMENTOS"))
            for candidato in candidatos:
                # Filtrar para garantir que está dentro de "01. FATURAMENTOS"
                if "01. FATURAMENTOS" in candidato.parent.as_posix():
                    caminho_alvo = candidato
                    print(f"✅ Pasta FATURAMENTOS encontrada em:\n   {caminho_alvo}")
                    break

            if caminho_alvo:
                break
        else:
            print(f"❌ Não encontrada: {base}")

    if not caminho_alvo:
        print("❌ Pasta FATURAMENTOS não foi encontrada automaticamente.")
        print("\nPossíveis soluções:")
        print("• Verifique se o OneDrive está sincronizando a pasta da empresa")
        print("• Clique com botão direito na pasta FATURAMENTOS → Propriedades → Localização")
        print("  e me diga o caminho exato que aparece")
        raise FileNotFoundError("Pasta FATURAMENTOS não localizada automaticamente")

    # Debug final: listar alguns arquivos para confirmar
    print(f"\nArquivos .xlsx encontrados na pasta ({len(list(caminho_alvo.glob('*.xlsx')))}):")
    for arq in sorted(caminho_alvo.glob("*.xlsx"))[:10]:  # mostra só os 10 primeiros
        print(f"   • {arq.name}")
    if len(list(caminho_alvo.glob("*.xlsx"))) > 10:
        print("   ... (mais arquivos)")

    print("========================================\n")
    return caminho_alvo



    """
    Salva wb1 normalmente e salva wb2 EXCLUSIVAMENTE como 3.xlsx
    na pasta do cliente (ex: WILSON). Nunca salva em FATURAMENTOS.
    """
    try:
        if wb1:
            wb1.save()
            wb1.close()

        if wb2:
            if not arquivo_saida:
                raise RuntimeError(
                    "Caminho de saída não informado. "
                    "wb2 NÃO será salvo para evitar salvar em FATURAMENTOS."
                )

            wb2.save(arquivo_saida)
            wb2.close()

    finally:
        if app:
            app.quit()

def obter_dn_da_pasta(caminho_arquivo):
    """
    Extrai números do nome da pasta do navio
    Ex: '123 - NAVIO' -> '123'
    """
    pasta = os.path.basename(os.path.dirname(caminho_arquivo))

    numeros = re.findall(r"\d+", pasta)

    if not numeros:
        return None

    return numeros[0]  # primeiro bloco numérico


# ===== Licença =====#


def data_online():
    context = ssl.create_default_context(cafile=certifi.where())

    req = urllib.request.Request(
        "https://www.cloudflare.com",
        headers={"User-Agent": "Mozilla/5.0"}
    )

    with urllib.request.urlopen(req, context=context, timeout=5) as r:
        data_str = r.headers["Date"]

    dt_utc = datetime.strptime(
        data_str, "%a, %d %b %Y %H:%M:%S %Z"
    ).replace(tzinfo=timezone.utc)

    dt_local = dt_utc.astimezone()
    return dt_utc, dt_local


def validar_licenca():
    hoje_utc, hoje_local = data_online()

    limite = datetime(hoje_utc.year, hoje_utc.month, 30, tzinfo=timezone.utc)
    if hoje_utc > limite:
        sys.exit("⛔ Licença expirada")

    print(f"📅 Data local: {hoje_local.date()}")


def data_por_extenso(valor):
    if isinstance(valor, datetime):
        data = valor

    elif isinstance(valor, date):
        data = datetime(valor.year, valor.month, valor.day)

    elif isinstance(valor, str):
        try:
            data = datetime.strptime(valor, "%d/%m/%Y")
        except:
            return ""  # não inventa data

    else:
        return ""  # nunca usa datetime.now()

    return data.strftime("%d de %B de %Y")

def MMO(arquivo1, wb2):
    ws = wb2.sheets["REPORT VIGIA"]
    if str(ws["E25"].value).strip().upper() != "MMO": return
    df = pd.read_excel(arquivo1, sheet_name="Resumo", header=None)
    col_g = df[6].dropna()
    if col_g.empty: return
    try:
        ultimo_float = locale.atof(str(col_g.iloc[-1]).replace("R$", "").strip())
    except: ultimo_float = float(col_g.iloc[-1])
    ws["F25"].value = ultimo_float
    ws["F25"].number_format = "#.##0,00"


def processar_front(ws1, ws_front):
    """
    Atualiza somente a aba FRONT VIGIA
    """

    # data atual por extenso (rodapé)
    meses = ["", "janeiro", "fevereiro", "março", "abril", "maio", "junho",
             "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"]

    hoje = datetime.now()
    ws_front.range("C39").value = (
        f"Santos, {hoje.day} de {meses[hoje.month]} de {hoje.year}"
    )

    # pega datas extremas do RESUMO
    data_min, data_max = obter_datas_extremos(ws1)

    # mostra no FRONT
    if data_min:
        ws_front.range("D16").value = data_por_extenso(data_min)

    if data_max:
        ws_front.range("D17").value = data_por_extenso(data_max)

    # 👉 retorna as datas para o main
    return data_min, data_max


# ===== Funções REPORT =====#

def inserir_linhas_report(ws_report, linha_inicial, periodos):
    """
    Insere linhas copiando a linha inicial para acomodar periodos > 1
    """
    if periodos <= 1:
        return

    row_height = ws_report.api.Rows(linha_inicial).RowHeight

    for i in range(periodos - 1):
        destino = linha_inicial + 1 + i
        ws_report.api.Rows(destino).Insert()
        ws_report.api.Rows(linha_inicial).Copy(ws_report.api.Rows(destino))
        ws_report.api.Rows(destino).RowHeight = row_height

# ===== COLUNA E ===== #


def obter_periodos(ws_resumo):
    """
    Lê a coluna AA da aba Resumo usando xlwings
    (sem pandas, sem conflito de arquivo)
    """
    valores = ws_resumo.range("AA:AA").value

    # Remove None
    valores = [v for v in valores if v is not None]

    try:
        ultimo = str(valores[-1]).replace("R$", "").replace(",", ".").strip()
        return int(float(ultimo))
    except:
        return 1



def gerar_coluna_E_ajustada(ws1, periodos, coluna_horario="C"):
    """
    Gera a lista de ciclos para preencher a coluna E do REPORT VIGIA.
    
    - Se C3 for 06h, 12h, 18h ou 00h, começa a lista por este ciclo.
    - Se C3 for "Total" ou vazio, assume primeiro ciclo 00x06 e continua a sequência normal.
    - Repete a sequência até completar 'periodos'.
    """
    # Mapear horários para ciclos
    horario_para_ciclo = {
        "06h": "06x12", "06H": "06x12",
        "12h": "12x18", "12H": "12x18",
        "18h": "18x24", "18H": "18x24",
        "00h": "00x06", "00H": "00x06"
    }

    # Sequência padrão completa
    sequencia_padrao = ["06x12", "12x18", "18x24", "00x06"]

    # Ler primeira e segunda células da coluna
    primeiro_horario = str(ws1.range(f"{coluna_horario}2").value).strip()
    segundo_valor = ws1.range(f"{coluna_horario}3").value
    segundo_valor = str(segundo_valor).strip() if segundo_valor is not None else ""

    # Determinar primeiro ciclo
    if segundo_valor.lower() == "total" or segundo_valor not in horario_para_ciclo:
        primeiro_ciclo = "00x06"  # primeiro ciclo é sempre 00x06
    else:
        primeiro_ciclo = horario_para_ciclo[segundo_valor]

    # Rotacionar sequência padrão para iniciar pelo primeiro ciclo
    idx_inicio = sequencia_padrao.index(primeiro_ciclo)
    sequencia = sequencia_padrao[idx_inicio:] + sequencia_padrao[:idx_inicio]

    # Gerar lista final até completar periodos
    ciclos_linha = []
    for c in cycle(sequencia):
        if len(ciclos_linha) >= periodos:
            break
        ciclos_linha.append(c)

    return ciclos_linha



def preencher_coluna_E_por_ciclos(ws_report, ciclos_linha, linha_inicial=22):
    for i, ciclo in enumerate(ciclos_linha):
        ws_report.range(f"E{linha_inicial + i}").value = ciclo


 
# ===== COLUNA G ===== #


def mapear_valores_por_ciclo(ws1, coluna_horario="C", coluna_valor="Z"):
    """
    Lê os valores do wb1 e agrupa por ciclo.
    ws1 : planilha do arquivo 1
    coluna_horario : coluna que contém os horários (06h, 12h, etc.)
    coluna_valor : coluna que contém os valores a preencher
    """
    horario_para_ciclo = {"06h":"06x12", "12h":"12x18", "18h":"18x24", "00h":"00x06"}
    sequencia_ciclos = ["06x12", "12x18", "18x24", "00x06"]

    last_row = ws1.used_range.last_cell.row
    horarios = ws1.range(f"{coluna_horario}1:{coluna_horario}{last_row}").value
    valores = ws1.range(f"{coluna_valor}1:{coluna_valor}{last_row}").value

    # Normaliza horários para minúsculo
    horarios = [str(h).strip().lower() if h is not None else None for h in horarios]

    valores_por_ciclo = {c: [] for c in sequencia_ciclos}

    for h, v in zip(horarios, valores):
        if h in horario_para_ciclo:
            ciclo = horario_para_ciclo[h]
            valores_por_ciclo[ciclo].append(v)

    return valores_por_ciclo

def preencher_coluna_G_por_ciclo(ws_report, ciclos_linha, valores_por_ciclo, coluna="G", linha_inicial=22):
    """
    Preenche a coluna G do REPORT VIGIA alinhando os valores da coluna Z
    à sequência de ciclos já definida na coluna E.
    """
    indices_ciclo = {c: 0 for c in valores_por_ciclo}

    for i, ciclo_val in enumerate(ciclos_linha):
        linha = linha_inicial + i
        lista_valores = valores_por_ciclo.get(ciclo_val, [])
        idx = indices_ciclo[ciclo_val]

        valor = lista_valores[idx] if idx < len(lista_valores) else None
        indices_ciclo[ciclo_val] += 1

        cel = ws_report.range(f"{coluna}{linha}")
        cel.value = valor

        # Formatação
        try:
            cel.api.NumberFormat = 'R$ #.##0,00'
            cel.api.HorizontalAlignment = xw.constants.HAlign.xlHAlignRight
            cel.api.VerticalAlignment = xw.constants.VAlign.xlVAlignCenter
            cel.api.Font.Name = "Calibri"
            cel.api.Font.Size = 18
        except:
            pass

    return len(ciclos_linha)

# ===== COLUNA C ===== #

def montar_datas_report_vigia(ws_report, ws_resumo, linha_inicial=22, periodos=None):
    """
    Preenche a coluna C (DATE) do REPORT VIGIA.
    - O dia só avança quando o ciclo da coluna E for 00x06.
    - Mantém a sequência correta independentemente do primeiro horário.
    """
    if periodos is None:
        raise ValueError("É necessário informar 'periodos' para preencher as datas")

    data_inicio, data_fim = obter_datas_extremos(ws_resumo)
    if not data_inicio or not data_fim:
        raise ValueError("Não foi possível determinar as datas extremas na aba RESUMO")

    data_atual = data_inicio

    for i in range(periodos):
        linha = linha_inicial + i
        ciclo = ws_report.range(f"E{linha}").value

        if ciclo in (None, ""):
            break

        # Coloca a data atual na coluna C
        ws_report.range(f"C{linha}").value = data_atual

        # Se o ciclo for 00x06, incrementa o dia para a próxima linha
        if isinstance(ciclo, str) and ciclo.strip().lower() == "00x06":
            data_atual += timedelta(days=1)

    return periodos


# ===== DATA INICIAL E FINAL DO FRONT =====#

MESES_EN = {
    1: "JAN", 2: "FEB", 3: "MAR", 4: "APR",
    5: "MAY", 6: "JUN", 7: "JUL", 8: "AUG",
    9: "SEP", 10: "OCT", 11: "NOV", 12: "DEC"
}

def obter_datas_extremos(ws_resumo):
    last_row = ws_resumo.used_range.last_cell.row
    valores = ws_resumo.range(f"B1:B{last_row}").value

    datas = []
    hoje = date.today()

    for v in valores:
        if v in (None, "", "Total"):
            continue

        # datetime vindo do Excel
        if isinstance(v, datetime):
            d = v.date()

            # 🚫 ignora fórmulas HOJE()
            if d == hoje:
                continue

            datas.append(d)
            continue

        # string
        if isinstance(v, str):
            v = v.strip().lower()

            # 19/10/2025
            try:
                datas.append(datetime.strptime(v, "%d/%m/%Y").date())
                continue
            except:
                pass

            # 19/out/25
            try:
                dia, mes_txt, ano = v.split("/")
                mes = MESES_EN.get(int(mes_txt))
                if mes:
                    ano = int(ano)
                    if ano < 100:
                        ano += 2000
                    datas.append(date(ano, mes, int(dia)))
            except:
                pass

    if not datas:
        return None, None

    return min(datas), max(datas)


# ===== ABAS ESPECIFICAS =====#

def MMO(arquivo1, wb2):
    ws = wb2.sheets["REPORT VIGIA"]
    if str(ws["E25"].value).strip().upper() != "MMO": return
    df = pd.read_excel(arquivo1, sheet_name="Resumo", header=None)
    col_g = df[6].dropna()
    if col_g.empty: return
    try:
        ultimo_float = locale.atof(str(col_g.iloc[-1]).replace("R$", "").strip())
    except: ultimo_float = float(col_g.iloc[-1])
    ws["F25"].value = ultimo_float
    ws["F25"].number_format = "#.##0,00"

def OC(arquivo1, wb2):
    ws = wb2.sheets["FRONT VIGIA"]
    if str(ws["G16"].value).strip().upper() == "O.C.:":
        ws["H16"].value = input("OC: ")

def credit_note(wb, valor_c21):
    if "Credit Note" in [s.name for s in wb.sheets]:
        wb.sheets["Credit Note"]["C21"].value = valor_c21

def quitacao(wb, valor_c21):
    """
    Preenche a aba Quitação com:
    - valor_c21 em C22
    - próximo número de NF baseado nos PDFs da pasta do mês atual
    Usa wb1.pasta_navio (já guardada em abrir_workbooks) para encontrar:
    Central de Documentos - Documentos > 2.2 CONTABILIDADE 2025 > 12 - DEZEMBRO
    Ignora arquivos com 'CANCELADA', 'CANCELADO', etc.
    """
    if "Quitação" not in [s.name for s in wb.sheets]:
        print("⚠️ Aba 'Quitação' não encontrada. Pulando quitacao().")
        return

    ws = wb.sheets["Quitação"]
    ws["C22"].value = valor_c21

    # ================ PEGAR PASTA DO NAVIO GUARDADA NO WB1 ================
    try:
        # wb pode ser wb2, mas wb1 tem a pasta_navio salva
        import sys
        current_module = sys.modules[__name__]
        if hasattr(current_module, 'wb1') and hasattr(current_module.wb1, 'pasta_navio'):
            pasta_navio = current_module.wb1.pasta_navio
        elif 'wb1' in globals() and hasattr(globals()['wb1'], 'pasta_navio'):
            pasta_navio = globals()['wb1'].pasta_navio
        else:
            raise AttributeError("pasta_navio não encontrada")
    except:
        print("❌ Não conseguiu acessar a pasta do navio. Usando NF.: 1 como fallback.")
        ws["H22"].value = "NF.: 1"
        return

    print(f"Usando pasta do navio para localizar contabilidade:\n   {pasta_navio}")

    # ================ SUBIR ATÉ A RAIZ DA EMPRESA ================
    raiz_empresa = None
    for pai in pasta_navio.parents:
        if pai.name == "SANPORT LOGÍSTICA PORTUÁRIA LTDA":
            raiz_empresa = pai
            break

    if not raiz_empresa:
        print("❌ Raiz 'SANPORT LOGÍSTICA PORTUÁRIA LTDA' não encontrada.")
        ws["H22"].value = "NF.: 1"
        return

    print(f"✅ Raiz da empresa encontrada: {raiz_empresa}")

    # ================ MONTAR CAMINHO EXATO DA CONTABILIDADE ================
    ano = datetime.now().year  # 2025
    mes_num = datetime.now().month  # 12 = Dezembro
    nomes_meses = ["JANEIRO", "FEVEREIRO", "MARÇO", "ABRIL", "MAIO", "JUNHO",
                   "JULHO", "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO", "DEZEMBRO"]
    nome_mes = nomes_meses[mes_num - 1]

    pasta_pdfs = raiz_empresa / "Central de Documentos - Documentos" / f"2.2 CONTABILIDADE {ano}" / f"{mes_num:02d} - {nome_mes}"

    if not pasta_pdfs.exists():
        print(f"⚠️ Pasta do mês não encontrada:\n   {pasta_pdfs}")
        print("   Criando automaticamente para futuro uso...")
        pasta_pdfs.mkdir(parents=True, exist_ok=True)
        ws["H22"].value = "NF.: 1"
        print("✅ Primeira NF do mês → NF.: 1")
        return

    print(f"🔍 Lendo PDFs em:\n   {pasta_pdfs}")

    # ================ ENCONTRAR O MAIOR NÚMERO VÁLIDO ================
    numeros = []

    for arquivo in pasta_pdfs.glob("*.pdf"):
        nome = arquivo.stem.upper()

        # Ignora arquivos cancelados
        if any(term in nome for term in ["CANCELADA", "CANCELADO", "CANCELAD"]):
            continue

        # Extrai o número inicial (ex: 7749.pdf → 7749)
        parte = nome.split()[0].split("-")[0].split("_")[0].split("(")[0].strip()
        if parte.isdigit():
            numeros.append(int(parte))

    if not numeros:
        proximo = 1
        print("   Nenhum PDF válido encontrado → NF.: 1")
    else:
        proximo = max(numeros) + 1
        print(f"   Maior número encontrado: {max(numeros)} → próximo = {proximo}")
        print(f"   Total de PDFs válidos: {len(numeros)}")

    ws["H22"].value = f"NF.: {proximo}"
    print(f"✅ Quitação preenchida → NF.: {proximo}")

def cargonave(ws):
    valor_c9 = ws.range("C9").value
    return str(valor_c9).strip().upper() == "A/C AGÊNCIA MARÍTIMA CARGONAVE LTDA."

def arredondar_para_baixo_50(ws_front_vigia):
    if not cargonave(ws_front_vigia): return
    valor = ws_front_vigia.range("E37").value
    if valor is None: return
    try: resultado = (int(valor)//50)*50
    except: return
    ws_front_vigia.range("H28").value = resultado

def obter_nome_navio_da_pasta(caminho_arquivo):
    """
    Ex: '123 - NAVIO' -> 'NAVIO'
    """
    pasta = os.path.basename(os.path.dirname(caminho_arquivo))

    if "-" in pasta:
        return pasta.split("-", 1)[1].strip()

    return pasta.strip()

def obter_aba_nf_opcional(wb):
    for sheet in wb.sheets:
        nome = sheet.name.strip().lower()
        if nome == "nf" or nome.startswith("nf") or "nota" in nome:
            return sheet
    return None

def escrever_nf(wb_faturamento, nome_navio, dn):
    # tenta localizar aba NF
    ws_nf = None
    for sheet in wb_faturamento.sheets:
        if sheet.name.strip().lower() == "nf":
            ws_nf = sheet
            break

    if ws_nf is None:
        print("⚠️ Aba NF não encontrada — seguindo sem escrever NF")
        return  # NÃO quebra o programa

    ano = datetime.now().year

    texto = (
        f"SERVIÇO PRESTADO DE ATENDIMENTO/APOIO AO M/V {nome_navio}\n"
        f"DN {dn}/{ano}"
    )

    # escreve na primeira célula
    cel = ws_nf.range("A1")
    cel.value = texto

    # mescla para ficar bonito
    ws_nf.range("A1:E2").merge()

    # formatação
    cel.api.HorizontalAlignment = -4108  # center
    cel.api.VerticalAlignment = -4108
    cel.api.WrapText = True
    cel.api.Font.Name = "Calibri"
    cel.api.Font.Size = 14
    cel.api.Font.Bold = True

    print("✅ Texto da NF escrito com sucesso")


def main():
    print("🚀 Iniciando execução...")

    validar_licenca()

    resultado = abrir_workbooks()
    if not resultado:
        sys.exit("❌ Operação cancelada.")

    app, wb_navio, wb_cliente, wb_3, ws_navio, ws_front = resultado

    print("📂 Workbooks abertos com sucesso!")

    try:
        # ======================================================
        # DADOS BÁSICOS (lê SOMENTE do arquivo 1)
        # ======================================================
        dn = obter_dn_da_pasta(wb_navio.fullname)
        if not dn:
            sys.exit("❌ DN não identificada.")

        nome_navio = obter_nome_navio_da_pasta(wb_navio.fullname)
        ano_atual = datetime.now().year
        texto_dn = f"DN: {dn}/{ano_atual}"

        berco = input("WAREHOUSE / BERÇO: ").strip().upper()

        # ======================================================
        # FRONT VIGIA (escreve SOMENTE no arquivo 3)
        # ======================================================
        print("⚙️ Preenchendo FRONT VIGIA...")

        ws_front.range("D15").value = nome_navio
        ws_front.range("C21").value = texto_dn
        ws_front.range("D18").value = berco

        # Lê do NAVIO, escreve no 3
        data_inicio, data_fim = processar_front(
            ws_navio=ws_navio,
            ws_front=ws_front
        )

        # ======================================================
        # MMO
        # ======================================================
        print("⚙️ Processando MMO...")
        MMO(wb_3.fullname, wb_3)

        # ======================================================
        # NOTAS FISCAIS
        # ======================================================
        escrever_nf(wb_3, nome_navio, dn)

        # ======================================================
        # REPORT VIGIA
        # ======================================================
        print("⚙️ Processando REPORT VIGIA...")

        try:
            ws_report = wb_3.sheets["REPORT VIGIA"]
        except:
            ws_report = wb_3.sheets.add("REPORT VIGIA")

        ws_resumo = wb_3.sheets["Resumo"]

        periodos = obter_periodos(ws_resumo)

        inserir_linhas_report(
            ws_report,
            linha_inicial=22,
            periodos=periodos
        )

        ciclos_linha = gerar_coluna_E_ajustada(
            ws_navio,
            periodos,
            coluna_horario="C"
        )

        preencher_coluna_E_por_ciclos(
            ws_report,
            ciclos_linha,
            linha_inicial=22
        )

        valores_por_ciclo = mapear_valores_por_ciclo(
            ws_navio,
            coluna_horario="C",
            coluna_valor="Z"
        )

        preencher_coluna_G_por_ciclo(
            ws_report,
            ciclos_linha,
            valores_por_ciclo,
            coluna="G",
            linha_inicial=22
        )

        montar_datas_report_vigia(
            ws_report=ws_report,
            ws_resumo=ws_resumo,
            linha_inicial=22,
            periodos=periodos
        )

        # ======================================================
        # FINANCEIRO
        # ======================================================
        print("⚙️ Processando Financeiro...")

        OC(wb_3.fullname, wb_3)
        credit_note(wb_3, texto_dn)
        quitacao(wb_3, texto_dn)

        # ======================================================
        # AJUSTES FINAIS
        # ======================================================
        print("⚙️ Ajustes finais...")

        arredondar_para_baixo_50(ws_front)
        cargonave(ws_front)

        # ======================================================
        # SALVAR 3.XLSX
        # ======================================================
        print("💾 Salvando 3.xlsx na pasta do NAVIO...")

        arquivo_saida = wb_3.pasta_navio / "3.xlsx"
        wb_3.save(str(arquivo_saida))

        print(f"✅ 3.xlsx gerado com sucesso:\n   {arquivo_saida}")
        print("✅ FATURAMENTOS permaneceu intacto.")

    except Exception as e:
        print(f"❌ Erro durante execução: {e}")
        raise

    finally:
        fechar_workbooks(app, wb_navio, wb_cliente, wb_3)
        print("🎉 Processo concluído com sucesso!")


if __name__ == "__main__":
    main()
