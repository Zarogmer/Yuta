# ==============================
# IMPORTS # ==============================
import sys
import re
import ssl
import certifi
import urllib.request
import shutil
import tempfile
import time
import os
import msvcrt

from pathlib import Path
from itertools import cycle
from datetime import datetime, date, timedelta, timezone

import tkinter as tk
from tkinter import Tk, filedialog

import pandas as pd
import xlwings as xw
import openpyxl
from copy import copy  # para copiar estilos
from openpyxl.styles import Font

import holidays

# instância de feriados do Brasil
feriados_br = holidays.Brazil()

# adicionar feriados personalizados
feriados_personalizados = [
    date(2025, 1, 1),
    date(2025, 4, 21),
    date(2025, 5, 1),
    # ... outros feriados locais
]

for d in feriados_personalizados:
    feriados_br[d] = "Feriado personalizado"



# ==============================
# FUNÇÕES AUXILIARES GLOBAIS
# ==============================



# ---------------------------
# 1️⃣ Copiar arquivo para pasta temporária e ler Excel
# ---------------------------
def copiar_para_temp_e_ler_excel(caminho_original: Path | str) -> pd.DataFrame:
    caminho_original = Path(caminho_original)
    if not caminho_original.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {caminho_original}")

    with tempfile.TemporaryDirectory() as temp_dir:
        caminho_temp = Path(temp_dir) / caminho_original.name
        print(f"Copiando {caminho_original.name} para pasta temporária local...")
        shutil.copy2(caminho_original, caminho_temp)
        print(f"Lendo arquivo temporário: {caminho_temp}")
        df = pd.read_excel(caminho_temp, engine="openpyxl")
    return df

# ---------------------------
# 2️⃣ Localizar pasta FATURAMENTOS automaticamente
# ---------------------------
def obter_pasta_faturamentos() -> Path:
    print("\n=== BUSCANDO PASTA FATURAMENTOS AUTOMATICAMENTE ===")
    possiveis_bases = [
        Path(r"C:\Users\Carol\SANPORT LOGÍSTICA PORTUÁRIA LTDA"),
        Path(r"C:\Users\Carol\OneDrive - SANPORT LOGÍSTICA PORTUÁRIA LTDA"),
        Path.home() / "SANPORT LOGÍSTICA PORTUÁRIA LTDA",
        Path.home() / "OneDrive" / "SANPORT LOGÍSTICA PORTUÁRIA LTDA",
    ]

    caminho_alvo = None
    for base in possiveis_bases:
        if base.exists():
            print(f"✅ Encontrada pasta base: {base}")
            candidatos = list(base.rglob("FATURAMENTOS"))
            for candidato in candidatos:
                if "01. FATURAMENTOS" in candidato.parent.as_posix():
                    caminho_alvo = candidato
                    print(f"✅ Pasta FATURAMENTOS encontrada em:\n   {caminho_alvo}")
                    break
            if caminho_alvo:
                break
        else:
            print(f"❌ Não encontrada: {base}")

    if not caminho_alvo:
        raise FileNotFoundError("Pasta FATURAMENTOS não localizada automaticamente")

    arquivos_xlsx = list(caminho_alvo.glob("*.xlsx"))
    print(f"\nArquivos .xlsx encontrados na pasta ({len(arquivos_xlsx)}):")
    for arq in sorted(arquivos_xlsx)[:10]:
        print(f"   • {arq.name}")
    if len(arquivos_xlsx) > 10:
        print("   ... (mais arquivos)")
    print("========================================\n")
    return caminho_alvo

# ---------------------------
# 3️⃣ Abrir workbooks NAVIO e cliente com xlwings
# ---------------------------
def abrir_workbooks(pasta_faturamentos: Path):
    root = tk.Tk()
    root.withdraw()

    pasta_navio_str = filedialog.askdirectory(title="Selecione a pasta do NAVIO (onde está o 1.xlsx)")
    if not pasta_navio_str:
        print("Seleção cancelada pelo usuário.")
        return None, None, None, None, None

    pasta_navio = Path(pasta_navio_str)
    pasta_cliente = pasta_navio.parent
    nome_cliente = pasta_cliente.name.strip()

    arquivos_1 = list(pasta_navio.glob("1*.xls*"))
    if not arquivos_1:
        raise FileNotFoundError(f"Nenhum arquivo iniciando com '1' encontrado em:\n{pasta_navio}")

    arquivo1 = arquivos_1[0]
    arquivo2 = pasta_faturamentos / f"{nome_cliente}.xlsx"
    if not arquivo2.exists():
        raise FileNotFoundError(f"Arquivo de faturamento não encontrado:\n{arquivo2}")

    # abrir com xlwings
    app = xw.App(visible=False)
    wb1 = wb2 = None
    try:
        wb1 = app.books.open(str(arquivo1))
        wb2 = app.books.open(str(arquivo2))

        ws1 = wb1.sheets[0]
        nomes_abas = [s.name for s in wb2.sheets]

        if nome_cliente in nomes_abas:
            ws_front = wb2.sheets[nome_cliente]
        elif "FRONT VIGIA" in nomes_abas:
            ws_front = wb2.sheets["FRONT VIGIA"]
        else:
            raise RuntimeError(f"Nenhuma aba válida encontrada em {arquivo2}")

        return app, wb1, wb2, ws1, ws_front
    except Exception as e:
        if wb1: wb1.close()
        if wb2: wb2.close()
        if app: app.quit()
        raise e


def obter_dn_da_pasta(pasta: Path) -> str:
    """
    Extrai o número DN do nome da pasta.
    Se não encontrar, retorna '0000' e exibe aviso.
    """
    numeros = re.findall(r"\d+", pasta.name)
    if not numeros:
        print(
            f"⚠️ Não foi possível identificar o DN no nome da pasta '{pasta.name}', usando '0000' como padrão"
        )
        return "0000"
    return numeros[0]


def obter_nome_navio_da_pasta(pasta: Path) -> str:
    nome_limpo = re.sub(r"^\d+[\s\-_]*", "", pasta.name, flags=re.IGNORECASE).strip()
    return nome_limpo if nome_limpo else "NAVIO NÃO IDENTIFICADO"


def fechar_workbooks(app=None, wb_navio=None, wb_cliente=None, arquivo_saida: Path | None = None):
    """
    Fecha e salva workbooks de forma defensiva.
    - app: instância xlwings.App (ou None)
    - wb_navio: workbook do navio (opcional)
    - wb_cliente: workbook do cliente (opcional) -> salvo em arquivo_saida se fornecido
    - arquivo_saida: Path onde salvar o Excel e gerar PDF (opcional)
    """
    # salva/expor e fecha com muitos try/except para evitar exceções RPC
    try:
        if wb_cliente and arquivo_saida:
            try:
                # tenta salvar o workbook do cliente
                wb_cliente.save(str(arquivo_saida))
            except Exception as e:
                print(f"⚠️ Erro ao salvar Excel: {e}")

            try:
                # tenta exportar para PDF (pode falhar se Excel estiver em estado inválido)
                pdf_saida = arquivo_saida.with_suffix(".pdf")
                wb_cliente.api.ExportAsFixedFormat(0, str(pdf_saida))
            except Exception as e:
                print(f"⚠️ Erro ao exportar PDF: {e}")

            try:
                print(f"💾 Arquivo Excel salvo em: {arquivo_saida}")
                print(f"📑 PDF gerado em: {pdf_saida}")
            except Exception:
                pass

        # fecha workbooks individualmente, cada um em try/except
        try:
            if wb_navio:
                wb_navio.close()
        except Exception as e:
            print(f"⚠️ Erro ao fechar wb_navio: {e}")

        try:
            if wb_cliente:
                wb_cliente.close()
        except Exception as e:
            print(f"⚠️ Erro ao fechar wb_cliente: {e}")

    finally:
        # encerra a app somente se ela existir e parecer válida
        try:
            if app:
                # proteção extra: algumas versões de xlwings/COM podem lançar se já encerrado
                app.quit()
        except Exception as e:
            print(f"⚠️ Erro ao encerrar app Excel: {e}")


def selecionar_arquivo_navio() -> str | None:
    """
    Abre uma janela para o usuário selecionar o arquivo NAVIO (.xlsx).
    A janela aparece em primeiro plano.
    """
    root = Tk()
    root.lift()                         # coloca a janela na frente
    root.attributes("-topmost", True)   # força ficar em primeiro plano
    root.focus_force()                  # força foco
    root.update()                       # aplica imediatamente
    caminho = filedialog.askopenfilename(
        title="Selecione o arquivo do NAVIO",
        filetypes=[("Arquivos Excel", "*.xlsx")]
    )
    root.destroy()  # fecha a janela principal depois da seleção

    if not caminho:
        return None

    print(f"📂 Arquivo NAVIO selecionado: {Path(caminho).name}")
    return caminho


# ==============================
# LICENÇA E DATA
# ==============================


def data_online():
    context = ssl.create_default_context(cafile=certifi.where())
    req = urllib.request.Request(
        "https://www.cloudflare.com", headers={"User-Agent": "Mozilla/5.0"}
    )
    with urllib.request.urlopen(req, context=context, timeout=5) as r:
        data_str = r.headers["Date"]
    dt_utc = datetime.strptime(data_str, "%a, %d %b %Y %H:%M:%S %Z").replace(
        tzinfo=timezone.utc
    )
    dt_local = dt_utc.astimezone()
    return dt_utc, dt_local


def validar_licenca():
    hoje_utc, hoje_local = data_online()

    # 🔥 define uma data fixa de expiração: 5 de janeiro de 2026
    limite = datetime(2026, 1, 5, tzinfo=timezone.utc)

    if hoje_utc > limite:
        sys.exit("⛔ Licença expirada")

    print(f"📅 Data local: {hoje_local.date()}")



# ==============================
# CLASSE 1: FATURAMENTO COMPLETO


class FaturamentoCompleto:
    def __init__(self, g_logic=1):
        self.app = None
        self.wb1 = None
        self.wb2 = None
        self.ws1 = None
        self.ws_front = None
        self.nome_navio = None
        self.g_logic = g_logic


    # ---------------- fluxo principal ----------------


    def executar(self):
        print("🚀 Iniciando execução...")
        pasta_faturamentos = obter_pasta_faturamentos()   # pega a pasta automaticamente
        resultado = abrir_workbooks(pasta_faturamentos)   # passa como argumento
        ...

        if not resultado:
            raise SystemExit("❌ Erro ou pasta inválida")
        self.app, self.wb1, self.wb2, self.ws1, self.ws_front = resultado
        print("📂 Workbooks abertos com sucesso!")
        try:
            self.processar()
        except Exception as e:
            print(f"❌ Erro durante o processamento: {e}")
            # tenta salvar/fechar com segurança usando a função global
            try:
                pasta_saida = Path(self.wb1.fullname).parent if self.wb1 else None
                arquivo_saida = (pasta_saida / "3.xlsx") if pasta_saida else None
                fechar_workbooks(self.app, self.wb1, self.wb2, arquivo_saida)
            except Exception as e2:
                print(f"⚠️ Erro ao fechar após falha: {e2}")
            # re-levanta para o menu tratar ou encerra conforme seu fluxo
            raise


    def processar(self):
        """
        Fluxo principal do faturamento:
        1) Preenche FRONT VIGIA
        2) Atualiza o REPORT VIGIA (E, G e C)
        """

        try:

            # ---------- FRONT ----------
            self.preencher_front_vigia()

            # ---------- REPORT ----------
            ws_report = self.wb2.sheets["REPORT VIGIA"]

            ...
            self.processar_MMO(self.wb1, self.wb2)

            # 1️⃣ quantidade de períodos (APENAS PARA INSERIR LINHAS)
            qtd_periodos = self.obter_periodos(self.ws1)


            self.inserir_linhas_report(
                ws_report,
                linha_inicial=22,
                periodos=qtd_periodos
            )

            # 2️⃣ COLUNA E → LISTA DE PERÍODOS
            periodos = self.preencher_coluna_E(
                ws_report,
                linha_inicial=22,
                debug=True
            )


            # 3️⃣ COLUNA G → VALORES (BASEADO NA E)
            self.preencher_coluna_G(
                ws_report,
                self.ws1,              # ws_resumo
                linha_inicial=22,
                periodos=periodos,     # 🔥 lista
                debug=True
            )

            # 4️⃣ COLUNA C → DATAS (USA QUANTIDADE DA E)
            self.montar_datas_report_vigia(
                ws_report,
                self.ws1,
                linha_inicial=22,
                periodos=len(periodos)  # ⚠️ INT
            )

            self.arredondar_para_baixo_50_se_cargonave(self.ws_front)

            self._OC()

            print("✅ REPORT VIGIA atualizado com sucesso!")


        except Exception as e:
            print(f"❌ Erro ao atualizar REPORT: {e}")
            raise

        # ---------- SALVA E FECHA ----------
        try:
            pasta_saida = Path(self.wb1.fullname).parent if self.wb1 else None
            arquivo_saida = (pasta_saida / "3.xlsx") if pasta_saida else None

            fechar_workbooks(self.app, self.wb1, self.wb2, arquivo_saida)

            print(f"💾 Arquivo Excel salvo em: {arquivo_saida}")

        except Exception as e:
            print(f"⚠️ Erro ao salvar/fechar workbooks: {e}")
            raise




# ===== FRONT ======#


    def preencher_front_vigia(self):
        """
        Preenche a aba FRONT VIGIA com:
        - Nome do navio
        - DN
        - Warehouse / Berço (entrada do usuário)
        - Data inicial e final
        - Rodapé com data atual
        """
        try:
            # Pasta do navio
            pasta_navio = Path(self.wb1.fullname).parent
            dn = obter_dn_da_pasta(pasta_navio)
            self.nome_navio = obter_nome_navio_da_pasta(pasta_navio)
            ano = datetime.now().year
            texto_dn = f"DN: {dn}/{ano}"

            # ---------------- FRONT ----------------
            try:
                # Nome do navio
                self.ws_front.range("D15").value = self.nome_navio
                # DN
                self.ws_front.range("C21").value = texto_dn
                # Warehouse / Berço
                self.ws_front.range("D18").value = input("WAREHOUSE / BERÇO: ").upper()
            except Exception as e:
                print(f"⚠️ Erro ao preencher FRONT (nome/DN/berço): {e}")

            # ---------------- DATAS ----------------
            data_min, data_max = self.obter_datas_extremos(self.ws1)
            try:
                if data_min:
                    self.ws_front.range("D16").value = self.data_por_extenso(data_min)
                if data_max:
                    self.ws_front.range("D17").value = self.data_por_extenso(data_max)
            except Exception as e:
                print(f"⚠️ Erro ao preencher datas no FRONT: {e}")

            # ---------------- RODAPÉ ----------------
            try:
                hoje = datetime.now()
                meses = ["", "janeiro","fevereiro","março","abril","maio","junho",
                        "julho","agosto","setembro","outubro","novembro","dezembro"]
                self.ws_front.range("C39").value = f"Santos, {hoje.day} de {meses[hoje.month]} de {hoje.year}"
            except Exception as e:
                print(f"⚠️ Erro ao preencher rodapé do FRONT: {e}")

            print("✅ FRONT VIGIA preenchido com sucesso!")

        except Exception as e:
            print(f"❌ Erro geral ao preencher FRONT VIGIA: {e}")
            raise


#==================== REPORT =====================#

    def inserir_linhas_report(self, ws_report, linha_inicial, periodos):
        if periodos <= 1:
            return
        app = ws_report.book.app
        app.screen_updating = False
        app.display_alerts = False
        app.enable_events = False
        app.calculation = 'manual'
        try:
            linha_modelo = linha_inicial
            for i in range(periodos - 1):
                destino = linha_inicial + 1 + i
                ws_report.api.Rows(destino).Insert()
                ws_report.api.Rows(linha_modelo).Copy(ws_report.api.Rows(destino))
        finally:
            app.calculation = 'automatic'
            app.enable_events = True
            app.display_alerts = True
            app.screen_updating = True


    # ===== LINHA E=====#


    def gerar_ciclos_coluna_E(self, ws_resumo, linha_inicial=2):
        """
        Gera a lista de períodos para a coluna E do REPORT, baseada na data mais antiga.
        Sequência: 06x12 -> 12x18 -> 18x24 -> 00x06
        """
        sequencia_padrao = ["06x12", "12x18", "18x24", "00x06"]

        # 1️⃣ Encontrar a primeira data válida (mais antiga que hoje)
        last_row = ws_resumo.used_range.last_cell.row
        valores = ws_resumo.range(f"B{linha_inicial}:B{last_row}").value
        hoje = date.today()
        primeira_linha_data = None

        for idx, v in enumerate(valores):
            if v in (None, "", "Total"):
                continue
            if isinstance(v, datetime):
                d = v.date()
            elif isinstance(v, str):
                try:
                    d = datetime.strptime(v.strip(), "%d/%m/%Y").date()
                except:
                    continue
            else:
                continue

            if d < hoje:
                primeira_linha_data = linha_inicial + idx
                break

        if primeira_linha_data is None:
            return []  # nenhuma data antiga encontrada

        # 2️⃣ Contar espaços vazios ou "Total" antes da próxima data
        contador_vazio = 0
        for i in range(primeira_linha_data + 1, last_row + 1):
            valor = ws_resumo.range(f"B{i}").value
            if valor in (None, "", "Total"):
                contador_vazio += 1
            else:
                break

        # 3️⃣ Definir primeiro período
        if contador_vazio >= 4:
            primeiro_periodo = "06x12"
        elif contador_vazio == 3:
            primeiro_periodo = "12x18"
        elif contador_vazio == 2:
            primeiro_periodo = "18x24"
        else:
            primeiro_periodo = "00x06"

        # 4️⃣ Sequência cíclica
        idx_inicio = sequencia_padrao.index(primeiro_periodo)
        sequencia_ciclica = sequencia_padrao[idx_inicio:] + sequencia_padrao[:idx_inicio]

        # 5️⃣ Gerar lista completa de períodos
        total_periodos = self.obter_periodos(ws_resumo)
        ciclos = [sequencia_ciclica[i % 4] for i in range(total_periodos)]

        return ciclos


    
    def preencher_coluna_E(self, ws_report, linha_inicial=22, debug=False):
        """
        Preenche a coluna E do REPORT VIGIA com os períodos gerados.
        """
        try:
            ciclos = self.gerar_ciclos_coluna_E(self.ws1)
            for idx, p in enumerate(ciclos):
                ws_report.range(f"E{linha_inicial + idx}").value = p
            if debug:

                return ciclos
        except Exception as e:
            print(f"❌ Erro ao preencher coluna E: {e}")
            raise

 

    # ===== LINHA G=====#


    def normalizar_periodo(self, valor_c):
        if not valor_c:
            return None

        s = str(valor_c).strip().lower()
        if s.startswith("06"):
            return "06x12"
        if s.startswith("12"):
            return "12x18"
        if s.startswith("18"):
            return "18x24"
        if s.startswith("00"):
            return "00x06"
        return None
        

    def gerar_valores_coluna_G(self, ws_resumo, periodos_E, debug=False):
        mapa = self.extrair_valores_por_periodo(ws_resumo, debug=debug)
        contadores = {k: 0 for k in mapa}
        valores_g = []

        for p in periodos_E:
            if p in mapa and contadores[p] < len(mapa[p]):
                valor = mapa[p][contadores[p]]
                contadores[p] += 1
            else:
                valor = 0.0

            valores_g.append(valor)

        if debug:


         return valores_g


    def preencher_coluna_G(self, ws_report, ws_resumo, linha_inicial=22, periodos=None, debug=False):
        """
        Preenche a coluna G seguindo EXATAMENTE a ordem da coluna E
        e formatando como moeda com 2 casas decimais.
        """

        if not periodos:
            raise ValueError("periodos (lista da coluna E) é obrigatório")

        valores = self.gerar_valores_coluna_G(
            ws_resumo,
            periodos,
            debug=debug
        )


        for i, valor in enumerate(valores):
                cell = ws_report.range(f"G{linha_inicial + i}")
                cell.value = valor              # número cru, sem arredondar
                cell.api.NumberFormatLocal = 'R$ #.##0,00'


    def extrair_valores_por_periodo(self, ws_resumo, debug=False):
        last_row = ws_resumo.used_range.last_cell.row

        mapa = {
            "00x06": [],
            "06x12": [],
            "12x18": [],
            "18x24": []
        }

        for i in range(2, last_row + 1):
            c = ws_resumo.range(f"C{i}").value
            z = ws_resumo.range(f"Z{i}").value

            if not c or z is None:
                continue

            s_c = str(c).strip().lower()
            if s_c.startswith("total"):
                continue

            periodo = self.normalizar_periodo(s_c)
            if not periodo:
                continue

            try:
                # ✅ CASO 1: já é número no Excel
                if isinstance(z, (int, float)):
                    valor = float(z)

                # ✅ CASO 2: veio como texto "R$ 1.144,70"
                else:
                    valor = (
                        str(z)
                        .replace("R$", "")
                        .replace(".", "")
                        .replace(",", ".")
                        .strip()
                    )
                    valor = float(valor)

            except:
                continue


            mapa[periodo].append(valor)
            

        return mapa
    

    def extrair_numero_excel(self, z):
        """
        Garante conversão correta de valores do Excel
        independente de vir como float ou string pt-BR.
        """

        # 👉 Caso 1: Excel já entregou número
        if isinstance(z, (int, float)):
            return float(z)

        # 👉 Caso 2: Veio como texto (ex: "1.144,70")
        s = str(z).strip()

        if not s:
            raise ValueError("valor vazio")

        s = (
            s.replace("R$", "")
            .replace(" ", "")
            .replace(".", "")
            .replace(",", ".")
        )

        return float(s)


    # ===== LINHA C =====#


    def montar_datas_report_vigia(self, ws_report, ws_resumo, linha_inicial=22, periodos=None):
        if periodos is None:
            raise ValueError("É necessário informar 'periodos' para preencher as datas")
        data_inicio, _ = self.obter_datas_extremos(ws_resumo)
        if not data_inicio:
            raise ValueError("Não foi possível determinar a data inicial na aba RESUMO")
        data_atual = data_inicio
        for i in range(periodos):
            linha = linha_inicial + i
            ciclo = ws_report.range(f"E{linha}").value
            if ciclo in (None, ""):
                break
            ws_report.range(f"C{linha}").value = data_atual
            if isinstance(ciclo, str) and ciclo.strip().lower() == "00x06":
                data_atual += timedelta(days=1)
        return periodos


    # ===== ABAS ESPECIFICAS =====#

    def _OC(self):
        ws = self.wb2.sheets["FRONT VIGIA"]

        if str(ws["G16"].value).strip().upper() == "O.C.:":
            ws["H16"].value = input("OC: ")


    def processar_MMO(self, wb_navio, wb_cliente):
        """
        MMO:
        - LÊ: último valor da coluna G da aba 'Resumo' (NAVIO)
        - ESCREVE: F25 da aba 'REPORT VIGIA' (CLIENTE)
        """

        print("   🔹 Iniciando MMO...")

        # ---------- REPORT VIGIA (CLIENTE) ----------
        try:
            ws_report = wb_cliente.sheets["REPORT VIGIA"]
        except:
            print("   ⚠️ Aba 'REPORT VIGIA' não encontrada no CLIENTE. MMO ignorado.")
            return

        if str(ws_report.range("E25").value).strip().upper() != "MMO":
            print("   ℹ️ MMO não aplicável (E25 != 'MMO').")
            return

        # ---------- RESUMO (NAVIO) ----------
        try:
            ws_resumo = wb_navio.sheets["Resumo"]
        except:
            print("   ⚠️ Aba 'Resumo' não encontrada no NAVIO. MMO ignorado.")
            return

        # ---------- LÊ COLUNA G ----------
        valores = ws_resumo.range("G1:G1000").value
        valores_validos = [v for v in valores if v not in (None, "")]

        if not valores_validos:
            print("   ℹ️ Coluna G vazia no Resumo. MMO ignorado.")
            return

        ultimo_valor = valores_validos[-1]

        # ---------- CONVERSÃO CORRETA (IGUAL COLUNA G) ----------
        try:
            valor_float = float(ultimo_valor)
        except:
            print(f"   ⚠️ Valor inválido '{ultimo_valor}'. Usando 0.")
            valor_float = 0.0

        # 🔥 correção de escala (quando vem gigante)
        if valor_float > 1_000_000:
            valor_float = valor_float / 100

        # ---------- ESCREVE ----------
        celula = ws_report.range("F25")
        celula.value = valor_float
        celula.api.NumberFormatLocal = 'R$ #.##0,00'

        print(f"   ✅ MMO concluído → R$ {valor_float:,.2f}")


    def arredondar_para_baixo_50_se_cargonave(self, ws_front_vigia):
        """
        Aplica arredondamento para baixo em múltiplos de 50
        Somente para A/C AGÊNCIA MARÍTIMA CARGONAVE LTDA.
        """

        valor_empresa = ws_front_vigia.range("C9").value
        if not valor_empresa:
            return

        if str(valor_empresa).strip().upper() != "A/C AGÊNCIA MARÍTIMA CARGONAVE LTDA.":
            return

        valor = ws_front_vigia.range("E37").value
        if valor is None:
            return

        try:
            resultado = (int(valor) // 50) * 50
        except (ValueError, TypeError):
            return

        ws_front_vigia.range("H28").value = resultado


# ===== DATAS / UTILITARIOS =====#


    def data_por_extenso(self, valor):
        if isinstance(valor, datetime):
            data = valor
        elif isinstance(valor, date):
            data = datetime(valor.year, valor.month, valor.day)
        elif isinstance(valor, str):
            try:
                data = datetime.strptime(valor, "%d/%m/%Y")
            except:
                return ""
        else:
            return ""
        return data.strftime("%d de %B de %Y")


    def obter_datas_extremos(self, ws_resumo):
        last_row = ws_resumo.used_range.last_cell.row
        valores = ws_resumo.range(f"B1:B{last_row}").value
        datas = []
        hoje = date.today()
        for v in valores:
            if v in (None, "", "Total"):
                continue
            if isinstance(v, datetime):
                d = v.date()
                if d == hoje:
                    continue
                datas.append(d)
                continue
            if isinstance(v, str):
                s = v.strip()
                try:
                    d = datetime.strptime(s, "%d/%m/%Y").date()
                    if d != hoje:
                        datas.append(d)
                    continue
                except:
                    pass
        if not datas:
            return None, None
        return min(datas), max(datas)


    def obter_periodos(self, ws_resumo):
        """
        Retorna o último valor válido da coluna AA como inteiro.
        """
        last_row = ws_resumo.used_range.last_cell.row
        # Lê toda a coluna AA
        valores = ws_resumo.range(f"AA1:AA{last_row}").value

        if not valores:
            return 1  # padrão

        # Garante que 'valores' seja uma lista
        if not isinstance(valores, list):
            valores = [valores]

        # Procura o último valor não vazio
        for v in reversed(valores):
            if v is not None and v != "":
                try:
                    return int(float(v))
                except:
                    continue

        return 1



# ==============================
# CLASSE 2: FATURAMENTO DE ACORDO
# ==============================


class FaturamentoDeAcordo:


    def executar(self):
        print("🚀 Iniciando execução...")
        pasta_faturamentos = obter_pasta_faturamentos()   # pega a pasta automaticamente
        resultado = abrir_workbooks(pasta_faturamentos)   # passa como argumento

        if not resultado:
            raise SystemExit("❌ Erro ao abrir workbooks")

        app, wb1, wb2, ws1, ws_front = resultado

        try:
            pasta_navio = Path(wb1.fullname).parent
            dn = obter_dn_da_pasta(pasta_navio)
            nome_navio = obter_nome_navio_da_pasta(pasta_navio)

            print(f"📋 DN: {dn}")
            print(f"🚢 Navio: {nome_navio}")

            hoje = datetime.now()
            meses = [
                "", "janeiro","fevereiro","março","abril","maio","junho",
                "julho","agosto","setembro","outubro","novembro","dezembro"
            ]
            data_extenso = f"{hoje.day} de {meses[hoje.month]} de {hoje.year}"

            ws_front.range("D15").value = nome_navio
            ws_front.range("C21").value = f"DN: {dn}/{hoje.year}"
            ws_front.range("D16").value = data_extenso
            ws_front.range("D17").value = data_extenso
            ws_front.range("D18").value = "-"
            ws_front.range("C26").value = f'  DE ACORDO ( M/V "{nome_navio}" )'
            ws_front.range("C27").value = "  VOY SA02325"
            ws_front.range("G27").value = None

            cliente_c9 = str(ws_front.range("C9").value or "").strip()
            if "Unimar Agenciamentos" in cliente_c9:
                ws_front.range("G26").value = 400

            try:
                valor = float(ws_front.range("C35").value or 0)
                ws_front.range("C35").value = valor + 20
            except:
                ws_front.range("C35").value = 20

            ws_front.range("C39").value = f"Santos, {data_extenso}"

            # ---------- REMOVE OUTRAS ABAS ----------
            for sheet in list(wb2.sheets):
                if sheet.name != ws_front.name:
                    sheet.delete()

            print("\n✅ Faturamento De Acordo concluído!")

        finally:
            # ---------- SALVA E FECHA ----------
            try:
                # Definindo caminho do arquivo final igual ao FaturamentoCompleto
                pasta_saida = Path(wb1.fullname).parent if wb1 else Path.home() / "Desktop"
                arquivo_saida = pasta_saida / f"3 - DN_{dn}.xlsx"

                # Chamada unificada de salvar/fechar
                fechar_workbooks(
                    app=app,
                    wb_navio=wb1,
                    wb_cliente=wb2,
                    arquivo_saida=arquivo_saida
                )

            except Exception as e:
                print(f"⚠️ Erro ao salvar/fechar workbooks: {e}")
                raise

# ==============================
# CLASSE 3: Fazer Ponto
# ==============================

class ProgramaCopiarPeriodo:
    PERIODOS_MENU = {"1": "06h", "2": "12h", "3": "18h", "4": "00h"}
    MAPA_PERIODOS = {
        "06h": "06h", "6h": "06h", "06": "06h",
        "12h": "12h", "12": "12h",
        "18h": "18h", "18": "18h",
        "00h": "00h", "0h": "00h", "00": "00h", "24h": "00h"
    }
    EQUIVALENTES = {
        "06h": ["06h", "12h"],
        "12h": ["12h", "06h"],
        "18h": ["18h", "00h"],
        "00h": ["00h", "18h"]
    }
    BLOCOS = {"06h": 1, "12h": 1, "18h": 2, "00h": 2}

    def __init__(self, debug=False):
        self.wb1 = None
        self.ws1 = None
        self.datas = []
        self.debug = debug

    # ---------------------------
    # Abrir arquivo NAVIO
    # ---------------------------
    def abrir_arquivo_navio(self, caminho):
        self.wb1 = openpyxl.load_workbook(caminho)
        self.ws1 = self.wb1.active
        if self.debug:
            print(f"[DEBUG] Arquivo NAVIO '{caminho}' aberto.")

    # ---------------------------
    # Datas
    # ---------------------------
    def carregar_datas(self):
        datas = []
        for row in self.ws1.iter_rows(min_row=2, max_col=2):
            v = row[1].value
            if isinstance(v, (datetime, date)):
                datas.append(v.strftime("%d/%m/%Y"))
            elif isinstance(v, str) and "/" in v:
                datas.append(v.strip())
        self.datas = list(dict.fromkeys(datas))
        if not self.datas:
            raise Exception("Nenhuma data encontrada na coluna B.")

    def escolher_data(self):
        print("\nDatas disponíveis:")
        for i, d in enumerate(self.datas, 1):
            print(f"{i} - {d}")
        while True:
            try:
                escolha = int(input("Escolha a data: ")) - 1
                if 0 <= escolha < len(self.datas):
                    return self.datas[escolha]
                print("Opção inválida.")
            except ValueError:
                print("Opção inválida.")

    def escolher_periodo(self):
        print("\nHorário:")
        print("1 = 06h | 2 = 12h | 3 = 18h | 4 = 00h")
        while True:
            op = input("Opção: ").strip()
            if op in self.PERIODOS_MENU:
                return self.PERIODOS_MENU[op]
            print("Opção inválida.")

    # ---------------------------
    # Domingos / feriados
    # ---------------------------
    def is_dia_bloqueado(self, data_str):
        d = datetime.strptime(data_str, "%d/%m/%Y").date()
        return d.weekday() == 6 or d in feriados_br

    # ---------------------------
    # Localizar linhas
    # ---------------------------
    def encontrar_linha_data(self, data_str):
        for i, row in enumerate(self.ws1.iter_rows(min_row=2), start=2):
            valor = row[1].value
            if isinstance(valor, (datetime, date)) and valor.strftime("%d/%m/%Y") == data_str:
                return i
            
            elif valor == data_str:
                return i
        raise Exception(f"Data {data_str} não encontrada.")

    def encontrar_total_data(self, linha_data):
        i = linha_data + 1
        while True:
            valor_c = self.ws1.cell(row=i, column=3).value
            valor_a = self.ws1.cell(row=i, column=1).value
            if isinstance(valor_a, str) and valor_a.strip().lower() == "totalgeral":
                raise Exception("Total do dia não encontrado antes do Total Geral")
            if isinstance(valor_c, str) and valor_c.strip().lower() == "total":
                return i
            i += 1
            if i > self.ws1.max_row:
                raise Exception("Fim da planilha sem encontrar 'Total' do dia")

    def encontrar_linha_total_geral(self):
        for i, row in enumerate(self.ws1.iter_rows(min_row=1), start=1):
            valor = row[0].value
            if isinstance(valor, str) and "total" in valor.lower():
                return i
        raise Exception("Total Geral não encontrado.")

    # ---------------------------
    # Utilitario
    # ---------------------------

    def normalizar_texto(self, texto):
        return str(texto).lower().replace(" ", "")

    def normalizar_periodo(self, texto):
        t = self.normalizar_texto(texto)
        return self.MAPA_PERIODOS.get(t, None)

    def abrir_arquivo_navio(self, caminho: Path):
        try:
            self.caminho_navio = caminho  # <-- guardar caminho real
            self.wb1 = openpyxl.load_workbook(str(caminho))
            self.ws1 = self.wb1.active
            if self.debug:
                print(f"[DEBUG] Arquivo NAVIO '{caminho}' aberto com sucesso.")
        except Exception as e:
            print(f"❌ Erro ao abrir arquivo NAVIO: {e}")
            self.wb1 = None
            self.ws1 = None

    
    def selecionar_arquivo_navio(self) -> Path | None:
        root = Tk()
        root.lift()
        root.attributes("-topmost", True)
        root.focus_force()
        root.update()
        caminho = filedialog.askopenfilename(
            title="Selecione o arquivo NAVIO",
            filetypes=[("Arquivos Excel", "*.xlsx")]
        )
        root.destroy()
        if not caminho:
            return None
        print(f"📂 Arquivo NAVIO selecionado: {Path(caminho).name}")
        return Path(caminho)



    # ---------------------------
    # Encontrar linha modelo inteligente
    # ---------------------------
    
    def encontrar_modelo_periodo_inteligente(self, data_destino, periodo):
        datas_ordenadas = sorted(self.datas, key=lambda d: datetime.strptime(d, "%d/%m/%Y"))
        idx = datas_ordenadas.index(data_destino)
        bloco_alvo = self.BLOCOS[periodo]

        def procurar_na_data(d, mesmo_dia=False):
            linha_data = self.encontrar_linha_data(d)
            i = linha_data + 1
            while i <= self.ws1.max_row:
                valor_a = self.ws1.cell(row=i, column=1).value
                valor_c = self.ws1.cell(row=i, column=3).value

                # Pula total geral e total do dia
                if isinstance(valor_a, str) and self.normalizar_texto(valor_a) == "totalgeral":
                    break
                if isinstance(valor_c, str) and self.normalizar_texto(valor_c) == "total":
                    break

                # Pula células vazias ou não-string
                if not valor_c or not isinstance(valor_c, str):
                    i += 1
                    continue

                p = self.normalizar_periodo(valor_c)
                if not p:
                    i += 1
                    continue

                # Mesmo dia → aceita equivalentes
                if mesmo_dia and p in self.EQUIVALENTES[periodo]:
                    if self.debug:
                        print(f"✔ Mesmo dia: usando {p} de {d}")
                    return i

                # Outro dia → só se mesmo bloco
                if not mesmo_dia and self.BLOCOS[p] == bloco_alvo:
                    if self.debug:
                        print(f"✔ Outro dia: usando {p} de {d} (bloco {bloco_alvo})")
                    return i

                i += 1

            return None

        # 1️⃣ tenta na data escolhida
        linha = procurar_na_data(data_destino, mesmo_dia=True)
        if linha:
            return linha

        # 2️⃣ tenta em outras datas válidas (não bloqueadas)
        for offset in range(1, len(datas_ordenadas)):
            for novo_idx in (idx + offset, idx - offset):
                if 0 <= novo_idx < len(datas_ordenadas):
                    d = datas_ordenadas[novo_idx]
                    if self.is_dia_bloqueado(d):
                        continue
                    linha = procurar_na_data(d, mesmo_dia=False)
                    if linha:
                        return linha

        raise Exception(f"Nenhum modelo válido encontrado para {periodo}")

    # ---------------------------
    # Copiar e colar linha
    # ---------------------------
    
    def copiar_colar(self, data, periodo):
        if self.is_dia_bloqueado(data):
            print(f"⛔ {data} é domingo/feriado — período não será criado")
            return

        # Encontrar linhas importantes
        linha_data = self.encontrar_linha_data(data)
        linha_total_dia = self.encontrar_total_data(linha_data)
        linha_modelo = self.encontrar_modelo_periodo_inteligente(data, periodo)

        # 1️⃣ Inserir nova linha **acima do Total do dia**
        self.ws1.insert_rows(linha_total_dia)

        ultima_col = self.ws1.max_column

        # 2️⃣ Copiar valores da linha modelo (colunas 3 em diante) ignorando None ou 0
        for col in range(3, ultima_col + 1):
            origem = self.ws1.cell(row=linha_modelo, column=col)
            destino = self.ws1.cell(row=linha_total_dia, column=col)
            if origem.value not in (None, 0):
                destino.value = origem.value
                if origem.has_style:
                    destino.font = copy(origem.font)
                    destino.fill = copy(origem.fill)
                    destino.border = copy(origem.border)
                    destino.alignment = copy(origem.alignment)

        # Atualizar período na coluna C
        self.ws1.cell(row=linha_total_dia, column=3, value=periodo)

        # 3️⃣ Atualizar totais
        # Após inserir a linha, o Total do dia original está agora em linha_total_dia + 1
        self.somar_linha_no_total_do_dia(linha_total_dia, linha_total_dia + 1)
        self.somar_linha_no_total_geral(linha_total_dia)

        if self.debug:
            print(f"✔ Linha {linha_modelo} copiada para {linha_total_dia} com período {periodo}")

    # ---------------------------
    # Somar totais
    # ---------------------------

    def somar_linha_no_total_do_dia(self, linha_origem, linha_total_dia):
        ultima_col = self.ws1.max_column
        for col in range(3, ultima_col + 1):
            v_origem = self.ws1.cell(row=linha_origem, column=col).value
            v_total = self.ws1.cell(row=linha_total_dia, column=col).value
            try:
                v_origem_num = float(v_origem)
            except:
                v_origem_num = 0
            try:
                v_total_num = float(v_total)
            except:
                v_total_num = 0
            self.ws1.cell(row=linha_total_dia, column=col).value = v_total_num + v_origem_num
        if self.debug:
            print(f"➕ Linha {linha_origem} somada ao TOTAL DO DIA")

    def somar_linha_no_total_geral(self, linha_origem):
        linha_total_geral = self.encontrar_linha_total_geral()
        ultima_col = self.ws1.max_column
        for col in range(3, ultima_col + 1):
            valor_origem = self.ws1.cell(row=linha_origem, column=col).value
            try:
                valor_origem_num = float(valor_origem)
            except:
                continue
            v_total = self.ws1.cell(row=linha_total_geral, column=col).value
            try:
                v_total_num = float(v_total)
            except:
                v_total_num = 0
            self.ws1.cell(row=linha_total_geral, column=col).value = v_total_num + valor_origem_num
        if self.debug:
            print(f"➕ Linha {linha_origem} somada ao TOTAL GERAL")

    # ---------------------------
    # Executar
    # ---------------------------

    def executar(self, usar_arquivo_aberto=False):
        if not usar_arquivo_aberto:
            caminho = Path.home() / "Desktop" / "1.xlsx"
            if not caminho.exists():
                print(f"❌ Arquivo não encontrado: {caminho}")
                return
            self.caminho_navio = caminho
            self.wb1 = openpyxl.load_workbook(str(caminho))
            self.ws1 = self.wb1.active
        else:
            if not hasattr(self, "ws1") or self.ws1 is None:
                print("❌ Planilha não carregada. Use abrir_arquivo_navio() primeiro.")
                return
            if not hasattr(self, "caminho_navio"):
                print("❌ Caminho do arquivo não encontrado.")
                return

        # fluxo normal
        self.carregar_datas()
        data = self.escolher_data()
        periodo = self.escolher_periodo()
        print(f"\n✅ Executando FAZER PONTO no NAVIO - Data: {data}, Período: {periodo}")
        self.copiar_colar(data, periodo)

        # salvar no mesmo local do arquivo original
        arquivo_saida = self.caminho_navio.parent / "1_atualizado.xlsx"
        self.wb1.save(str(arquivo_saida))
        print(f"\n💾 Arquivo NAVIO atualizado salvo em: {arquivo_saida}")



class GerarRelatorio:
    pass


# ==============================
# MENU PRINCIPAL
# ==============================


class CentralSanport:
    def __init__(self):
        self.completo = FaturamentoCompleto()
        self.de_acordo = FaturamentoDeAcordo()
        self.programa_copiar = ProgramaCopiarPeriodo()
        self.relatorio = GerarRelatorio()   # ✅ nova função
        self.opcoes = [
            "FATURAMENTO",
            "DE ACORDO",
            "FAZER PONTO - X",
            "RELATÓRIO - X",   # ✅ adiciona no menu
            "SAIR DO PROGRAMA"
        ]
        

    # =========================
    # UTILITÁRIOS
    # =========================
    def limpar_tela(self):
        os.system("cls" if os.name == "nt" else "clear")

    def limpar_buffer_teclado(self):
        while msvcrt.kbhit():
            msvcrt.getch()

    def pausar_e_voltar(self, selecionado):
        print("\n🔁 Pressione ENTER para voltar ao menu...")
        while True:
            key = msvcrt.getch()
            if key in (b"\r", b"\n"):
                self.limpar_buffer_teclado()
                self.mostrar_menu(selecionado)
                return

    # =========================
    # MENU
    # =========================
    def mostrar_menu(self, selecionado):
        self.limpar_tela()

        print("╔" + "═" * 62 + "╗")
        print(f"║{' 🚢 CENTRAL DE PROCESSOS - SANPORT 🚢 '.center(60)}║")
        print("╚" + "═" * 62 + "╝\n")

        for i, opcao in enumerate(self.opcoes):
            if i == selecionado:
                print(f"          ►► {opcao} ◄◄")
            else:
                print(f"              {opcao}")

        print("\n" + "═" * 64)
        print("   ↑ ↓ = Navegar     ENTER = Selecionar")
        print("═" * 64)

    # =========================
    # EXECUÇÃO PRINCIPAL
    # =========================

    # =========================
    # Dentro da classe CentralSanport
    # =========================

    def rodar(self):
        selecionado = 0
        self.mostrar_menu(selecionado)

        while True:
            key = msvcrt.getch()

            # SETAS
            if key in (b"\xe0", b"\x00"):
                key = msvcrt.getch()

                if key == b"H":  # ↑
                    selecionado = max(0, selecionado - 1)
                    self.mostrar_menu(selecionado)

                elif key == b"P":  # ↓
                    selecionado = min(len(self.opcoes) - 1, selecionado + 1)
                    self.mostrar_menu(selecionado)

                continue

            # ENTER
            if key in (b"\r", b"\n"):
                self.limpar_tela()

                # ----------------------------
                # FATURAMENTO
                # ----------------------------
                if selecionado == 0:
                    print("╔" + "═" * 62 + "╗")
                    print("║" + " INICIANDO FATURAMENTO... ".center(60) + "║")
                    print("╚" + "═" * 62 + "╝\n")

                    try:
                        # Chama o FaturamentoCompleto
                        # NÃO precisa do input() dentro da classe
                        self.completo.executar()

                    except Exception as e:
                        print(f"\n❌ ERRO NO FATURAMENTO: {e}")

                    # Apenas aqui espera o ENTER para voltar
                    self.pausar_e_voltar(selecionado)

                # ----------------------------
                # DE ACORDO
                # ----------------------------
                elif selecionado == 1:
                    print("╔" + "═" * 62 + "╗")
                    print("║" + " INICIANDO DE ACORDO... ".center(60) + "║")
                    print("╚" + "═" * 62 + "╝\n")

                    try:
                        self.de_acordo.executar()
                    except Exception as e:
                        print(f"\n❌ ERRO: {e}")

                    self.pausar_e_voltar(selecionado)



                # ----------------------------
                # FAZER PONTO
                # ----------------------------
                elif selecionado == 2:  # FAZER PONTO
                    programa = ProgramaCopiarPeriodo(debug=True)

                    # Selecionar arquivo NAVIO
                    caminho_navio = programa.selecionar_arquivo_navio()
                    if not caminho_navio:
                        print("⛔ Nenhum arquivo selecionado")
                        self.pausar_e_voltar(selecionado)
                        continue

                    # Abrir arquivo NAVIO
                    programa.abrir_arquivo_navio(caminho_navio)

                    # Executar FAZER PONTO usando o arquivo já aberto
                    programa.executar(usar_arquivo_aberto=True)

                    self.pausar_e_voltar(selecionado)









                    # ----------------------------
                            # RELATÓRIO
                    # ----------------------------                
            
                
                
                
                
                elif selecionado == 3:   # posição da nova opção
                    print("╔" + "═" * 62 + "╗")
                    print("║" + " INICIANDO RELATÓRIO... ".center(60) + "║")
                    print("╚" + "═" * 62 + "╝\n")

                    try:
                        self.relatorio.executar()
                        print("\n✅ RELATÓRIO GERADO COM SUCESSO")
                    except Exception as e:
                        print(f"\n❌ ERRO NO RELATÓRIO: {e}")

                    self.pausar_e_voltar(selecionado)



                # ----------------------------
                # SAIR
                # ----------------------------
                elif selecionado == 4:
                    self.limpar_tela()
                    print("\n👋 Saindo do programa...")
                    break


if __name__ == "__main__":
    validar_licenca()              # 🔥 roda a verificação de licença primeiro
    CentralSanport().rodar()       # só abre o menu se a licença estiver válida
