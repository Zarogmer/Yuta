import xlwings as xw
from datetime import datetime, timedelta
from openpyxl.styles import Font, Alignment
import pandas as pd
import locale
import os
import sys
from datetime import date
import urllib.request
import sys
import ssl
import certifi
from datetime import datetime, timezone, date

def data_online():
    context = ssl.create_default_context(cafile=certifi.where())

    req = urllib.request.Request(
        "https://www.cloudflare.com",
        headers={"User-Agent": "Mozilla/5.0"}
    )

    with urllib.request.urlopen(req, context=context, timeout=5) as r:
        data_str = r.headers["Date"]

    # transforma em datetime UTC do servidor
    dt_utc = datetime.strptime(data_str, "%a, %d %b %Y %H:%M:%S %Z").replace(tzinfo=timezone.utc)

    # para uso interno/exibição: converte para hora local
    dt_local = dt_utc.astimezone()  # pega fuso local do sistema
    return dt_utc, dt_local

try:
    hoje_utc, hoje_local = data_online()

    # validação sempre UTC
    data_limite_utc = datetime(hoje_utc.year, hoje_utc.month, 20, tzinfo=timezone.utc)
    if hoje_utc > data_limite_utc:
        sys.exit("⛔ Licença expirada")

    # apenas exibição: sempre local
    print("Data local:", hoje_local.date())


except Exception as e:
    sys.exit(f"Erro ao verificar licença: {e}")

    sys.exit("❌ Não foi possível validar a data online")



# Abre Excel em background
app = xw.App(visible=False, add_book=False)

def main():

    print("==========Sanport==========")
    # Seu código principal aqui



def data_por_extenso(valor):
    if isinstance(valor, datetime):
        data = valor
    else:
        try:
            data = datetime.strptime(str(valor), "%d/%m/%Y")
        except:
            data = datetime.now()
    return data.strftime("%B %d, %Y")


def aplicar_fonte_xw(ws, celula, tamanho=15):
    """
    Aplica fonte Arial tamanho definido (padrão = 15)
    e centraliza a célula horizontal e verticalmente usando xlwings.
    """
    rng = ws.range(celula)
    rng.api.Font.Name = "Calibri"
    rng.api.Font.Size = tamanho
    rng.api.Font.Bold = False
    rng.api.Font.Italic = False
    rng.api.Font.Underline = False
    rng.api.HorizontalAlignment = xw.constants.HAlign.xlHAlignCenter
    rng.api.VerticalAlignment = xw.constants.VAlign.xlVAlignCenter





def formatar_coluna_c_data(ws, linha_inicial, periodos):
    """
    Aplica formatação de data (DD/MM/YYYY) e alinhamento centralizado
    na coluna C a partir da linha_inicial pelo número de 'periodos'.
    """
    alinhamento = Alignment(horizontal="center", vertical="center")
    formato_data = "dd/mm/yyyy"



    for offset in range(periodos):
        linha_atual = linha_inicial + offset
        celula = ws[f"C{linha_atual}"]

        celula.number_format = formato_data
        celula.alignment = alinhamento


def obter_data_mais_distante(ws_origem, ws_front_vigia):
    """
    Lê todas as datas válidas da coluna B,
    converte 'dd/mm/yyyy' para datetime,
    pega a data MAIS DISTANTE (maior)
    e cola no D17 em formato 'October 24, 2025'.
    """

    last_row = ws_origem.used_range.last_cell.row

    coluna_b = ws_origem.range(f"B1:B{last_row}").value

    datas_validas = []

    for v in coluna_b:
        if v in (None, "", "Total"):
            continue

        # Se já for datetime nativo (excel retorna assim às vezes)
        if isinstance(v, datetime):
            datas_validas.append(v)
            continue

        # Se for string tipo '05/09/2025'
        if isinstance(v, str):
            try:
                data = datetime.strptime(v.strip(), "%d/%m/%Y")
                datas_validas.append(data)
            except:
                pass  # ignora textos como "Data Operacional"

    if not datas_validas:
        return

    # Agora pega a maior data
    data_escolhida = max(datas_validas)

    data_extenso = data_por_extenso(data_escolhida)
    ws_front_vigia["D17"].value = data_extenso


    # 🔍 DEBUG CRÍTICO
    if not berco:
        return

    ws_front.range("D18").value = berco




def formatar_coluna_b_mes_extenso(ws, linha_inicial, periodos):
    """
    Formata SOMENTE as células existentes da coluna B no formato:
    'May 25, 2025'
    Sem criar linhas novas.
    """
    formato_custom = 'MMMM DD, YYYY'  # Mês extenso + dia + ano

    for i in range(periodos):
        linha = linha_inicial + i
        celula = ws[f"B{linha}"]

        # Formata apenas se a célula contém uma data válida
        if hasattr(celula.value, "year"):
            celula.number_format = formato_custom




#ABAS ESPECIFICAS   

def validar_c9(ws):
    valor_c9 = ws.range("C9").value
    if valor_c9 is None:
        return False

    clientes_validos = {
        "A/C AGÊNCIA MARÍTIMA CARGONAVE LTDA.",
    }

    return str(valor_c9).strip().upper() in clientes_validos

#
#def processar_cargill(wb, ws_origem):
  #  mmo = calcular_mmo(ws_origem)
   # wb.sheets["REPORT VIGIA"].range("G26").value = mmo
#



def arredondar_para_baixo_50(ws_front_vigia):

    if not validar_c9(ws_front_vigia):
        return

    valor = ws_front_vigia.range("E37").value

    if valor is None:
        return

    try:
        valor_int = int(valor)
    except:

        return

    resultado = (valor_int // 50) * 50
    ws_front_vigia.range("H28").value = resultado



def credit_note(wb, valor_c21):
    sheet_names = [s.name for s in wb.sheets]
    if "Credit Note" in sheet_names:
        ws3 = wb.sheets["Credit Note"]
        ws3["C21"].value = valor_c21

    else:
        pass  # Não faz nada se as abas não existirem

def quitacao(wb, valor_c21):
    sheet_names = [s.name for s in wb.sheets]
    if "Quitação" not in sheet_names:
        return

    ws4 = wb.sheets["Quitação"]
    ws4["C22"].value = valor_c21

    # Caminho da pasta "3" na Área de Trabalho
    pasta_pdfs = os.path.join(os.path.expanduser("~"), "Desktop", "JANEIRO")

    # Lista apenas arquivos .pdf
    pdfs = [f for f in os.listdir(pasta_pdfs) if f.lower().endswith(".pdf")]

    # Ordena (importante se forem 1.pdf, 2.pdf, 10.pdf, etc.)
    pdfs.sort(key=lambda x: int(os.path.splitext(x)[0]))

    # Número da NF = quantidade de PDFs processados até agora
    numero_nf = len(pdfs) + 1       

    ws4["H22"].value = f"NF.: {numero_nf}"




def atualizar_report_vigia(arquivo1, wb2):
    ws = wb2.sheets["REPORT VIGIA"]

    # Lê o valor de E25 (pode ser fórmula, pega o valor)
    valor_e25 = str(ws["E25"].value).strip().upper()

    # Só continua se tiver "MMO"
    if valor_e25 != "MMO":
        return

    # Lê o Excel do arquivo1
    df = pd.read_excel(arquivo1, sheet_name="Resumo", header=None)

    # Coluna G = índice 6
    col_g = df[6].dropna()

    if col_g.empty:
        return  # não há valores

    # Último valor preenchido
    ultimo = col_g.iloc[-1]

    # Converte para float corretamente, tratando "R$" e formato brasileiro
    try:
        ultimo_str = str(ultimo).replace("R$", "").strip()
        ultimo_float = locale.atof(ultimo_str)
    except:
        # se falhar, mantém como número bruto
        ultimo_float = float(ultimo)

    # Cola no Excel e formata
    ws["F25"].value = ultimo_float
    ws["F25"].number_format = "#.##0,00"




def OC(arquivo1, wb2):

    ws = wb2.sheets["FRONT VIGIA"]

    # Lê o valor atual da célula G16
    valor_g16 = str(ws["G16"].value).strip().upper()

    # Se G16 for "O.C.:", então escreve H16 normalmente
    if valor_g16 == "O.C.:":
        oc = input("OC: ")
        ws["H16"].value = f"{oc}"
    else:
        pass  # Não faz nada se as abas não existirem


try:
    data_online()
    main()
    from pathlib import Path

    desktop = Path.home() / "Desktop"

    arquivo1 = (desktop / "1.xlsx").resolve()
    arquivo2 = (desktop / "2.xlsx").resolve()

    # Abre workbooks
    wb1 = app.books.open(arquivo1)
    wb2 = app.books.open(arquivo2)

    ws1 = wb1.sheets[0]              # primeira aba do 1.xlsx (ativa no código original)
    ws_front = wb2.sheets["FRONT VIGIA"]   # primeira aba sensível (preservar shapes)
    # ws_report e demais serão acessadas pelo nome mais abaixo quando existirem

    # =========================
    # 1 – Transferências diretas (usando xlwings na FRONT VIGIA)
    # =========================
    ws_front.range("D15").value = ws1.range("A2").value
    ws_front.range("D16").value = data_por_extenso(ws1.range("B2").value)
    periodo = ws1.range("D2").value
    # =========================
    # 3 – Inputs do usuário
    # =========================
    numero_dn = input("DN: ").strip()
    berco = input("WAREHOUSE / BERÇO: ").strip().upper()
    ws_front["D18"].value = berco

    
    
    valor_c21 = f"DN: {numero_dn}/25"
    valor_c21_quit = f"DN: {numero_dn}/25"

    
    
    ws_front.range("C21").value = valor_c21

    # =========================
    # 4 – Se existir aba Quitação -> solicitar NF e gravar H21/H22
    #        (mas SÓ perguntar se a aba existir)
    # =========================
    # Chamada corrigida
    OC(arquivo1, wb2)
    obter_data_mais_distante(ws1, ws_front)
    atualizar_report_vigia(arquivo1, wb2)
    credit_note(wb2, valor_c21)
    quitacao(wb2, valor_c21)  


    # =========================
    # 6 – Data por extenso em C39 (FRONT VIGIA)
    # =========================
    meses = [
        "", "janeiro", "fevereiro", "março", "abril", "maio", "junho",
        "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"
    ]

    hoje = datetime.now()
    data_extenso = f"{hoje.day} de {meses[hoje.month]} de {hoje.year}"
    ws_front.range("C39").value = f" Santos, {data_extenso}"

    # =========================
    # 7 – REPORT VIGIA: clonar linha 22 para "periodos" vezes (preservando formatação)
    #    Usamos xlwings para manipular planilha REPORT VIGIA via Excel (mantém shapes)
    # =========================
    if "REPORT VIGIA" not in [s.name for s in wb2.sheets]:
        raise RuntimeError("Aba 'REPORT VIGIA' não encontrada em 2.xlsx")

    ws_report = wb2.sheets["REPORT VIGIA"]


    linha_origem = 22  # linha-base de onde você copia os estilos   
    # Captura altura da linha origem (RowHeight)
    row_height = ws_report.api.Rows(linha_origem).RowHeight
    # Localiza última célula preenchida da coluna AA no ws_report
    # Pega todas as células preenchidas na coluna AA manualmente
    # 1 – captura com end('down')
   

    df = pd.read_excel(arquivo1, sheet_name="Resumo", header=None)

    # Coluna AA = índice 26
    col_aa = df[26]

    # Remove vazios
    col_aa = col_aa.dropna()

    # Último valor preenchido
    ultimo = col_aa.iloc[-1]

    # Converte para inteiro
    periodos = int(float(str(ultimo).replace("R$", "").replace(",", ".").replace(" ", "")))


    # Inserir e copiar linha 22 para cada período (fazendo uma cópia real no Excel)
    for i in range(periodos):
        destino = linha_origem + 1 + i
        # Inserir linha em branco na posição destino
        ws_report.api.Rows(destino).Insert()
        # Copiar linha origem para destino (preserva fórmulas, formatação, merges)
        ws_report.api.Rows(linha_origem).Copy(ws_report.api.Rows(destino))
        # Ajustar altura
        ws_report.api.Rows(destino).RowHeight = row_height

    formatar_coluna_b_mes_extenso(ws_report, linha_inicial=22, periodos=periodos)

    
    # =========================
    # 8 – Formatar coluna C como data (DD/MM/YYYY) e preencher datas conforme ciclo
    # =========================

    df = pd.read_excel(arquivo1, sheet_name="Resumo")

    # ---- 1) coluna C ----
    coluna_c = df.iloc[:, 2].astype(str).str.strip()

    # ---- 2) cortar no primeiro "Total" ----
    lista_reduzida = []
    for valor in coluna_c:
        if valor.lower() == "total":
            break
        lista_reduzida.append(valor)

    # ---- 3) normalizar ----
    horarios = [h.lower() for h in lista_reduzida if h.lower() in ["00h", "06h", "12h", "18h"]]

    # ---- 4) ordem desejada (00h sempre por último) ----
    ordem = ["06h", "12h", "18h", "00h"]

    # ---- 5) pegar o menor horário ----
    horarios_validos = [h for h in ordem if h in horarios]

    if horarios_validos:
        menor_horario = horarios_validos[0]
    else:
        menor_horario = "06h"  # fallback

    # ---- 6) ciclo ----
    ciclo = ["06x12", "12x18", "18x00", "00x06"]
    mapa = {"06h": 0, "12h": 1, "18h": 2, "00h": 3}

    periodo_escolhido = ciclo[mapa[menor_horario]]

    indice = mapa[menor_horario]   # ← ISSO PRECISA EXISTIR




    # DESMESCLAR COLUNA E das linhas geradas (se necessário)
    # Percorrer merges e limpar apenas se intersecta as células alvo
    # xlwings não tem API direta para merged_ranges fácil; usamos COM:
    merged_items = ws_report.api.UsedRange.MergeCells  # True/False; para detalhes usamos MergeAreas
    # Simples approach: tentar unmerge nas células E{linha} para cada nova linha
    for i in range(periodos):
        linha = linha_origem + i
        ws_report.range(f"E{linha}").value = ciclo[(indice + i) % len(ciclo)]
        try:
            rng = ws_report.range(f"E{linha}")
            if rng.api.MergeCells:
                rng.api.UnMerge()
        except Exception:
            pass
            
    # Preencher coluna E com ciclo
    for i in range(periodos):
        linha = linha_origem + i
        ws_report.range(f"E{linha}").value = ciclo[(indice + i) % len(ciclo)]

    # Aplicar fonte Arial tamanho 20 na coluna E das linhas geradas
    for i in range(periodos):
        linha = linha_origem + i
        rng = ws_report.range(f"E{linha}")
        
        aplicar_fonte_xw(ws_report, f"E{linha}", tamanho=18)

    # =========================
    # 9 – Pegar valores da coluna Z do arquivo1 quando coluna C for "00h","06h","12h","18h"
    # =========================
    valores_validos = []
    filtros = ["00h", "06h", "12h", "18h"]

    # Ler colunas C e Z do wb1 até o fim (assumindo dados contínuos)
    dadosC = ws1.range("C2").options(expand='down').value  # lista
    dadosZ = ws1.range("Z2").options(expand='down').value

    # garantir listas
    if not isinstance(dadosC, list):
        dadosC = [dadosC] if dadosC is not None else []
    if not isinstance(dadosZ, list):
        dadosZ = [dadosZ] if dadosZ is not None else []

    # iterar paralelo
    for c_val, z_val in zip(dadosC, dadosZ):
        if c_val is None:
            break
        c_norm = str(c_val).strip().lower()
        if c_norm in filtros:
            valores_validos.append(z_val)

    # cortar para tamanho de periodos
    valores_validos = valores_validos[:periodos]

    # =========================
    # 10 – Colar na coluna G do REPORT VIGIA a partir de G22
    # =========================
    linha_destino = linha_origem
    for valor in valores_validos:
        cel = ws_report.range(f"G{linha_destino}")
        cel.value = valor
        cel.api.HorizontalAlignment = xw.constants.HAlign.xlHAlignRight
        cel.api.VerticalAlignment = xw.constants.VAlign.xlVAlignCenter
        # Formato moeda brasileiro com 2 casas decimais
        cel.api.NumberFormat = 'R$ #.##0,00'
        cel.api.Font.Name = "Calibri"
        cel.api.Font.Size = 18
        linha_destino += 1

    # =========================
    # Ajustar as datas na coluna C com base no ciclo
    # =========================
    # data_inicial vem de wb2 FRONT VIGIA D16
    di_val = ws_front.range("D16").value
    if isinstance(di_val, datetime):
        data_inicial = di_val
    else:
        # tentar converter string dd/mm/yyyy
        try:
            data_inicial = datetime.strptime(str(di_val), "%d/%m/%Y")
        except Exception:
            data_inicial = datetime.now()

    data_atual = data_inicial
    for i in range(periodos):
        linha = linha_origem + i
        periodo = ws_report.range(f"E{linha}").value

    # Preenche com STRING por extenso em inglês
        ws_report.range(f"C{linha}").value = data_por_extenso(data_atual)
        aplicar_fonte_xw(ws_report, f"C{linha}", tamanho=20)

        # Incremento da data baseado no período
        if periodo in ["00x06"]:
            data_atual += timedelta(days=1)

#========================
    # 11 – Arredondar para baixo o valor em E37 e colocar em H28 (FRONT VIGIA)
    # =========================
    arredondar_para_baixo_50(ws_front)
    validar_c9(ws_front)
    # =========================
    # SALVAR E FECHAR
    # =========================


    desktop = (Path.home() / "Desktop").resolve()
    novo_arquivo = desktop / "2_ATUALIZADO.xlsx"

    wb2.save(str(novo_arquivo))

    print("Arquivo salvo em:", novo_arquivo)


finally:
    try:
        wb1.close()
    except Exception:
        pass
    try:
        wb2.close()
    except Exception:
        pass
    app.quit()
