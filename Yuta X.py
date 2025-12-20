# 1️⃣ – Importações
import xlwings as xw
from datetime import datetime, timedelta, timezone
from openpyxl.styles import Alignment
import pandas as pd
import locale
import os
import sys
import urllib.request
import ssl
import certifi
from pathlib import Path
import tkinter as tk
from tkinter import filedialog
from datetime import date



# =========================
# Funções utilitárias
# =========================

def abrir_workbooks():
    """
    Abre:
    - 1.xlsx da pasta do NAVIO (selecionada pelo usuário)
    - XLSX do CLIENTE em Desktop/FATURAMENTOS/<CLIENTE>.xlsx
    """

    # --- Seleção da pasta do navio ---
    root = tk.Tk()
    root.withdraw()
    pasta_navio = filedialog.askdirectory(
        title="Selecione a pasta do NAVIO (onde está o 1.xlsx)"
    )

    if not pasta_navio:
        print("Nenhuma pasta selecionada. Encerrando.")
        return None, None, None, None, None

    pasta_navio = Path(pasta_navio)
    pasta_cliente = pasta_navio.parent   # PASTA PAI
    nome_cliente = pasta_cliente.name

    arquivo1 = pasta_navio / "1.xlsx"

    pasta_faturamentos = Path.home() / "Desktop" / "FATURAMENTOS"
    arquivo2 = pasta_faturamentos / f"{nome_cliente}.xlsx"

    app = xw.App(visible=False)
    wb1 = wb2 = None

    try:
        # --- Valida arquivos ---
        if not arquivo1.exists():
            raise FileNotFoundError(f"Arquivo 1.xlsx não encontrado:\n{arquivo1}")

        if not arquivo2.exists():
            raise FileNotFoundError(
                f"Arquivo de faturamento do cliente não encontrado:\n{arquivo2}"
            )

        # --- Abre workbooks ---
        wb1 = app.books.open(arquivo1)
        wb2 = app.books.open(arquivo2)

        ws1 = wb1.sheets[0]

        nomes_abas = [s.name for s in wb2.sheets]

        # Prioridade: aba com nome do cliente
        if nome_cliente in nomes_abas:
            ws_front = wb2.sheets[nome_cliente]

        # Fallback: FRONT VIGIA
        elif "FRONT VIGIA" in nomes_abas:
            ws_front = wb2.sheets["FRONT VIGIA"]

        else:
            raise RuntimeError(
                f"Nenhuma aba válida encontrada no faturamento.\n"
                f"Esperado: '{nome_cliente}' ou 'FRONT VIGIA'."
            )

        return app, wb1, wb2, ws1, ws_front

    except Exception as e:
        print(f"Erro ao abrir os arquivos: {e}")

        if wb1:
            wb1.close()
        if wb2:
            wb2.close()

        app.quit()
        return None, None, None, None, None

    
def fechar_workbooks(app, wb1=None, wb2=None, arquivo_saida=None):
    """
    Salva wb1 normalmente e salva wb2 EXCLUSIVAMENTE como 3.xlsx
    na pasta do cliente (ex: WILSON). Nunca salva em FATURAMENTOS.
    """
    try:
        if wb1:
            wb1.save()
            wb1.close()

        if wb2:
            if not arquivo_saida:
                raise RuntimeError(
                    "Caminho de saída não informado. "
                    "wb2 NÃO será salvo para evitar salvar em FATURAMENTOS."
                )

            wb2.save(arquivo_saida)
            wb2.close()

    finally:
        if app:
            app.quit()



def data_online():
    context = ssl.create_default_context(cafile=certifi.where())

    req = urllib.request.Request(
        "https://www.cloudflare.com",
        headers={"User-Agent": "Mozilla/5.0"}
    )

    with urllib.request.urlopen(req, context=context, timeout=5) as r:
        data_str = r.headers["Date"]

    # transforma em datetime UTC do servidor
    dt_utc = datetime.strptime(data_str, "%a, %d %b %Y %H:%M:%S %Z").replace(tzinfo=timezone.utc)

    # para uso interno/exibição: converte para hora local
    dt_local = dt_utc.astimezone()  # pega fuso local do sistema
    return dt_utc, dt_local

try:
    hoje_utc, hoje_local = data_online()

    # validação sempre UTC
    data_limite_utc = datetime(hoje_utc.year, hoje_utc.month, 22, tzinfo=timezone.utc)
    if hoje_utc > data_limite_utc:
        sys.exit("⛔ Licença expirada")

    # apenas exibição: sempre local
    print("Data local:", hoje_local.date())


except Exception as e:
    sys.exit(f"Erro ao verificar licença: {e}")

    sys.exit("❌ Não foi possível validar a data online")



def data_por_extenso(valor):
    if isinstance(valor, datetime):
        data = valor

    elif isinstance(valor, date):
        data = datetime(valor.year, valor.month, valor.day)

    elif isinstance(valor, str):
        try:
            data = datetime.strptime(valor, "%d/%m/%Y")
        except:
            return ""  # não inventa data

    else:
        return ""  # nunca usa datetime.now()

    return data.strftime("%d de %B de %Y")



def processar_front(ws1, ws_front):
    """
    Atualiza somente a aba FRONT VIGIA
    """

    # data atual por extenso (rodapé)
    meses = ["", "janeiro", "fevereiro", "março", "abril", "maio", "junho",
             "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"]

    hoje = datetime.now()
    ws_front.range("C39").value = (
        f"Santos, {hoje.day} de {meses[hoje.month]} de {hoje.year}"
    )

    # pega datas extremas do RESUMO
    data_min, data_max = obter_datas_extremos(ws1)

    # mostra no FRONT
    if data_min:
        ws_front.range("D16").value = data_por_extenso(data_min)

    if data_max:
        ws_front.range("D17").value = data_por_extenso(data_max)

    # 👉 retorna as datas para o main
    return data_min, data_max


def processar_report_completo(wb2, ws_origem, ws_front, arquivo1_path, linha_inicial=22):
    """
    Atualiza a aba REPORT VIGIA:
    1) Cria linhas conforme o número de períodos na coluna AA do Resumo
    2) Preenche coluna E com ciclos (06x12, 12x18, etc.)
    3) Preenche coluna G com valores correspondentes da coluna Z do Resumo
    """
    ws_report = wb2.sheets["REPORT VIGIA"]
    row_height = ws_report.api.Rows(linha_inicial).RowHeight

    # ===== 1) obter períodos pela coluna AA =====
    df_resumo = pd.read_excel(arquivo1_path, sheet_name="Resumo", header=None)
    col_aa = df_resumo[26].dropna()  # coluna AA é índice 26
    ultimo = col_aa.iloc[-1]
    try:
        periodos = int(float(str(ultimo).replace("R$", "").replace(",", ".").replace(" ", "")))
    except:
        periodos = 1

    # ===== 2) inserir linhas no REPORT VIGIA =====
    for i in range(periodos - 1):  # -1 porque a linha inicial já existe
        destino = linha_inicial + 1 + i
        ws_report.api.Rows(destino).Insert()
        ws_report.api.Rows(linha_inicial).Copy(ws_report.api.Rows(destino))
        ws_report.api.Rows(destino).RowHeight = row_height

    # ===== 3) obter horários e valores =====
    df2 = pd.read_excel(arquivo1_path, sheet_name="Resumo")
    col_c = df2.iloc[:, 2].astype(str).str.strip()
    col_z = df2.iloc[:, 25]  # coluna Z é índice 25

    # Filtra horários válidos e ignora "Total"
    horarios_validos = ["06h", "12h", "18h", "00h"]
    lista_horarios = []
    valores_por_ciclo = {"06x12": [], "12x18": [], "18x00": [], "00x06": []}
    mapa_horario_para_ciclo = {"06h": "06x12", "12h": "12x18", "18h": "18x00", "00h": "00x06"}

    for h, z in zip(col_c, col_z):
        h_lower = h.lower()
        if h_lower in horarios_validos:
            ciclo = mapa_horario_para_ciclo[h_lower]
            lista_horarios.append(h_lower)
            valores_por_ciclo[ciclo].append(z)

    # Define menor horário como ponto de partida
    ordem = ["06h", "12h", "18h", "00h"]
    menor_horario = next((h for h in ordem if h in lista_horarios), "06h")
    ciclo_base = ["06x12", "12x18", "18x00", "00x06"]
    indice_base = mapa_horario_para_ciclo[menor_horario]
    indice = ciclo_base.index(indice_base)

    # ===== 4) Preencher coluna E com ciclos =====
    ciclos_linha = []
    for i in range(periodos):
        ciclo_val = ciclo_base[(indice + i) % len(ciclo_base)]
        linha = linha_inicial + i
        ws_report.range(f"E{linha}").value = ciclo_val
        ciclos_linha.append(ciclo_val)
        cel = ws_report.range(f"E{linha}")
        try:
            if cel.api.MergeCells:
                cel.api.UnMerge()
        except:
            pass

    # ===== 5) Preencher coluna G com valores da coluna Z =====
    # Controle de índice por ciclo
    indices_ciclo = {c: 0 for c in valores_por_ciclo}
    for i, ciclo_val in enumerate(ciclos_linha):
        linha = linha_inicial + i
        idx = indices_ciclo[ciclo_val]
        lista_valores = valores_por_ciclo[ciclo_val]
        valor = lista_valores[idx] if idx < len(lista_valores) else None
        indices_ciclo[ciclo_val] += 1

        cel = ws_report.range(f"G{linha}")
        cel.value = valor
        cel.api.NumberFormat = 'R$ #.##0,00'
        cel.api.HorizontalAlignment = -4152
        cel.api.VerticalAlignment = -4108
        cel.api.Font.Name = "Calibri"
        cel.api.Font.Size = 18

    return periodos



MESES_PT = {
    "jan": 1, "fev": 2, "mar": 3, "abr": 4,
    "mai": 5, "jun": 6, "jul": 7, "ago": 8,
    "set": 9, "out": 10, "nov": 11, "dez": 12
}

from datetime import date, datetime

def obter_datas_extremos(ws_resumo):
    last_row = ws_resumo.used_range.last_cell.row
    valores = ws_resumo.range(f"B1:B{last_row}").value

    datas = []
    hoje = date.today()

    for v in valores:
        if v in (None, "", "Total"):
            continue

        # datetime vindo do Excel
        if isinstance(v, datetime):
            d = v.date()

            # 🚫 ignora fórmulas HOJE()
            if d == hoje:
                continue

            datas.append(d)
            continue

        # string
        if isinstance(v, str):
            v = v.strip().lower()

            # 19/10/2025
            try:
                datas.append(datetime.strptime(v, "%d/%m/%Y").date())
                continue
            except:
                pass

            # 19/out/25
            try:
                dia, mes_txt, ano = v.split("/")
                mes = MESES_PT.get(mes_txt[:3])
                if mes:
                    ano = int(ano)
                    if ano < 100:
                        ano += 2000
                    datas.append(date(ano, mes, int(dia)))
            except:
                pass

    if not datas:
        return None, None

    return min(datas), max(datas)




def obter_turno(ws, linha):
    """
    Procura o turno (06x12, 12x18, 18x00, 00x06)
    nas colunas C, D, E ou F da linha informada.
    """
    for col in ["C", "D", "E", "F"]:
        valor = ws.range(f"{col}{linha}").value
        if valor in ("06x12", "12x18", "18x00", "00x06"):
            return valor
    return None



def montar_datas_report_vigia(ws_report, data_inicio, linha_inicial=22):
    """
    Preenche a coluna C (DATE) do REPORT VIGIA.
    A data só avança quando encontra o período 00x06.
    """

    ultima_linha = ws_report.used_range.last_cell.row
    data_atual = data_inicio
    linha = linha_inicial

    while linha <= ultima_linha:
        periodo = ws_report.range(f"F{linha}").value  # coluna PERIODS

        if periodo in (None, ""):
            break

        # 👉 COLA A DATA NA COLUNA C
        ws_report.range(f"C{linha}").value = data_atual

        # vira o dia somente no 00x06
        if isinstance(periodo, str) and periodo.strip() == "00x06":
            data_atual += timedelta(days=1)

        linha += 1











def OC(arquivo1, wb2):
    ws = wb2.sheets["FRONT VIGIA"]
    if str(ws["G16"].value).strip().upper() == "O.C.:":
        ws["H16"].value = input("OC: ")

def credit_note(wb, valor_c21):
    if "Credit Note" in [s.name for s in wb.sheets]:
        wb.sheets["Credit Note"]["C21"].value = valor_c21

def quitacao(wb, valor_c21):
    if "Quitação" not in [s.name for s in wb.sheets]: return
    ws = wb.sheets["Quitação"]
    ws["C22"].value = valor_c21
    pasta_pdfs = os.path.join(os.path.expanduser("~"), "Desktop", "JANEIRO")
    pdfs = [f for f in os.listdir(pasta_pdfs) if f.lower().endswith(".pdf")]
    pdfs.sort(key=lambda x: int(os.path.splitext(x)[0]))
    ws["H22"].value = f"NF.: {len(pdfs)+1}"

def cargonave(ws):
    valor_c9 = ws.range("C9").value
    return str(valor_c9).strip().upper() == "A/C AGÊNCIA MARÍTIMA CARGONAVE LTDA."

def arredondar_para_baixo_50(ws_front_vigia):
    if not cargonave(ws_front_vigia): return
    valor = ws_front_vigia.range("E37").value
    if valor is None: return
    try: resultado = (int(valor)//50)*50
    except: return
    ws_front_vigia.range("H28").value = resultado

# =========================
# Programa principal
# =========================


def obter_sheet_por_nome_parcial(wb, nome_procurado):
    for s in wb.sheets:
        if nome_procurado.lower() in s.name.lower():
            return s
    raise Exception(f"Aba contendo '{nome_procurado}' não encontrada")


def main():
    print("🚀 Iniciando execução...")

    # ========= 1 – Validar licença =========
    try:
        hoje_utc, hoje_local = data_online()
        print(f"📅 Data local: {hoje_local.date()}")

        limite = datetime(hoje_utc.year, hoje_utc.month, 22, tzinfo=timezone.utc)
        if hoje_utc > limite:
            sys.exit("⛔ Licença expirada")

    except Exception as e:
        sys.exit(f"Erro ao verificar licença: {e}")

    # ========= 2 – Abrir arquivos =========
    # 2 – Abrir arquivos
    app, wb1, wb2, ws1, ws_front = abrir_workbooks()
    if not all([app, wb1, wb2, ws1, ws_front]):
        return

    # 3 – Inputs do usuário
    numero_dn = input("DN: ").strip()
    berco = input("WAREHOUSE / BERÇO: ").strip().upper()
    ws_front["D18"].value = berco

    # DN SOMENTE NO FRONT VIGIA
    valor_c21 = f"DN: {numero_dn}/25"
    ws_front.range("C21").value = valor_c21

    # 4 – Transferências iniciais
    ws_front.range("D15").value = ws1.range("A2").value
    ws_front.range("D16").value = data_por_extenso(ws1.range("B2").value)


    print("📂 Workbooks abertos com sucesso")

    # ========= 3 – FRONT (dados fixos) =========
    ws_front.range("D15").value = ws1.range("A2").value

    # ========= 4 – PROCESSAR FRONT =========
    print("⚙️ Processando FRONT VIGIA...")
    data_inicio, data_fim = processar_front(ws1, ws_front)

    if not data_inicio or not data_fim:
        sys.exit("❌ Datas extremas inválidas no RESUMO")

    print(f"📆 Datas extremas: {data_inicio} → {data_fim}")

    # ========= 5 – REPORT VIGIA =========
    ws_report = wb2.sheets["REPORT VIGIA"]

    print("⚙️ Processando REPORT VIGIA (linhas, ciclos e valores)...")
    total_linhas = processar_report_completo(
        wb2,
        ws1,
        ws_front,
        wb1.fullname,
        linha_inicial=22
    )

    # ========= 6 – PREENCHER DATAS (C22 ↓) =========
    print("🖊️ Preenchendo datas no REPORT VIGIA...")

    data_atual = data_inicio
    linha = 22

    while linha < 22 + total_linhas:
        turno = obter_turno(ws_report, linha)

        if not turno:
            linha += 1
            continue

        # escreve data na COLUNA C
        ws_report.range(f"C{linha}").value = data_atual.strftime("%d/%m/%Y")

        # se virar 00x06 → incrementa dia
        if turno == "00x06":
            data_atual += timedelta(days=1)

        linha += 1

    # ========= 7 – DEBUG FINAL =========
    print("🔍 Turnos detectados:")
    print(ws_report.range("C22:F40").value)

    print("🔍 Datas escritas (C22:C40):")
    print(ws_report.range("C22:C40").value)



    OC(str(wb1.fullname), wb2)



    # ❌ NÃO colar DN fora do FRONT VIGIA

    credit_note(wb2, valor_c21)
    quitacao(wb2, valor_c21)

    # 6 – Ajustes finais
    arredondar_para_baixo_50(ws_front)
    cargonave(ws_front)

    # Pasta do cliente (onde está o 1.xlsx)
    pasta_saida = Path(wb1.fullname).parent
    arquivo_saida = pasta_saida / "3.xlsx"

    # NÃO salvar wb2 aqui
    # wb2.save(arquivo_saida)  ❌ REMOVE ISSO

    fechar_workbooks(app, wb1, wb2, arquivo_saida)

    print(f"Processo finalizado. Arquivo final salvo em: {arquivo_saida}")

if __name__ == "__main__":
    
    main()
    
# Fim do código
