# 1️⃣ – Importações
import xlwings as xw
from datetime import datetime, timedelta, timezone
from openpyxl.styles import Alignment
import pandas as pd
import os
import sys
import urllib.request
import ssl
import certifi
from pathlib import Path
import tkinter as tk
from datetime import date
import re
import locale
from itertools import cycle
from tkinter import Tk, filedialog
import shutil
import tempfile

# =========================
# Funções utilitárias
# =========================




def abrir_workbooks():
    """
    Usuário seleciona uma pasta no Desktop.
    A pasta DEVE conter os arquivos:
    - 1.xlsx
    - 2.xlsx
    """
    root = Tk()
    root.withdraw()

    desktop = Path.home() / "Desktop"

    pasta = filedialog.askdirectory(
        title="Selecione a pasta no Desktop que contém 1.xlsx e 2.xlsx",
        initialdir=desktop
    )

    if not pasta:
        print("❌ Nenhuma pasta selecionada.")
        return None

    pasta = Path(pasta)

    arquivo1 = pasta / "1.xlsx"
    arquivo2 = pasta / "2.xlsx"

    if not arquivo1.exists() or not arquivo2.exists():
        print("❌ A pasta selecionada NÃO contém 1.xlsx e 2.xlsx")
        return None

    print("📂 Pasta selecionada:", pasta)
    print("• Abrindo:", arquivo1.name)
    print("• Abrindo:", arquivo2.name)

    app = xw.App(visible=False)
    app.display_alerts = False
    app.screen_updating = False

    try:
        wb1 = app.books.open(str(arquivo1))
        wb2 = app.books.open(str(arquivo2))

        ws1 = wb1.sheets[0]

        # FRONT VIGIA obrigatória
        if "FRONT VIGIA" not in [s.name for s in wb2.sheets]:
            raise RuntimeError("Aba 'FRONT VIGIA' não encontrada no 2.xlsx")

        ws_front = wb2.sheets["FRONT VIGIA"]

        return app, wb1, wb2, ws1, ws_front

    except Exception as e:
        print(f"❌ Erro ao abrir arquivos: {e}")
        try:
            wb1.close()
            wb2.close()
            app.quit()
        except:
            pass
        return None

    
def fechar_workbooks(app, wb1=None, wb2=None, arquivo_saida=None):
    """
    Salva wb1 normalmente e salva wb2 EXCLUSIVAMENTE como 3.xlsx
    na pasta do cliente (ex: WILSON). Nunca salva em FATURAMENTOS.
    """
    try:
        if wb1:
            wb1.save()
            wb1.close()

        if wb2:
            if not arquivo_saida:
                raise RuntimeError(
                    "Caminho de saída não informado. "
                    "wb2 NÃO será salvo para evitar salvar em FATURAMENTOS."
                )

            wb2.save(arquivo_saida)
            wb2.close()

    finally:
        if app:
            app.quit()




def data_por_extenso(valor):
    if isinstance(valor, datetime):
        data = valor

    elif isinstance(valor, date):
        data = datetime(valor.year, valor.month, valor.day)

    elif isinstance(valor, str):
        try:
            data = datetime.strptime(valor, "%d/%m/%Y")
        except:
            return ""  # não inventa data

    else:
        return ""  # nunca usa datetime.now()

    return data.strftime("%d de %B de %Y")

def processar_front(ws1, ws_front):
    """
    Atualiza somente a aba FRONT VIGIA
    """

    # data atual por extenso (rodapé)
    meses = ["", "janeiro", "fevereiro", "março", "abril", "maio", "junho",
             "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"]

    hoje = datetime.now()
    ws_front.range("C39").value = (
        f"Santos, {hoje.day} de {meses[hoje.month]} de {hoje.year}"
    )

    # pega datas extremas do RESUMO
    data_min, data_max = obter_datas_extremos(ws1)

    # mostra no FRONT
    if data_min:
        ws_front.range("D16").value = data_por_extenso(data_min)

    if data_max:
        ws_front.range("D17").value = data_por_extenso(data_max)

    # 👉 retorna as datas para o main
    return data_min, data_max


# ===== Funções REPORT =====#

def inserir_linhas_report(ws_report, linha_inicial, periodos):
    """
    Insere linhas copiando a linha inicial para acomodar periodos > 1
    """
    if periodos <= 1:
        return

    row_height = ws_report.api.Rows(linha_inicial).RowHeight

    for i in range(periodos - 1):
        destino = linha_inicial + 1 + i
        ws_report.api.Rows(destino).Insert()
        ws_report.api.Rows(linha_inicial).Copy(ws_report.api.Rows(destino))
        ws_report.api.Rows(destino).RowHeight = row_height


# ===== COLUNA E ===== #

def obter_periodos(ws_resumo):
    """
    Lê a coluna AA da aba Resumo usando xlwings
    (sem pandas, sem conflito de arquivo)
    """
    valores = ws_resumo.range("AA:AA").value

    # Remove None
    valores = [v for v in valores if v is not None]

    try:
        ultimo = str(valores[-1]).replace("R$", "").replace(",", ".").strip()
        return int(float(ultimo))
    except:
        return 1


def gerar_coluna_E_ajustada(ws1, periodos, coluna_horario="C"):
    """
    Gera a lista de ciclos para preencher a coluna E do REPORT VIGIA.
    
    - Se C3 for 06h, 12h, 18h ou 00h, começa a lista por este ciclo.
    - Se C3 for "Total" ou vazio, assume primeiro ciclo 00x06 e continua a sequência normal.
    - Repete a sequência até completar 'periodos'.
    """
    # Mapear horários para ciclos
    horario_para_ciclo = {
        "06h": "06x12", "06H": "06x12",
        "12h": "12x18", "12H": "12x18",
        "18h": "18x24", "18H": "18x24",
        "00h": "00x06", "00H": "00x06"
    }

    # Sequência padrão completa
    sequencia_padrao = ["06x12", "12x18", "18x24", "00x06"]

    # Ler primeira e segunda células da coluna
    primeiro_horario = str(ws1.range(f"{coluna_horario}2").value).strip()
    segundo_valor = ws1.range(f"{coluna_horario}3").value
    segundo_valor = str(segundo_valor).strip() if segundo_valor is not None else ""

    # Determinar primeiro ciclo
    if segundo_valor.lower() == "total" or segundo_valor not in horario_para_ciclo:
        primeiro_ciclo = "00x06"  # primeiro ciclo é sempre 00x06
    else:
        primeiro_ciclo = horario_para_ciclo[segundo_valor]

    # Rotacionar sequência padrão para iniciar pelo primeiro ciclo
    idx_inicio = sequencia_padrao.index(primeiro_ciclo)
    sequencia = sequencia_padrao[idx_inicio:] + sequencia_padrao[:idx_inicio]

    # Gerar lista final até completar periodos
    ciclos_linha = []
    for c in cycle(sequencia):
        if len(ciclos_linha) >= periodos:
            break
        ciclos_linha.append(c)

    return ciclos_linha


def preencher_coluna_E_por_ciclos(ws_report, ciclos_linha, linha_inicial=22):
    for i, ciclo in enumerate(ciclos_linha):
        ws_report.range(f"E{linha_inicial + i}").value = ciclo


 
# ===== COLUNA G ===== #


def mapear_valores_por_ciclo(ws1, coluna_horario="C", coluna_valor="Z"):
    """
    Lê os valores do wb1 e agrupa por ciclo.
    ws1 : planilha do arquivo 1
    coluna_horario : coluna que contém os horários (06h, 12h, etc.)
    coluna_valor : coluna que contém os valores a preencher
    """
    horario_para_ciclo = {"06h":"06x12", "12h":"12x18", "18h":"18x24", "00h":"00x06"}
    sequencia_ciclos = ["06x12", "12x18", "18x24", "00x06"]

    last_row = ws1.used_range.last_cell.row
    horarios = ws1.range(f"{coluna_horario}1:{coluna_horario}{last_row}").value
    valores = ws1.range(f"{coluna_valor}1:{coluna_valor}{last_row}").value

    # Normaliza horários para minúsculo
    horarios = [str(h).strip().lower() if h is not None else None for h in horarios]

    valores_por_ciclo = {c: [] for c in sequencia_ciclos}

    for h, v in zip(horarios, valores):
        if h in horario_para_ciclo:
            ciclo = horario_para_ciclo[h]
            valores_por_ciclo[ciclo].append(v)

    return valores_por_ciclo

def preencher_coluna_G_por_ciclo(ws_report, ciclos_linha, valores_por_ciclo, coluna="G", linha_inicial=22):
    """
    Preenche a coluna G do REPORT VIGIA alinhando os valores da coluna Z
    à sequência de ciclos já definida na coluna E.
    """
    indices_ciclo = {c: 0 for c in valores_por_ciclo}

    for i, ciclo_val in enumerate(ciclos_linha):
        linha = linha_inicial + i
        lista_valores = valores_por_ciclo.get(ciclo_val, [])
        idx = indices_ciclo[ciclo_val]

        valor = lista_valores[idx] if idx < len(lista_valores) else None
        indices_ciclo[ciclo_val] += 1

        cel = ws_report.range(f"{coluna}{linha}")
        cel.value = valor

        # Formatação
        try:
            cel.api.NumberFormat = 'R$ #.##0,00'
            cel.api.HorizontalAlignment = xw.constants.HAlign.xlHAlignRight
            cel.api.VerticalAlignment = xw.constants.VAlign.xlVAlignCenter
            cel.api.Font.Name = "Calibri"
            cel.api.Font.Size = 18
        except:
            pass

    return len(ciclos_linha)

# ===== COLUNA C ===== #


def montar_datas_report_vigia(ws_report, ws_resumo, linha_inicial=22, periodos=None):
    """
    Preenche a coluna C (DATE) do REPORT VIGIA.
    - O dia só avança quando o ciclo da coluna E for 00x06.
    - Mantém a sequência correta independentemente do primeiro horário.
    """
    if periodos is None:
        raise ValueError("É necessário informar 'periodos' para preencher as datas")

    data_inicio, data_fim = obter_datas_extremos(ws_resumo)
    if not data_inicio or not data_fim:
        raise ValueError("Não foi possível determinar as datas extremas na aba RESUMO")

    data_atual = data_inicio

    for i in range(periodos):
        linha = linha_inicial + i
        ciclo = ws_report.range(f"E{linha}").value

        if ciclo in (None, ""):
            break

        # Coloca a data atual na coluna C
        ws_report.range(f"C{linha}").value = data_atual

        # Se o ciclo for 00x06, incrementa o dia para a próxima linha
        if isinstance(ciclo, str) and ciclo.strip().lower() == "00x06":
            data_atual += timedelta(days=1)

    return periodos


# ===== DATA INICIAL E FINAL DO FRONT =====#

MESES_EN = {
    1: "JAN", 2: "FEB", 3: "MAR", 4: "APR",
    5: "MAY", 6: "JUN", 7: "JUL", 8: "AUG",
    9: "SEP", 10: "OCT", 11: "NOV", 12: "DEC"
}

def obter_datas_extremos(ws_resumo):
    last_row = ws_resumo.used_range.last_cell.row
    valores = ws_resumo.range(f"B1:B{last_row}").value

    datas = []
    hoje = date.today()

    for v in valores:
        if v in (None, "", "Total"):
            continue

        # datetime vindo do Excel
        if isinstance(v, datetime):
            d = v.date()

            # 🚫 ignora fórmulas HOJE()
            if d == hoje:
                continue

            datas.append(d)
            continue

        # string
        if isinstance(v, str):
            v = v.strip().lower()

            # 19/10/2025
            try:
                datas.append(datetime.strptime(v, "%d/%m/%Y").date())
                continue
            except:
                pass

            # 19/out/25
            try:
                dia, mes_txt, ano = v.split("/")
                mes = MESES_EN.get(int(mes_txt))
                if mes:
                    ano = int(ano)
                    if ano < 100:
                        ano += 2000
                    datas.append(date(ano, mes, int(dia)))
            except:
                pass

    if not datas:
        return None, None

    return min(datas), max(datas)


# ===== ABAS ESPECIFICAS =====#


def OC(arquivo1, wb2):
    ws = wb2.sheets["FRONT VIGIA"]
    if str(ws["G16"].value).strip().upper() == "O.C.:":
        ws["H16"].value = input("OC: ")

def credit_note(wb, valor_c21):
    if "Credit Note" in [s.name for s in wb.sheets]:
        wb.sheets["Credit Note"]["C21"].value = valor_c21

def quitacao(wb, valor_c21):
    if "Quitação" not in [s.name for s in wb.sheets]: return
    ws = wb.sheets["Quitação"]
    ws["C22"].value = valor_c21
    pasta_pdfs = os.path.join(os.path.expanduser("~"), "Desktop", "JANEIRO")
    pdfs = [f for f in os.listdir(pasta_pdfs) if f.lower().endswith(".pdf")]
    pdfs.sort(key=lambda x: int(os.path.splitext(x)[0]))
    ws["H22"].value = f"NF.: {len(pdfs)+1}"

def MMO(arquivo1, wb2):
    """
    Processa MMO sem abrir arquivo na rede (zero permission denied).
    wb2 é o wb_navio (tem "Resumo")
    Escreve em "REPORT VIGIA" do wb2
    """
    print("   Iniciando MMO...")

    try:
        ws_report = wb2.sheets["REPORT VIGIA"]
    except:
        print("   ⚠️ Aba 'REPORT VIGIA' não encontrada. Pulando MMO.")
        return

    if str(ws_report["E25"].value).strip().upper() != "MMO":
        print("   MMO não necessário (E25 != 'MMO').")
        return

    try:
        ws_resumo = wb2.sheets["Resumo"]
    except:
        print("   ⚠️ Aba 'Resumo' não encontrada. Pulando MMO.")
        return

    print("   Lendo coluna G...")
    valores_g = ws_resumo.range("G1:G1000").value
    valores_limpos = [v for v in valores_g if v is not None]

    if not valores_limpos:
        print("   Coluna G vazia. Pulando MMO.")
        return

    ultimo_valor = valores_limpos[-1]

    try:
        texto = str(ultimo_valor).replace("R$", "").replace(" ", "").strip()
        texto = texto.replace(".", "").replace(",", ".")
        ultimo_float = float(texto)
    except:
        print(f"   Erro ao converter '{ultimo_valor}'. Usando 0.")
        ultimo_float = 0.0

    ws_report["F25"].value = ultimo_float
    ws_report["F25"].number_format = "#,##0.00"

    print(f"   ✅ MMO concluído: R$ {ultimo_float:,.2f} escrito em F25")
    
def cargonave(ws):
    valor_c9 = ws.range("C9").value
    return str(valor_c9).strip().upper() == "A/C AGÊNCIA MARÍTIMA CARGONAVE LTDA."

def arredondar_para_baixo_50(ws_front_vigia):
    if not cargonave(ws_front_vigia): return
    valor = ws_front_vigia.range("E37").value
    if valor is None: return
    try: resultado = (int(valor)//50)*50
    except: return
    ws_front_vigia.range("H28").value = resultado

def obter_nome_navio_da_pasta(caminho_arquivo):
    """
    Ex: '123 - NAVIO' -> 'NAVIO'
    """
    pasta = os.path.basename(os.path.dirname(caminho_arquivo))

    if "-" in pasta:
        return pasta.split("-", 1)[1].strip()

    return pasta.strip()

def obter_aba_nf_opcional(wb):
    for sheet in wb.sheets:
        nome = sheet.name.strip().lower()
        if nome == "nf" or nome.startswith("nf") or "nota" in nome:
            return sheet
    return None

def escrever_nf(wb_faturamento, nome_navio, dn):
    # tenta localizar aba NF
    ws_nf = None
    for sheet in wb_faturamento.sheets:
        if sheet.name.strip().lower() == "nf":
            ws_nf = sheet
            break

    if ws_nf is None:
        print("⚠️ Aba NF não encontrada — seguindo sem escrever NF")
        return  # NÃO quebra o programa

    ano = datetime.now().year

    texto = (
        f"SERVIÇO PRESTADO DE ATENDIMENTO/APOIO AO M/V {nome_navio}\n"
        f"DN {dn}/{ano}"
    )

    # escreve na primeira célula
    cel = ws_nf.range("A1")
    cel.value = texto

    # mescla para ficar bonito
    ws_nf.range("A1:E2").merge()

    # formatação
    cel.api.HorizontalAlignment = -4108  # center
    cel.api.VerticalAlignment = -4108
    cel.api.WrapText = True
    cel.api.Font.Name = "Calibri"
    cel.api.Font.Size = 14
    cel.api.Font.Bold = True

    print("✅ Texto da NF escrito com sucesso")




def main():
    print("🚀 Iniciando execução...")

    # ========= 1 – Licença =========

    # ========= 2 – Localizar FATURAMENTOS =========
    
    # ========= 3 – Abrir arquivos =========
    resultado = abrir_workbooks()
    if not resultado:
        sys.exit("❌ Erro ou pasta inválida")

    app, wb1, wb2, ws1, ws_front = resultado


    print("📂 Workbooks abertos com sucesso!")

    try:
        # ========= 4 – DN e Navio =========
        dn = obter_dn_da_pasta(wb1.fullname)
        if not dn:
            sys.exit("❌ DN não identificada pela pasta")

        nome_navio = obter_nome_navio_da_pasta(wb1.fullname)
        ano_atual = datetime.now().year
        texto_dn = f"DN: {dn}/{ano_atual}"

        # Preenchimento FRONT VIGIA
        ws_front.range("D15").value = nome_navio
        ws_front.range("C21").value = texto_dn

        berco = input("WAREHOUSE / BERÇO: ").strip().upper()
        ws_front["D18"].value = berco

        # ========= 4 – Processar FRONT VIGIA =========
        print("⚙️ Processando FRONT VIGIA...")
        data_inicio, data_fim = processar_front(ws1, ws_front)

        if not data_inicio or not data_fim:
            sys.exit("❌ Datas extremas inválidas no RESUMO")

        print(f"📆 Datas extremas: {data_inicio} → {data_fim}")

        # ========= 5 – MMO =========
        print("⚙️ Processando MMO...")
        MMO(wb1.fullname, wb2)  # ou se você mudou para MMO(wb1, wb2), deixa assim
        # ========= 6 – NF =========
        escrever_nf(wb2, nome_navio, dn)

        # ========= 7 – REPORT VIGIA =========
        print("⚙️ Processando REPORT VIGIA...")
        ws_resumo = wb1.sheets["Resumo"]
        periodos = obter_periodos(ws_resumo)

        ws_report = wb2.sheets["REPORT VIGIA"]

        # Inserir linhas extras se necessário
        inserir_linhas_report(ws_report, linha_inicial=22, periodos=periodos)

        # Coluna E - Ciclos
        ciclos_linha = gerar_coluna_E_ajustada(ws1, periodos, coluna_horario="C")
        preencher_coluna_E_por_ciclos(ws_report, ciclos_linha, linha_inicial=22)

        # Coluna G - Valores
        valores_por_ciclo = mapear_valores_por_ciclo(ws1, coluna_horario="C", coluna_valor="Z")
        preencher_coluna_G_por_ciclo(ws_report, ciclos_linha, valores_por_ciclo, coluna="G", linha_inicial=22)

        # Coluna C - Datas (respeitando ciclos 00x06)
        montar_datas_report_vigia(
            ws_report=ws_report,
            ws_resumo=ws_resumo,
            linha_inicial=22,
            periodos=periodos
        )

        # ========= 8 – Financeiro =========
        print("⚙️ Processando Financeiro...")
        OC(str(wb1.fullname), wb2)
        credit_note(wb2, texto_dn)
        quitacao(wb2, texto_dn)  # descomentado se precisar

        # ========= 9 – Ajustes finais =========
        print("⚙️ Aplicando ajustes finais...")
        arredondar_para_baixo_50(ws_front)
        cargonave(ws_front)

        # ========= 10 – Salvar e fechar =========
        pasta_saida = Path(wb1.fullname).parent
        arquivo_saida = pasta_saida / "3.xlsx"

        fechar_workbooks(app, wb1, wb2, arquivo_saida)

        print(f"✅ Processo concluído com sucesso!")
        print(f"   Arquivo salvo em: {arquivo_saida}")

    except Exception as e:
        print(f"❌ Erro durante o processamento: {e}")
        # Garante que o Excel feche mesmo em caso de erro
        try:
            if wb1:
                wb1.close()
            if wb2:
                wb2.close()
            if app:
                app.quit()
        except:
            pass
        sys.exit(1)


if __name__ == "__main__":
    main()
