import sys
import time
from datetime import datetime
from pathlib import Path
import xlwings as xw
from datetime import datetime, timedelta, timezone
from openpyxl.styles import Alignment
import pandas as pd
import os
import sys
import urllib.request
import ssl
import certifi
from pathlib import Path
import tkinter as tk
from datetime import date
import re
from itertools import cycle
from tkinter import Tk, filedialog
import shutil
import tempfile
import time



# ==============================
# CLASSE 1: FATURAMENTO COMPLETO
# ==============================


class FaturamentoCompleto:
    def executar(self):
        print("🚀 Iniciando execução do Faturamento Completo...")

        # 1️⃣ – Importações


# =========================
# Funções utilitárias
# =========================


def copiar_para_temp_e_ler_excel(caminho_original: Path | str) -> pd.DataFrame:
    """
    Copia o arquivo para pasta temporária local e lê com pandas.
    Resolve a maioria dos PermissionError em pastas OneDrive/rede.
    """
    caminho_original = Path(caminho_original)
    if not caminho_original.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {caminho_original}")

    with tempfile.TemporaryDirectory() as temp_dir:
        temp_dir_path = Path(temp_dir)
        caminho_temp = temp_dir_path / caminho_original.name

        print(f"Copiando {caminho_original.name} para pasta temporária local...")
        shutil.copy2(caminho_original, caminho_temp)

        print(f"Lendo arquivo temporário: {caminho_temp}")
        df = pd.read_excel(caminho_temp, engine="openpyxl")  # engine explícito ajuda

    return df


def obter_pasta_faturamentos() -> Path:
    """
    Localiza automaticamente a pasta FATURAMENTOS dentro da estrutura OneDrive da SANPORT.
    Funciona mesmo com espaços e hífens no nome da pasta.
    """
    print("\n=== BUSCANDO PASTA FATURAMENTOS AUTOMATICAMENTE ===")

    # Possíveis locais base onde o OneDrive sincroniza a pasta da empresa
    possiveis_bases = [
        Path(r"C:\Users\Carol\SANPORT LOGÍSTICA PORTUÁRIA LTDA"),
        Path(
            r"C:\Users\Carol\OneDrive - SANPORT LOGÍSTICA PORTUÁRIA LTDA"
        ),  # caso seja OneDrive pessoal
        Path.home() / "SANPORT LOGÍSTICA PORTUÁRIA LTDA",
        Path.home() / "OneDrive" / "SANPORT LOGÍSTICA PORTUÁRIA LTDA",
    ]

    caminho_alvo = None

    for base in possiveis_bases:
        if base.exists():
            print(f"✅ Encontrada pasta base: {base}")

            # Procurar recursivamente por uma pasta chamada "FATURAMENTOS" dentro de "01. FATURAMENTOS"
            candidatos = list(base.rglob("FATURAMENTOS"))
            for candidato in candidatos:
                # Filtrar para garantir que está dentro de "01. FATURAMENTOS"
                if "01. FATURAMENTOS" in candidato.parent.as_posix():
                    caminho_alvo = candidato
                    print(f"✅ Pasta FATURAMENTOS encontrada em:\n   {caminho_alvo}")
                    break

            if caminho_alvo:
                break
        else:
            print(f"❌ Não encontrada: {base}")

    if not caminho_alvo:
        print("❌ Pasta FATURAMENTOS não foi encontrada automaticamente.")
        print("\nPossíveis soluções:")
        print("• Verifique se o OneDrive está sincronizando a pasta da empresa")
        print(
            "• Clique com botão direito na pasta FATURAMENTOS → Propriedades → Localização"
        )
        print("  e me diga o caminho exato que aparece")
        raise FileNotFoundError("Pasta FATURAMENTOS não localizada automaticamente")

    # Debug final: listar alguns arquivos para confirmar
    print(
        f"\nArquivos .xlsx encontrados na pasta ({len(list(caminho_alvo.glob('*.xlsx')))}):"
    )
    for arq in sorted(caminho_alvo.glob("*.xlsx"))[:10]:  # mostra só os 10 primeiros
        print(f"   • {arq.name}")
    if len(list(caminho_alvo.glob("*.xlsx"))) > 10:
        print("   ... (mais arquivos)")

    print("========================================\n")
    return caminho_alvo


def abrir_workbooks(pasta_faturamentos: Path):
    """
    Abre:
    - 1.xlsx (ou similar) da pasta do NAVIO selecionada pelo usuário
    - XLSX do cliente na pasta FATURAMENTOS
    """
    root = tk.Tk()
    root.withdraw()

    # 1️⃣ Usuário seleciona a pasta do NAVIO
    pasta_navio_str = filedialog.askdirectory(
        title="Selecione a pasta do NAVIO (onde está o 1.xlsx)"
    )

    if not pasta_navio_str:
        print("Seleção cancelada pelo usuário.")
        return None, None, None, None, None

    pasta_navio = Path(pasta_navio_str)
    pasta_cliente = pasta_navio.parent
    nome_cliente = pasta_cliente.name.strip()  # remove espaços acidentais

    # Procura arquivo que começa com "1" (1.xlsx, 1.xls, etc.)
    arquivos_1 = list(pasta_navio.glob("1*.xls*"))
    if not arquivos_1:
        raise FileNotFoundError(
            f"Nenhum arquivo iniciando com '1' encontrado em:\n{pasta_navio}"
        )

    arquivo1 = arquivos_1[
        0
    ]  # pega o primeiro (se tiver mais de um, pode melhorar depois)

    # 2️⃣ Arquivo de faturamento do cliente (caminho fixo)
    arquivo2 = pasta_faturamentos / f"{nome_cliente}.xlsx"

    if not arquivo2.exists():
        raise FileNotFoundError(f"Arquivo de faturamento não encontrado:\n{arquivo2}")

    app.api.Calculate()
    time.sleep(0.5)

    app = xw.App(visible=False)
    wb1 = wb2 = None

    try:
        # Aqui você pode usar xlwings direto, mas se quiser evitar lock, pode copiar também:
        wb1 = app.books.open(str(arquivo1))
        wb2 = app.books.open(str(arquivo2))

        ws1 = wb1.sheets[0]

        nomes_abas = [s.name for s in wb2.sheets]

        if nome_cliente in nomes_abas:
            ws_front = wb2.sheets[nome_cliente]
        elif "FRONT VIGIA" in nomes_abas:
            ws_front = wb2.sheets["FRONT VIGIA"]
        else:
            raise RuntimeError(
                f"Nenhuma aba válida encontrada.\n"
                f"Esperado: '{nome_cliente}' ou 'FRONT VIGIA'"
            )

        return app, wb1, wb2, ws1, ws_front

    except Exception as e:
        if wb1:
            wb1.close()
        if wb2:
            wb2.close()
        if app:
            app.quit()
        raise e


def fechar_workbooks(app, wb1=None, wb2=None, arquivo_saida=None):
    """
    Salva wb1 normalmente e salva wb2 EXCLUSIVAMENTE como 3.xlsx
    na pasta do cliente (ex: WILSON). Nunca salva em FATURAMENTOS.
    """
    try:
        if wb1:
            wb1.save()
            wb1.close()

        if wb2:
            if not arquivo_saida:
                raise RuntimeError(
                    "Caminho de saída não informado. "
                    "wb2 NÃO será salvo para evitar salvar em FATURAMENTOS."
                )

            wb2.save(arquivo_saida)
            wb2.close()

    finally:
        if app:
            app.quit()


def obter_dn_da_pasta(caminho_arquivo):
    """
    Extrai números do nome da pasta do navio
    Ex: '123 - NAVIO' -> '123'
    """
    pasta = os.path.basename(os.path.dirname(caminho_arquivo))

    numeros = re.findall(r"\d+", pasta)

    if not numeros:
        return None

    return numeros[0]  # primeiro bloco numérico


# ===== Licença =====#


def data_online():
    context = ssl.create_default_context(cafile=certifi.where())

    req = urllib.request.Request(
        "https://www.cloudflare.com", headers={"User-Agent": "Mozilla/5.0"}
    )

    with urllib.request.urlopen(req, context=context, timeout=5) as r:
        data_str = r.headers["Date"]

    dt_utc = datetime.strptime(data_str, "%a, %d %b %Y %H:%M:%S %Z").replace(
        tzinfo=timezone.utc
    )

    dt_local = dt_utc.astimezone()
    return dt_utc, dt_local


def validar_licenca():
    hoje_utc, hoje_local = data_online()

    limite = datetime(hoje_utc.year, hoje_utc.month, 30, tzinfo=timezone.utc)
    if hoje_utc > limite:
        sys.exit("⛔ Licença expirada")

    print(f"📅 Data local: {hoje_local.date()}")


def data_por_extenso(valor):
    if isinstance(valor, datetime):
        data = valor

    elif isinstance(valor, date):
        data = datetime(valor.year, valor.month, valor.day)

    elif isinstance(valor, str):
        try:
            data = datetime.strptime(valor, "%d/%m/%Y")
        except:
            return ""  # não inventa data

    else:
        return ""  # nunca usa datetime.now()

    return data.strftime("%d de %B de %Y")


def processar_front(ws1, ws_front):
    """
    Atualiza somente a aba FRONT VIGIA
    """

    # data atual por extenso (rodapé)
    meses = [
        "",
        "janeiro",
        "fevereiro",
        "março",
        "abril",
        "maio",
        "junho",
        "julho",
        "agosto",
        "setembro",
        "outubro",
        "novembro",
        "dezembro",
    ]

    hoje = datetime.now()
    ws_front.range("C39").value = (
        f"Santos, {hoje.day} de {meses[hoje.month]} de {hoje.year}"
    )

    # pega datas extremas do RESUMO
    data_min, data_max = obter_datas_extremos(ws1)

    # mostra no FRONT
    if data_min:
        ws_front.range("D16").value = data_por_extenso(data_min)

    if data_max:
        ws_front.range("D17").value = data_por_extenso(data_max)

    # 👉 retorna as datas para o main
    return data_min, data_max


# ===== Funções REPORT =====#


def inserir_linhas_report(ws_report, linha_inicial, periodos):
    """
    Insere linhas copiando a linha inicial para acomodar periodos > 1
    """
    if periodos <= 1:
        return

    row_height = ws_report.api.Rows(linha_inicial).RowHeight

    for i in range(periodos - 1):
        destino = linha_inicial + 1 + i
        ws_report.api.Rows(destino).Insert()
        ws_report.api.Rows(linha_inicial).Copy(ws_report.api.Rows(destino))
        ws_report.api.Rows(destino).RowHeight = row_height


# ===== COLUNA E ===== #


def obter_periodos(ws_resumo):
    """
    Lê a coluna AA da aba Resumo usando xlwings
    (sem pandas, sem conflito de arquivo)
    """
    valores = ws_resumo.range("AA:AA").value

    # Remove None
    valores = [v for v in valores if v is not None]

    try:
        ultimo = str(valores[-1]).replace("R$", "").replace(",", ".").strip()
        return int(float(ultimo))
    except:
        return 1


def gerar_coluna_E_ajustada(ws1, periodos, coluna_horario="C"):
    """
    Gera a lista de ciclos para preencher a coluna E do REPORT VIGIA.

    - Se C3 for 06h, 12h, 18h ou 00h, começa a lista por este ciclo.
    - Se C3 for "Total" ou vazio, assume primeiro ciclo 00x06 e continua a sequência normal.
    - Repete a sequência até completar 'periodos'.
    """
    # Mapear horários para ciclos
    horario_para_ciclo = {
        "06h": "06x12",
        "06H": "06x12",
        "12h": "12x18",
        "12H": "12x18",
        "18h": "18x24",
        "18H": "18x24",
        "00h": "00x06",
        "00H": "00x06",
    }

    # Sequência padrão completa
    sequencia_padrao = ["06x12", "12x18", "18x24", "00x06"]

    # Ler primeira e segunda células da coluna
    primeiro_horario = str(ws1.range(f"{coluna_horario}2").value).strip()
    segundo_valor = ws1.range(f"{coluna_horario}3").value
    segundo_valor = str(segundo_valor).strip() if segundo_valor is not None else ""

    # Determinar primeiro ciclo
    if segundo_valor.lower() == "total" or segundo_valor not in horario_para_ciclo:
        primeiro_ciclo = "00x06"  # primeiro ciclo é sempre 00x06
    else:
        primeiro_ciclo = horario_para_ciclo[segundo_valor]

    # Rotacionar sequência padrão para iniciar pelo primeiro ciclo
    idx_inicio = sequencia_padrao.index(primeiro_ciclo)
    sequencia = sequencia_padrao[idx_inicio:] + sequencia_padrao[:idx_inicio]

    # Gerar lista final até completar periodos
    ciclos_linha = []
    for c in cycle(sequencia):
        if len(ciclos_linha) >= periodos:
            break
        ciclos_linha.append(c)

    return ciclos_linha


def preencher_coluna_E_por_ciclos(ws_report, ciclos_linha, linha_inicial=22):
    for i, ciclo in enumerate(ciclos_linha):
        ws_report.range(f"E{linha_inicial + i}").value = ciclo


# ===== COLUNA G ===== #


def mapear_valores_por_ciclo(ws1, coluna_horario="C", coluna_valor="Z"):
    """
    Lê os valores do wb1 e agrupa por ciclo.
    ws1 : planilha do arquivo 1
    coluna_horario : coluna que contém os horários (06h, 12h, etc.)
    coluna_valor : coluna que contém os valores a preencher
    """
    horario_para_ciclo = {
        "06h": "06x12",
        "12h": "12x18",
        "18h": "18x24",
        "00h": "00x06",
    }
    sequencia_ciclos = ["06x12", "12x18", "18x24", "00x06"]

    last_row = ws1.used_range.last_cell.row
    horarios = ws1.range(f"{coluna_horario}1:{coluna_horario}{last_row}").value
    valores = ws1.range(f"{coluna_valor}1:{coluna_valor}{last_row}").value

    # Normaliza horários para minúsculo
    horarios = [str(h).strip().lower() if h is not None else None for h in horarios]

    valores_por_ciclo = {c: [] for c in sequencia_ciclos}

    for h, v in zip(horarios, valores):
        if h in horario_para_ciclo:
            ciclo = horario_para_ciclo[h]
            valores_por_ciclo[ciclo].append(v)

    return valores_por_ciclo


def preencher_coluna_G_por_ciclo(
    ws_report, ciclos_linha, valores_por_ciclo, coluna="G", linha_inicial=22
):
    """
    Preenche a coluna G do REPORT VIGIA alinhando os valores da coluna Z
    à sequência de ciclos já definida na coluna E.
    """
    indices_ciclo = {c: 0 for c in valores_por_ciclo}

    for i, ciclo_val in enumerate(ciclos_linha):
        linha = linha_inicial + i
        lista_valores = valores_por_ciclo.get(ciclo_val, [])
        idx = indices_ciclo[ciclo_val]

        valor = lista_valores[idx] if idx < len(lista_valores) else None
        indices_ciclo[ciclo_val] += 1

        cel = ws_report.range(f"{coluna}{linha}")
        cel.value = valor

        # Formatação
        try:
            cel.api.NumberFormat = "R$ #.##0,00"
            cel.api.HorizontalAlignment = xw.constants.HAlign.xlHAlignRight
            cel.api.VerticalAlignment = xw.constants.VAlign.xlVAlignCenter
            cel.api.Font.Name = "Calibri"
            cel.api.Font.Size = 18
        except:
            pass

    return len(ciclos_linha)


# ===== COLUNA C ===== #


def montar_datas_report_vigia(ws_report, ws_resumo, linha_inicial=22, periodos=None):
    """
    Preenche a coluna C (DATE) do REPORT VIGIA.
    - O dia só avança quando o ciclo da coluna E for 00x06.
    - Mantém a sequência correta independentemente do primeiro horário.
    """
    if periodos is None:
        raise ValueError("É necessário informar 'periodos' para preencher as datas")

    data_inicio, data_fim = obter_datas_extremos(ws_resumo)
    if not data_inicio or not data_fim:
        raise ValueError("Não foi possível determinar as datas extremas na aba RESUMO")

    data_atual = data_inicio

    for i in range(periodos):
        linha = linha_inicial + i
        ciclo = ws_report.range(f"E{linha}").value

        if ciclo in (None, ""):
            break

        # Coloca a data atual na coluna C
        ws_report.range(f"C{linha}").value = data_atual

        # Se o ciclo for 00x06, incrementa o dia para a próxima linha
        if isinstance(ciclo, str) and ciclo.strip().lower() == "00x06":
            data_atual += timedelta(days=1)

    return periodos


# ===== DATA INICIAL E FINAL DO FRONT =====#

MESES_EN = {
    1: "JAN",
    2: "FEB",
    3: "MAR",
    4: "APR",
    5: "MAY",
    6: "JUN",
    7: "JUL",
    8: "AUG",
    9: "SEP",
    10: "OCT",
    11: "NOV",
    12: "DEC",
}


def obter_datas_extremos(ws_resumo):
    last_row = ws_resumo.used_range.last_cell.row
    valores = ws_resumo.range(f"B1:B{last_row}").value

    datas = []
    hoje = date.today()

    for v in valores:
        if v in (None, "", "Total"):
            continue

        # datetime vindo do Excel
        if isinstance(v, datetime):
            d = v.date()

            # 🚫 ignora fórmulas HOJE()
            if d == hoje:
                continue

            datas.append(d)
            continue

        # string
        if isinstance(v, str):
            v = v.strip().lower()

            # 19/10/2025
            try:
                datas.append(datetime.strptime(v, "%d/%m/%Y").date())
                continue
            except:
                pass

            # 19/out/25
            try:
                dia, mes_txt, ano = v.split("/")
                mes = MESES_EN.get(int(mes_txt))
                if mes:
                    ano = int(ano)
                    if ano < 100:
                        ano += 2000
                    datas.append(date(ano, mes, int(dia)))
            except:
                pass

    if not datas:
        return None, None

    return min(datas), max(datas)


# ===== ABAS ESPECIFICAS =====#


def OC(wb1, wb2):
    ws = wb2.sheets["FRONT VIGIA"]
    if str(ws["G16"].value).strip().upper() == "O.C.:":
        ws["H16"].value = input("OC: ")


def credit_note(wb, valor_c21):
    if "Credit Note" in [s.name for s in wb.sheets]:
        wb.sheets["Credit Note"]["C21"].value = valor_c21


def obter_proxima_nf(pasta_nfs):
    if not os.path.exists(pasta_nfs):
        print(f"⚠️ Pasta de NFs não encontrada:\n{pasta_nfs}")
        return 1

    numeros = []

    for arquivo in os.listdir(pasta_nfs):
        nome, ext = os.path.splitext(arquivo)

        # ignora arquivos que não sejam PDF (se quiser)
        if ext.lower() != ".pdf":
            continue

        # captura apenas o número inicial do nome
        match = re.match(r"(\d+)", nome)
        if match:
            numeros.append(int(match.group(1)))

    if not numeros:
        return 1

    return max(numeros) + 1


def colar_nf(ws, celula, numero_nf):
    ws[celula].value = f"NF.: {numero_nf}"


def MMO(wb1, wb2):
    """
    wb1 = NAVIO (tem 'Resumo')
    wb2 = CLIENTE (tem 'REPORT VIGIA')
    """

    print("   Iniciando MMO...")

    # --- REPORT VIGIA (destino) ---
    try:
        ws_report = wb2.sheets["REPORT VIGIA"]
    except:
        print("   ⚠️ Aba 'REPORT VIGIA' não encontrada. Pulando MMO.")
        return

    if str(ws_report["E25"].value).strip().upper() != "MMO":
        print("   MMO não necessário (E25 != 'MMO').")
        return

    # --- RESUMO (origem) ---
    try:
        ws_resumo = wb1.sheets["Resumo"]
    except:
        print("   ⚠️ Aba 'Resumo' não encontrada no NAVIO. Pulando MMO.")
        return

    print("   Lendo coluna G do Resumo...")

    valores_g = ws_resumo.range("G1:G1000").value
    valores_limpos = [v for v in valores_g if v not in (None, "")]

    if not valores_limpos:
        print("   Coluna G vazia. Pulando MMO.")
        return

    ultimo_valor = valores_limpos[-1]

    try:
        texto = str(ultimo_valor)
        texto = texto.replace("R$", "").replace(" ", "")
        texto = texto.replace(".", "").replace(",", ".")
        ultimo_float = float(texto)
    except Exception as e:
        print(f"   Erro ao converter '{ultimo_valor}': {e}")
        ultimo_float = 0.0

    ws_report["F25"].value = ultimo_float
    ws_report["F25"].number_format = "#,##0.00"

    print(f"   ✅ MMO concluído: R$ {ultimo_float:,.2f} escrito em F25")


def cargonave(ws):
    valor_c9 = ws.range("C9").value
    return str(valor_c9).strip().upper() == "A/C AGÊNCIA MARÍTIMA CARGONAVE LTDA."


def arredondar_para_baixo_50(ws_front_vigia):
    if not cargonave(ws_front_vigia):
        return
    valor = ws_front_vigia.range("E37").value
    if valor is None:
        return
    try:
        resultado = (int(valor) // 50) * 50
    except:
        return
    ws_front_vigia.range("H28").value = resultado


def obter_nome_navio_da_pasta(caminho_arquivo):
    """
    Ex: '123 - NAVIO' -> 'NAVIO'
    """
    pasta = os.path.basename(os.path.dirname(caminho_arquivo))

    if "-" in pasta:
        return pasta.split("-", 1)[1].strip()

    return pasta.strip()


def obter_aba_nf_opcional(wb):
    for sheet in wb.sheets:
        nome = sheet.name.strip().lower()
        if nome == "nf" or nome.startswith("nf") or "nota" in nome:
            return sheet
    return None


def escrever_nf(wb_faturamento, nome_navio, dn):
    # tenta localizar aba NF
    ws_nf = None
    for sheet in wb_faturamento.sheets:
        if sheet.name.strip().lower() == "nf":
            ws_nf = sheet
            break

    if ws_nf is None:
        print("⚠️ Aba NF não encontrada — seguindo sem escrever NF")
        return  # NÃO quebra o programa

    ano = datetime.now().year

    texto = (
        f"SERVIÇO PRESTADO DE ATENDIMENTO/APOIO AO M/V {nome_navio}\n" f"DN {dn}/{ano}"
    )

    # escreve na primeira célula
    cel = ws_nf.range("A1")
    cel.value = texto

    # mescla para ficar bonito
    ws_nf.range("A1:E2").merge()

    # formatação
    cel.api.HorizontalAlignment = -4108  # center
    cel.api.VerticalAlignment = -4108
    cel.api.WrapText = True
    cel.api.Font.Name = "Calibri"
    cel.api.Font.Size = 14
    cel.api.Font.Bold = True

    print("✅ Texto da NF escrito com sucesso")


def quitacao(wb, valor_c21):
    if "Quitação" not in [s.name for s in wb.sheets]:
        return

    ws = wb.sheets["Quitação"]

    # Colar valor financeiro
    ws["C22"].value = valor_c21

    # Caminho real das NFs
    pasta_nfs = r"C:\Users\Carol\SANPORT LOGÍSTICA PORTUÁRIA LTDA\Central de Documentos - Documentos\2.2 CONTABILIDADE 2025\12 - DEZEMBRO"

    proxima_nf = obter_proxima_nf(pasta_nfs)

    # Colar NF
    colar_nf(ws, "H22", proxima_nf)


def main():
    print("🚀 Iniciando execução...")

    # ========= 1 – Licença =========
    validar_licenca()

    # ========= 2 – Localizar FATURAMENTOS =========
    pasta_faturamentos = obter_pasta_faturamentos()

    # ========= 3 – Abrir arquivos =========
    resultado = abrir_workbooks(pasta_faturamentos)
    if not resultado:
        sys.exit("❌ Usuário cancelou ou erro ao abrir workbooks")

    app, wb1, wb2, ws1, ws_front = resultado

    try:
        # ========= 4 – DN e Navio =========
        dn = obter_dn_da_pasta(wb1.fullname)
        if not dn:
            sys.exit("❌ DN não identificada pela pasta")

        nome_navio = obter_nome_navio_da_pasta(wb1.fullname)
        ano_atual = datetime.now().year
        texto_dn = f"DN: {dn}/{ano_atual}"

        # Preenchimento FRONT VIGIA
        ws_front.range("D15").value = nome_navio
        ws_front.range("C21").value = texto_dn

        berco = input("WAREHOUSE / BERÇO: ").strip().upper()
        ws_front["D18"].value = berco

        # ========= 4 – Processar FRONT VIGIA =========
        print("⚙️ Processando FRONT VIGIA...")
        data_inicio, data_fim = processar_front(ws1, ws_front)

        if not data_inicio or not data_fim:
            sys.exit("❌ Datas extremas inválidas no RESUMO")

        ws_resumo = wb1.sheets["Resumo"]
        # ========= 5 – MMO =========
        print("⚙️ Processando MMO...")

        app.api.Calculate()
        time.sleep(0.5)

        MMO(wb1, wb2)  # ou se você mudou para MMO(wb1, wb2), deixa assim
        # ========= 6 – NF =========
        escrever_nf(wb2, nome_navio, dn)

        # ========= 7 – REPORT VIGIA =========
        print("⚙️ Processando REPORT VIGIA...")
        periodos = obter_periodos(ws_resumo)

        ws_report = wb2.sheets["REPORT VIGIA"]

        # Inserir linhas extras se necessário
        inserir_linhas_report(ws_report, linha_inicial=22, periodos=periodos)

        # Coluna E - Ciclos
        ciclos_linha = gerar_coluna_E_ajustada(ws1, periodos, coluna_horario="C")
        preencher_coluna_E_por_ciclos(ws_report, ciclos_linha, linha_inicial=22)

        # Coluna G - Valores
        valores_por_ciclo = mapear_valores_por_ciclo(
            ws1, coluna_horario="C", coluna_valor="Z"
        )
        preencher_coluna_G_por_ciclo(
            ws_report, ciclos_linha, valores_por_ciclo, coluna="G", linha_inicial=22
        )

        # Coluna C - Datas (respeitando ciclos 00x06)
        montar_datas_report_vigia(
            ws_report=ws_report,
            ws_resumo=ws_resumo,
            linha_inicial=22,
            periodos=periodos,
        )

        # ========= 8 – Financeiro =========
        print("⚙️ Processando Financeiro...")
        OC(str(wb1.fullname), wb2)
        credit_note(wb2, texto_dn)
        quitacao(wb2, texto_dn)  # descomentado se precisar

        # ========= 9 – Ajustes finais =========
        print("⚙️ Aplicando ajustes finais...")
        arredondar_para_baixo_50(ws_front)
        cargonave(ws_front)

        # ========= 10 – Salvar e fechar =========
        pasta_saida = Path(wb1.fullname).parent
        arquivo_saida = pasta_saida / "3.xlsx"

        fechar_workbooks(app, wb1, wb2, arquivo_saida)

        print(f"✅ Processo concluído com sucesso!")
        print(f"   Arquivo salvo em: {arquivo_saida}")

    except Exception as e:
        print(f"❌ Erro durante o processamento: {e}")
        # Garante que o Excel feche mesmo em caso de erro
        try:
            if wb1:
                wb1.close()
            if wb2:
                wb2.close()
            if app:
                app.quit()
        except:
            pass
        sys.exit(1)


# ==============================
# CLASSE 2: FATURAMENTO DE ACORDO (simples)
# ==============================

class FaturamentoDeAcordo:
    def executar(self):
        print("🚀 Iniciando Gerador de FRONT VIGIA - SANPORT (versão De Acordo)\n")

        # =========================
        # FUNÇÕES AUXILIARES
        # =========================

# =========================
# FUNÇÕES AUXILIARES (fora da classe)
# =========================

        def selecionar_pasta_com_arquivos():
            root = tk.Tk()
            root.withdraw()
            root.attributes('-topmost', True)
            root.update()
            pasta_str = filedialog.askdirectory(
                title="Selecione a pasta que contém 2.xlsx",
                initialdir=str(Path.home() / "Desktop")
            )
            root.destroy()
            if not pasta_str:
                print("❌ Nenhuma pasta selecionada. Encerrando.")
                sys.exit(0)
            pasta = Path(pasta_str)
            print(f"📂 Pasta selecionada: {pasta.name}")
            return pasta

        def obter_dn_da_pasta(pasta: Path):
            numeros = re.findall(r"\d+", pasta.name)
            if not numeros:
                raise ValueError("Não foi possível identificar o DN no nome da pasta.")
            return numeros[0]

        def obter_nome_navio_da_pasta(pasta: Path):
            nome_limpo = re.sub(r"^\d+[\s\-_]*", "", pasta.name, flags=re.IGNORECASE).strip()
            return nome_limpo if nome_limpo else "NAVIO NÃO IDENTIFICADO"

        # =========================
        # CLASSE INTERNA: FRONT VIGIA PROCESSOR
        # =========================

        class FrontVigiaProcessor:
            def __init__(self, ws_front, dn, nome_navio):
                self.ws = ws_front
                self.dn = dn
                self.nome_navio = nome_navio

            def executar(self):
                hoje = datetime.now()
                meses = ["", "janeiro", "fevereiro", "março", "abril", "maio", "junho",
                         "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"]
                data_extenso = f"{hoje.day} de {meses[hoje.month]} de {hoje.year}"
                texto_dn = f"DN: {self.dn}/{hoje.year}"

                # Preenchimentos
                self.ws.range("D15").value = self.nome_navio
                self.ws.range("C21").value = texto_dn
                self.ws.range("D16").value = data_extenso
                self.ws.range("D17").value = data_extenso
                self.ws.range("D18").value = "-"
                self.ws.range("C26").value = f'  DE ACORDO ( M/V "{self.nome_navio}" )'
                self.ws.range("C27").value = f'  VOY SA02325'

                # Limpeza G27
                self.ws.range("G27").value = None
                print("🧹 Célula G27 limpa (conteúdo apagado)")

                # Regra Unimar
                cliente_c9 = str(self.ws.range("C9").value or "").strip()
                if "Unimar Agenciamentos" in cliente_c9:
                    self.ws.range("G26").value = 400

                # Taxa administrativa +20
                try:
                    valor_atual = float(self.ws.range("C35").value or 0)
                    self.ws.range("C35").value = valor_atual + 20
                except:
                    self.ws.range("C35").value = 20

                # Rodapé
                self.ws.range("C39").value = f"Santos, {hoje.day} de {meses[hoje.month]} de {hoje.year}"

                print("✅ FRONT VIGIA preenchido com sucesso!")

        # =========================
        # EXECUÇÃO PRINCIPAL DO DE ACORDO
        # =========================

        try:
            pasta_selecionada = selecionar_pasta_com_arquivos()
            arquivo2 = pasta_selecionada / "2.xlsx"

            if not arquivo2.exists():
                raise FileNotFoundError("Arquivo '2.xlsx' não encontrado na pasta selecionada.")

            dn = obter_dn_da_pasta(pasta_selecionada)
            nome_navio = obter_nome_navio_da_pasta(pasta_selecionada)
            print(f"📋 DN: {dn}")
            print(f"🚢 Navio: {nome_navio}")

            app = xw.App(visible=False)
            wb2 = app.books.open(str(arquivo2))

            if "FRONT VIGIA" not in [s.name for s in wb2.sheets]:
                raise RuntimeError("Aba 'FRONT VIGIA' não encontrada no arquivo 2.xlsx")
            
            ws_front = wb2.sheets["FRONT VIGIA"]

            processor = FrontVigiaProcessor(ws_front, dn, nome_navio)
            processor.executar()

            # Salvamento na Área de Trabalho
            desktop = Path.home() / "Desktop"
            arquivo_excel = desktop / f"3 - DN_{dn}.xlsx"
            arquivo_pdf = desktop / f"3 - DN_{dn}.pdf"

            # Remove outras abas
            for sheet in list(wb2.sheets):
                if sheet.name != "FRONT VIGIA":
                    sheet.delete()

            wb2.save(str(arquivo_excel))
            print(f"📄 Excel salvo: {arquivo_excel.name}")

            ws_front.api.ExportAsFixedFormat(
                Type=0,
                Filename=str(arquivo_pdf),
                Quality=0,
                IncludeDocProperties=True,
                IgnorePrintAreas=False,
                OpenAfterPublish=False
            )
            print(f"📄 PDF salvo: {arquivo_pdf.name}")

            wb2.close()
            app.quit()

            print(f"\n✅ CONCLUÍDO COM SUCESSO!")
            print(f"Arquivos gerados na Área de Trabalho:")
            print(f"   📊 {arquivo_excel.name}")
            print(f"   📄 {arquivo_pdf.name}")

        except Exception as e:
            print(f"\n❌ ERRO: {e}")
            try:
                if 'app' in locals():
                    app.quit()
            except:
                pass
            sys.exit(1)


# ==============================
# CLASSE 3: CENTRAL SANPORT (MENU PRINCIPAL)
# ==============================

class CentralSanport:
    def __init__(self):
        self.de_acordo = FaturamentoDeAcordo()
        self.completo = FaturamentoCompleto()

    @staticmethod
    def mostrar_menu():
        print("=" * 60)
        print("     🚢 CENTRAL DE ATIVIDADE - SANPORT")
        print("     Bem-vindo(a)!")
        print("=" * 60)
        print("1 - Faturamento Completo ")
        print("2 - Faturamento De Acordo")
        print("0 - Sair")
        print("-" * 60)

    def rodar(self):
        while True:
            self.mostrar_menu()
            opcao = input("\nDigite a opção (1, 2 ou 0): ").strip()

            if opcao == "1":
                print("\n🚀 Iniciando Faturamento Completo...\n")
                try:
                    self.completo.executar()
                except Exception as e:
                    print(f"\n❌ Erro no Faturamento Completo: {e}")
                finally:
                    print("\n⏎ Voltando ao menu principal...\n")
                    input("Pressione ENTER para continuar...")  # Pausa pra você ler o resultado

            elif opcao == "2":
                print("\n🚀 Iniciando Faturamento De Acordo...\n")
                try:
                    self.de_acordo.executar()
                except Exception as e:
                    print(f"\n❌ Erro no Faturamento De Acordo: {e}")
                finally:
                    print("\n⏎ Voltando ao menu principal...\n")
                    input("Pressione ENTER para continuar...")  # Pausa pra você ver que terminou

            elif opcao == "0":
                print("\n👋 Saindo do programa. Até logo!\n")
                break

            else:
                print("❌ Opção inválida. Tente novamente.\n")


# ==============================
# EXECUÇÃO PRINCIPAL — SÓ ESSA LINHA NO FINAL DO ARQUIVO!!!
# ==============================

if __name__ == "__main__":
    CentralSanport().rodar()